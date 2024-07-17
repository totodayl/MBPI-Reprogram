from openpyxl import load_workbook
import json
import psycopg2
import calendar
from psycopg2 import sql
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import Qt, pyqtSignal
from PyQt5.QtWidgets import *
from datetime import timedelta, datetime, time, date
import holidays as hd
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import platform
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas


# For Clickable Icons
class ClickableLabel(QtWidgets.QLabel):
    clicked = pyqtSignal()

    def mousePressEvent(self, event):
        self.clicked.emit()


class Ui_LoginWindow(object):
    def setupUi(self, LoginWindow):
        LoginWindow.setObjectName("LoginWindow")
        LoginWindow.resize(800, 600)
        LoginWindow.setWindowIcon(QtGui.QIcon("logo-logo.png"))
        self.login_window = QtWidgets.QWidget(LoginWindow)
        self.login_window.setStyleSheet("""background-color : qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:1, 
        stop:0 rgba(176, 0, 0, 255), stop:0.738636 rgba(255, 113, 250, 255))""")
        self.login_window.setObjectName("MainWindow")
        self.username = QtWidgets.QLineEdit(self.login_window)
        self.username.setGeometry(QtCore.QRect(310, 140, 171, 31))
        self.username.setAutoFillBackground(False)
        self.username.setStyleSheet("background-color: rgb(255, 255, 255); border-radius: 5px")
        self.username.setObjectName("username")
        self.username.setText("postgres")
        self.password = QtWidgets.QLineEdit(self.login_window)
        self.password.setGeometry(QtCore.QRect(310, 230, 171, 31))
        self.password.setAutoFillBackground(False)
        self.password.setStyleSheet("background-color: rgb(255, 255, 255); border-radius: 5px")
        self.password.setObjectName("password")
        self.password.setEchoMode(QtWidgets.QLineEdit.Password)  # Make the input hidden
        self.password.setText("mbpi")
        self.login_btn = QtWidgets.QPushButton(self.login_window)
        self.login_btn.setGeometry(QtCore.QRect(360, 340, 75, 23))
        self.login_btn.setStyleSheet("\n"
                                     "color: rgb(0, 0, 0);\n"
                                     "background-color: rgb(255, 255, 255); \n"
                                     "border-radius: 10px;")
        self.login_btn.setObjectName("Login")
        self.login_btn.clicked.connect(self.login)
        LoginWindow.setCentralWidget(self.login_window)
        self.retranslateUi(LoginWindow)
        QtCore.QMetaObject.connectSlotsByName(LoginWindow)

    def retranslateUi(self, LoginWindow):
        _translate = QtCore.QCoreApplication.translate
        LoginWindow.setWindowTitle(_translate("LoginWindow", "MBPI"))
        self.login_btn.setText(_translate("LoginWindow", "Login"))

    def login(self):
        username = self.username.text()
        pass1 = self.password.text()

        self.computerName = platform.node()

        # Connect to the Database
        try:
            self.conn = psycopg2.connect(
                host="192.168.1.13",
                port=5432,
                dbname='postgres',
                user=f'{username}',
                password=f'{pass1}'
            )

            self.username.deleteLater()
            self.password.deleteLater()
            self.login_btn.deleteLater()
        except psycopg2.Error:
            QMessageBox.information(self.login_window, "INVALID CREDENTIALS", "Check your Username and Password")
            return
        self.cursor = self.conn.cursor()
        self.launch_main()

    # This is the main window after login screen
    def launch_main(self):
        LoginWindow.move(37, 100)
        LoginWindow.setFixedSize(1200, 750)
        self.login_window.setStyleSheet("background-color: rgb(60,60,60);")
        self.main_widget = QtWidgets.QWidget(self.login_window)
        self.main_widget.setStyleSheet("""
        background-color: rgb(240,240,240);     
        
        """)
        self.main_widget.setGeometry(210, 0, 991, 751)
        self.main_widget.show()

        self.production_btn = QtWidgets.QPushButton(self.login_window)
        self.production_btn.setGeometry(30, 200, 180, 40)
        self.production_btn.setCursor(Qt.PointingHandCursor)
        self.production_btn.setStyleSheet("""
        border-top-left-radius: 10px; 
        border-bottom-left-radius: 10px; 
        background-color: rgb(125,125,125);
        """)
        self.production_btn.clicked.connect(self.production)
        self.production_btn.show()

        self.production_lbl = QtWidgets.QLabel(self.production_btn)
        self.production_lbl.setText("Extruder")
        self.production_lbl.setGeometry(50, 5, 100, 30)
        self.production_lbl.setFont(QtGui.QFont("Arial", 12))
        self.production_lbl.setStyleSheet("color: blue;")
        self.production_lbl.setCursor(Qt.PointingHandCursor)
        self.production_lbl.show()

        self.production_icon = ClickableLabel(self.production_btn)
        self.production_icon.setGeometry(10, 3, 30, 30)
        self.production_icon.setPixmap(QtGui.QIcon('setting.png').pixmap(30, 30))  # Set icon
        self.production_icon.setScaledContents(True)  # Scale icon to fit the label
        self.production_icon.setCursor(Qt.PointingHandCursor)
        self.production_icon.show()

        self.qualityControl_btn = QtWidgets.QPushButton(self.login_window)
        self.qualityControl_btn.setGeometry(30, 260, 180, 40)
        self.qualityControl_btn.setCursor(Qt.PointingHandCursor)
        self.qualityControl_btn.setStyleSheet("""
                border-top-left-radius: 10px; 
                border-bottom-left-radius: 10px; 
                background-color: rgb(125,125,125);
                """)
        self.qualityControl_btn.clicked.connect(self.quality_control)
        self.qualityControl_btn.show()

        self.qc_icon = ClickableLabel(self.qualityControl_btn)
        self.qc_icon.setGeometry(10, 3, 30, 30)
        self.qc_icon.setPixmap(QtGui.QIcon('qc.png').pixmap(30, 30))  # Set icon
        self.qc_icon.setScaledContents(True)  # Scale icon to fit the label
        self.qc_icon.setCursor(Qt.PointingHandCursor)
        self.qc_icon.show()

        self.qualityControl_lbl = QtWidgets.QLabel(self.qualityControl_btn)
        self.qualityControl_lbl.setText("Quality Control")
        self.qualityControl_lbl.setGeometry(50, 5, 120, 30)
        self.qualityControl_lbl.setFont(QtGui.QFont("Arial", 12))
        self.qualityControl_lbl.setStyleSheet("color: blue;")
        self.qualityControl_lbl.setCursor(Qt.PointingHandCursor)
        self.qualityControl_lbl.show()

        self.logo = QLabel(self.login_window)
        self.logo.setGeometry(23, 10, 170, 121)
        pixmap = QtGui.QPixmap('MBPI-ADJUST.png')
        pixmap = pixmap.scaled(170, 121)
        self.logo.setPixmap(pixmap)
        self.logo.show()

        self.logo2 = QLabel(self.login_window)
        self.logo2.setGeometry(0, 145, 210, 15)
        pixmap = QtGui.QPixmap('LOGO-NEW.png')
        pixmap = pixmap.scaled(210, 15)
        self.logo2.setPixmap(pixmap)
        self.logo2.show()





        self.production()


    def production(self):

        # Delete If there are existing Widgets
        try:
            self.info_widget.deleteLater()
            self.temp_table.deleteLater()
            self.time_table.deleteLater()
            self.material_table.deleteLater()
            self.group_box.deleteLater()
            self.export_btn.deleteLater()
            print("Widgets Cleared")

        except Exception as e:
            print(e)

        def show_form():

            try:
                selected = self.extruder_table.selectedItems()
                selected = [i.text() for i in selected]

                # Query the whole columns
                self.cursor.execute(f"SELECT * FROM extruder WHERE process_id = {selected[0]}")
                selected = self.cursor.fetchall()

                # Clear all the widget first
                self.extruder_table.deleteLater()
                self.view_btn.deleteLater()
                self.add_btn.deleteLater()
                self.update_btn.deleteLater()
                self.print_btn.deleteLater()
                main_time_table.deleteLater()
                material_table.deleteLater()
                lotNumber_table.deleteLater()



            except:
                QMessageBox.information(self.production_widget, "ERROR", "No Selected Item")
                return

            selected = selected[0]

            # Unpack all the items for convenience
            machine, ordered_qty, product_output, customer = selected[1:5]
            formula_id, product_code, order_id, total_time, time_start = selected[5:10]
            time_end, output_percent, loss, loss_percent, purging = selected[10:15]
            resin, remarks, screw_config, feed_rate, rpm = selected[15:20]
            screen_size, operator, supervisor, materials, temperature = selected[20:25]
            purge_duration, outputs, output_per_hour = selected[25:28]
            production_ID = selected[28]
            total_input = selected[29]

            def exportToExcel():
                from openpyxl.styles import Font
                from openpyxl.styles import Alignment

                process_id = selected[0]

                self.cursor.execute(f"""
                           SELECT * FROM extruder 
                           WHERE process_id = '{process_id}';

                           """)
                items = self.cursor.fetchall()[0]
                # Unpack the Items
                machine_number = items[1]
                quantity_order = items[2]
                customer = items[4]
                code = items[6]
                total_input = items[-2]
                total_time = items[8]
                time_start = items[9]
                time_end = items[10]
                outputPerHour = items[27]
                total_output = items[3]
                outputPercent = items[11]
                loss = items[12]
                lossPercent = items[13]
                purging = items[14]
                resin = items[15]
                remarks = items[16]
                operator = items[21]
                supervisor = items[22]
                outputs = items[-7]
                materials = items[-10]
                lot_number = items[30]
                purge_duration = time(hour=items[-8])

                wb = load_workbook(
                    r"\\mbpi-server-01\IT\AMIEL\Extruder System\dist\Extruder Template.xlsx")
                worksheet = wb.active

                font = Font(size=8, bold=True, name='Arial')
                center_Alignment = Alignment(horizontal='center', vertical='center')

                worksheet["F5"] = "Extruder Machine No. " + machine_number[-1]
                worksheet["A8"] = machine_number[-1]
                worksheet["B8"] = quantity_order  # quantity order
                worksheet["C8"].font = font
                worksheet["C8"].alignment = center_Alignment
                worksheet["C8"] = customer  # customer

                worksheet["F8"] = code  # product code
                worksheet["G9"] = total_input  # total input
                worksheet["H9"] = total_time  # total time used
                worksheet["I9"] = outputPerHour  # output Per Hour
                worksheet["K9"] = total_output  # total Output
                worksheet["L9"] = outputPercent  # Total Output Percentage
                worksheet["M9"] = loss
                worksheet["N9"] = lossPercent

                total_sec = timedelta()
                for row in range(len(time_start)):
                    worksheet["A" + str(12 + row)] = time_start[row].strftime("%d-%b-%Y %H:%M")
                    worksheet["D" + str(12 + row)] = time_end[row].strftime("%d-%b-%Y %H:%M")
                    worksheet["F" + str(12 + row)] = time_end[row] - time_start[row]
                    worksheet["G" + str(12 + row)] = outputs[row]
                    total_sec = total_sec + (time_end[row] - time_start[row])

                try:
                    hour = str(int(total_sec.total_seconds() // 3600))
                    minute = str((int(total_sec.total_seconds() % 3600) // 60))

                    total_time_used = time(int(hour), int(minute))

                    worksheet["F25"] = total_time_used
                except ValueError:
                    worksheet["F25"] = hour + ":" + minute

                for key in list(materials.keys()):
                    worksheet["I" + str(12 + list(materials.keys()).index(key))] = key
                    worksheet["K" + str(12 + list(materials.keys()).index(key))] = materials[key]

                for ln in range(len(lot_number)):
                    worksheet["M" + str(12 + ln)] = lot_number[ln]

                worksheet["B27"] = purging
                worksheet["E28"] = purge_duration
                worksheet["B29"] = resin
                worksheet["G26"] = remarks
                worksheet["M28"] = operator
                worksheet["M29"] = supervisor

                wb.save(r"\\mbpi-server-01\IT\AMIEL\Extruder System\dist\text.xlsx")
                print("load successful")

            # Convert string of json to JSON
            materials = str(materials).replace("'", '"')
            materials = json.loads(materials)

            # Main Widget
            self.info_widget = QtWidgets.QWidget(self.production_widget)
            self.info_widget.setGeometry(20, 0, 951, 450)
            self.info_widget.setStyleSheet("background-color: rgb(0,109,189);")
            self.info_widget.show()

            # Set Font Style and Size
            label_stylesheet = "color: rgb(195, 164, 86)"
            font = QtGui.QFont("Arial", 14)  # Set Font for Labels

            # Create 3 widgets for division
            info_widget1 = QtWidgets.QWidget(self.info_widget)
            info_widget1.setGeometry(0, 80, 158, 380)

            info_widget2 = QtWidgets.QWidget(self.info_widget)
            info_widget2.setGeometry(158, 80, 158, 380)

            info_widget3 = QtWidgets.QWidget(self.info_widget)
            info_widget3.setGeometry(316, 80, 158, 380)

            info_widget4 = QtWidgets.QWidget(self.info_widget)
            info_widget4.setGeometry(474, 80, 158, 380)

            info_widget5 = QtWidgets.QWidget(self.info_widget)
            info_widget5.setGeometry(632, 80, 158, 380)

            info_widget6 = QtWidgets.QWidget(self.info_widget)
            info_widget6.setGeometry(790, 80, 158, 380)

            # create Vertical Layouts
            left_vertical_layout = QVBoxLayout()

            self.ordered_company = QtWidgets.QLabel(self.info_widget)
            self.ordered_company.setFont(QtGui.QFont("Arial", 30))
            self.ordered_company.setGeometry(0, 30, 950, 50)
            self.ordered_company.setText(customer)
            self.ordered_company.setAlignment(Qt.AlignCenter)
            self.ordered_company.show()

            # Extruder Label
            self.machine_label = QtWidgets.QLabel(self.info_widget)
            self.machine_label.setText("Extruder:")
            self.machine_label.setStyleSheet(label_stylesheet)
            self.machine_label.setFont(font)

            # Show Extruder Value
            self.extruder_val = QtWidgets.QLabel(self.info_widget)
            self.extruder_val.setText(machine)
            self.extruder_val.setFont(font)

            # Product Code Label
            self.code_label = QtWidgets.QLabel(self.info_widget)
            self.code_label.setText("Product Code:")
            self.code_label.setFont(font)
            self.code_label.setStyleSheet(label_stylesheet)

            # Show Product Code Value
            self.product_code_val = QtWidgets.QLabel(self.info_widget)
            self.product_code_val.setText(product_code)
            self.product_code_val.setFont(font)

            # Quantity Order Label
            self.order_label = QtWidgets.QLabel(self.info_widget)
            self.order_label.setText("Quantity Order:")
            self.order_label.setFont(font)
            self.order_label.setStyleSheet(label_stylesheet)

            # Show Order Value
            self.order_val = QtWidgets.QLabel(self.info_widget)
            self.order_val.setText(str(ordered_qty))
            self.order_val.setFont(font)

            # Show Output Label
            self.output_label = QtWidgets.QLabel(self.info_widget)
            self.output_label.setText("Output:")
            self.output_label.setFont(font)
            self.output_label.setStyleSheet(label_stylesheet)

            # Show Output Value
            self.output_val = QtWidgets.QLabel(self.info_widget)
            self.output_val.setText(str(product_output))
            self.output_val.setFont(font)

            # Show formula Label
            self.formula_label = QtWidgets.QLabel(self.info_widget)
            self.formula_label.setText("Formula ID:")
            self.formula_label.setFont(font)
            self.formula_label.setStyleSheet(label_stylesheet)

            # Show Formula ID Value
            self.formulaID_val = QtWidgets.QLabel(self.info_widget)
            self.formulaID_val.setText(str(formula_id))
            self.formulaID_val.setFont(font)

            self.resin_label = QtWidgets.QLabel(self.info_widget)
            self.resin_label.setText("Resin:")
            self.resin_label.setFont(font)
            self.resin_label.setStyleSheet(label_stylesheet)

            # Show Resin Value
            self.resin_val = QtWidgets.QLabel(self.info_widget)
            self.resin_val.setText(str(resin))
            self.resin_val.setFont(font)

            # Lot Label
            self.lot_label = QtWidgets.QLabel(self.info_widget)
            self.lot_label.setText("LOT Number:")
            self.lot_label.setFont(font)
            self.lot_label.setStyleSheet(label_stylesheet)

            # Show Lot Number Value
            self.lotNum_val = QtWidgets.QLabel(self.info_widget)
            self.lotNum_val.setText(str(resin))
            self.lotNum_val.setFont(font)

            self.feedrate_label = QtWidgets.QLabel(self.info_widget)
            self.feedrate_label.setText("Feed Rate:")
            self.feedrate_label.setFont(font)
            self.feedrate_label.setStyleSheet(label_stylesheet)

            # Show Feed Rate Value
            self.feedrate_val = QtWidgets.QLabel(self.info_widget)
            self.feedrate_val.setText(str(feed_rate))
            self.feedrate_val.setFont(font)

            # RPM label
            self.rpm_label = QtWidgets.QLabel(self.info_widget)
            self.rpm_label.setText("RPM:")
            self.rpm_label.setFont(font)
            self.rpm_label.setStyleSheet(label_stylesheet)

            # Show RPM Value
            self.rpm_val = QtWidgets.QLabel(self.info_widget)
            self.rpm_val.setText(str(rpm))
            self.rpm_val.setFont(font)

            # Screen Size Label
            self.screen_size_label = QtWidgets.QLabel(self.info_widget)
            self.screen_size_label.setFont(font)
            self.screen_size_label.setText("Screen Size:")
            self.screen_size_label.setStyleSheet(label_stylesheet)

            # Show Screen Size Value
            self.screenSize_val = QtWidgets.QLabel(self.info_widget)
            self.screenSize_val.setText(str(screen_size))
            self.screenSize_val.setFont(font)

            self.screwconfig_label = QtWidgets.QLabel(self.info_widget)
            self.screwconfig_label.setFont(font)
            self.screwconfig_label.setText("Screw Config:")
            self.screwconfig_label.setStyleSheet(label_stylesheet)

            # Show Screw Config Value
            self.screwConf_val = QtWidgets.QLabel(self.info_widget)
            self.screwConf_val.setText(str(screw_config))
            self.screwConf_val.setFont(font)

            # Output % Label
            self.output_percentage_lbl = QtWidgets.QLabel(self.info_widget)
            self.output_percentage_lbl.setText("Output %:")
            self.output_percentage_lbl.setFont(font)
            self.output_percentage_lbl.setStyleSheet(label_stylesheet)

            # Show Output Percentage Value
            self.outputPercent_val = QtWidgets.QLabel(self.info_widget)
            self.outputPercent_val.setText(str(output_percent))
            self.outputPercent_val.setFont(font)

            # Loss Label
            self.loss_label = QtWidgets.QLabel(self.info_widget)
            self.loss_label.setText("Loss:")
            self.loss_label.setFont(QtGui.QFont(font))
            self.loss_label.setStyleSheet(label_stylesheet)

            # Show Loss Value
            self.loss_val = QtWidgets.QLabel(self.info_widget)
            self.loss_val.setText(str(loss))
            self.loss_val.setFont(font)

            # loss percentage Label
            self.loss_percent_label = QtWidgets.QLabel(self.info_widget)
            self.loss_percent_label.setText("Loss %:")
            self.loss_percent_label.setFont(font)
            self.loss_percent_label.setStyleSheet(label_stylesheet)

            # Show Loss Percentage Value
            self.lossPercent_val = QtWidgets.QLabel(self.info_widget)
            self.lossPercent_val.setText(str(loss_percent))
            self.lossPercent_val.setFont(font)

            # Purge Duration Label
            self.purge_duration_label = QtWidgets.QLabel(self.info_widget)
            self.purge_duration_label.setText("Purge Duration:")
            self.purge_duration_label.setFont(font)
            self.purge_duration_label.setStyleSheet(label_stylesheet)

            # Show Resin Value
            self.purgeDuration_val = QtWidgets.QLabel(self.info_widget)
            self.purgeDuration_val.setText(str(purge_duration))
            self.purgeDuration_val.setFont(font)

            # Production ID
            self.productionID_value = QtWidgets.QLabel(self.info_widget)
            self.productionID_value.setText(str(production_ID))
            self.productionID_value.setFont(font)

            self.productionID_label = QtWidgets.QLabel(self.info_widget)
            self.productionID_label.setText("Production ID:")
            self.productionID_label.setFont(font)
            self.productionID_label.setStyleSheet(label_stylesheet)

            # Order ID
            self.orderID_label = QtWidgets.QLabel(self.info_widget)
            self.orderID_label.setText("Order ID:")
            self.orderID_label.setFont(font)
            self.orderID_label.setStyleSheet(label_stylesheet)

            self.orderID_value = QtWidgets.QLabel(self.info_widget)
            self.orderID_value.setText(str(order_id))
            self.orderID_value.setFont(font)

            # Total Time
            self.total_time = QtWidgets.QLabel(self.info_widget)
            self.total_time.setText("Total Hours:")
            self.total_time.setFont(font)
            self.total_time.setStyleSheet(label_stylesheet)

            self.total_time_value = QtWidgets.QLabel(self.info_widget)
            self.total_time_value.setText(str(total_time))
            self.total_time_value.setFont(font)

            # Output Per Hour
            self.outputPerHour_label = QtWidgets.QLabel(self.info_widget)
            self.outputPerHour_label.setText("Output / Hour:")
            self.outputPerHour_label.setFont(font)
            self.outputPerHour_label.setStyleSheet(label_stylesheet)

            self.outputPerHour_val = QtWidgets.QLabel(self.info_widget)
            self.outputPerHour_val.setText(str(output_per_hour))
            self.outputPerHour_val.setFont(font)

            # Total Input
            self.total_input = QtWidgets.QLabel(self.info_widget)
            self.total_input.setFont(font)
            self.total_input.setText(str(total_input))

            self.totalInput_label = QtWidgets.QLabel(self.info_widget)
            self.totalInput_label.setFont(font)
            self.totalInput_label.setText("Total Input:")
            self.totalInput_label.setStyleSheet(label_stylesheet)

            # Purging Labels
            self.purging_label = QtWidgets.QLabel(self.info_widget)
            self.purging_label.setText("Purging To:")
            self.purging_label.setFont(font)
            self.purging_label.setStyleSheet(label_stylesheet)

            self.purging_val = QtWidgets.QLabel(self.info_widget)
            self.purging_val.setFont(font)
            self.purging_val.setText(purging)

            # Operator Labels
            self.operator_label = QtWidgets.QLabel(self.info_widget)
            self.operator_label.setFont(font)
            self.operator_label.setText("Operator:")
            self.operator_label.setStyleSheet(label_stylesheet)

            self.operator_value = QtWidgets.QLabel(self.info_widget)
            self.operator_value.setFont(font)
            self.operator_value.setText(operator)

            # Supervisor Labels
            self.supervisor_label = QtWidgets.QLabel(self.info_widget)
            self.supervisor_label.setText("Supervisor")
            self.supervisor_label.setFont(font)
            self.supervisor_label.setStyleSheet(label_stylesheet)

            self.supervisor_value = QtWidgets.QLabel(self.info_widget)
            self.supervisor_value.setText(supervisor)
            self.supervisor_value.setFont(font)

            info_vbox1 = QVBoxLayout(info_widget1)
            info_vbox2 = QVBoxLayout(info_widget2)
            info_vbox3 = QVBoxLayout(info_widget3)
            info_vbox4 = QVBoxLayout(info_widget4)
            info_vbox5 = QVBoxLayout(info_widget5)
            info_vbox6 = QVBoxLayout(info_widget6)

            # First VBOX for Labels
            info_vbox1.addWidget(self.machine_label)
            info_vbox1.addWidget(self.productionID_label)
            info_vbox1.addWidget(self.code_label)
            info_vbox1.addWidget(self.orderID_label)
            info_vbox1.addWidget(self.formula_label)
            info_vbox1.addWidget(self.order_label)
            info_vbox1.addWidget(self.total_time)

            # Second VBOX for Inputs
            info_vbox2.addWidget(self.extruder_val)
            info_vbox2.addWidget(self.productionID_value)
            info_vbox2.addWidget(self.product_code_val)
            info_vbox2.addWidget(self.orderID_value)
            info_vbox2.addWidget(self.formulaID_val)
            info_vbox2.addWidget(self.order_val)
            info_vbox2.addWidget(self.total_time_value)

            # third VBOX for Second Label
            info_vbox3.addWidget(self.totalInput_label)
            info_vbox3.addWidget(self.output_label)
            info_vbox3.addWidget(self.output_percentage_lbl)
            info_vbox3.addWidget(self.outputPerHour_label)
            info_vbox3.addWidget(self.loss_label)
            info_vbox3.addWidget(self.loss_percent_label)
            info_vbox3.addWidget(self.feedrate_label)
            info_vbox3.addWidget(self.rpm_label)

            # Fourth Vbox for the Second Row Inputs Value
            info_vbox4.addWidget(self.total_input)
            info_vbox4.addWidget(self.output_val)
            info_vbox4.addWidget(self.outputPercent_val)
            info_vbox4.addWidget(self.outputPerHour_val)
            info_vbox4.addWidget(self.loss_val)
            info_vbox4.addWidget(self.lossPercent_val)
            info_vbox4.addWidget(self.feedrate_val)
            info_vbox4.addWidget(self.rpm_val)

            # Fifth VBOX for Labels
            info_vbox5.addWidget(self.purging_label)
            info_vbox5.addWidget(self.resin_label)
            info_vbox5.addWidget(self.screwconfig_label)
            info_vbox5.addWidget(self.screen_size_label)
            info_vbox5.addWidget(self.purge_duration_label)
            info_vbox5.addWidget(self.operator_label)
            info_vbox5.addWidget(self.supervisor_label)

            info_vbox6.addWidget(self.purging_val)
            info_vbox6.addWidget(self.resin_val)
            info_vbox6.addWidget(self.screwConf_val)
            info_vbox6.addWidget(self.screenSize_val)
            info_vbox6.addWidget(self.purgeDuration_val)
            info_vbox6.addWidget(self.operator_value)
            info_vbox6.addWidget(self.supervisor_value)

            info_widget1.show()
            info_widget2.show()
            info_widget3.show()
            info_widget4.show()
            info_widget5.show()
            info_widget6.show()

            # Create 3 tables for Time, Materials and Temperature
            try:
                self.time_table = QtWidgets.QTableWidget(self.production_widget)
                self.time_table.setGeometry(20, 450, 330, 300)
                self.time_table.setColumnCount(3)
                self.time_table.setRowCount(len(time_start) + 2)
                self.time_table.setRowCount(9)
                self.time_rows = len(time_end)
                self.time_cols = 3
                self.time_table.setHorizontalHeaderLabels(["Time Start", "Time End", "Time Diff"])
                total_timediff = timedelta()

                for i in range(self.time_rows):
                    for j in range(self.time_cols):
                        if j == 0:
                            item = QtWidgets.QTableWidgetItem(
                                time_start[i].strftime("%b-%d-%Y %H:%M"))  # Convert to string
                            item.setFlags(item.flags() & ~Qt.ItemIsEditable)  # Make the cells unable to be edited
                            self.time_table.setItem(i, j, item)
                        elif j == 1:
                            item = QtWidgets.QTableWidgetItem(
                                time_end[i].strftime("%b-%d-%Y %H:%M"))  # Convert to string
                            item.setFlags(item.flags() & ~Qt.ItemIsEditable)  # Make the cells unable to be edited
                            self.time_table.setItem(i, j, item)
                        else:
                            datetime1 = time_start[i]
                            datetime2 = time_end[i]
                            t_diff = datetime2 - datetime1
                            # Convert timedelta to a string representation (e.g., "X days, HH:MM:SS")
                            total_timediff = total_timediff + t_diff
                            t_diff_str = str(t_diff)
                            item = QtWidgets.QTableWidgetItem(t_diff_str)
                            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                            self.time_table.setItem(i, j, item)

                # Convert it to Hours and Minutes only
                total_hours = total_timediff.days * 24 + total_timediff.seconds // 3600
                total_minutes = (total_timediff.seconds % 3600) // 60

                # Populate the Time Table
                self.time_table.setItem(self.time_rows + 1, 1, QtWidgets.QTableWidgetItem("Total"))
                self.time_table.setItem(self.time_rows + 1, 2,
                                        QtWidgets.QTableWidgetItem(str(total_hours) + ":" + str(total_minutes)))
                self.time_table.show()

                # Material Table
                self.material_table = QtWidgets.QTableWidget(self.production_widget)
                self.material_table.setRowCount(len(materials))
                self.material_table.setColumnCount(2)
                self.material_table.setRowCount(9)
                self.material_table.setGeometry(350, 450, 225, 300)
                self.material_table.setHorizontalHeaderLabels(["Materials", "Quantity(Kg)"])

                # Populate the Materials Table
                for key in list(materials.keys()):
                    key_item = QtWidgets.QTableWidgetItem(str(key))
                    value_item = QtWidgets.QTableWidgetItem(str(materials[key]))
                    key_item.setFlags(key_item.flags() & ~Qt.ItemIsEditable)  # Make the cells unable to be edited
                    value_item.setFlags(value_item.flags() & ~Qt.ItemIsEditable)  # Make the cells unable to be edited
                    self.material_table.setItem(list(materials.keys()).index(key), 0, key_item)
                    self.material_table.setItem(list(materials.keys()).index(key), 1, value_item)

                self.material_table.show()

                # Temperature Table
                self.temp_table = QtWidgets.QTableWidget(self.production_widget)
                self.temp_table.setGeometry(575, 450, 140, 300)
                self.temp_table.setRowCount(12)
                self.temp_table.setColumnCount(1)
                self.temp_table.setVerticalHeaderLabels(["Z" + str(i + 1) for i in range(12)])
                self.temp_table.setHorizontalHeaderLabels(["Temperature"])

                # Populate the Table
                for i in range(len(temperature)):
                    item = QtWidgets.QTableWidgetItem(str(temperature[i]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                    self.temp_table.setItem(i, 0, item)

                self.temp_table.show()
            except Exception as e:
                print(e)

            self.group_box = QtWidgets.QGroupBox(self.production_widget)
            self.group_box.setGeometry(715, 450, 276, 301)
            self.group_box.setTitle("Remarks")
            self.group_box.setFont(QtGui.QFont("Arial", 10))

            self.show_remarks = QtWidgets.QTextEdit(self.group_box)
            self.show_remarks.setGeometry(0, 20, 230, 130)
            self.show_remarks.setAutoFillBackground(True)
            self.show_remarks.setText(remarks)
            self.show_remarks.setEnabled(False)
            self.show_remarks.show()

            self.group_box.show()

            self.export_btn = QtWidgets.QPushButton(self.production_widget)
            self.export_btn.setGeometry(730, 610, 100, 30)
            self.export_btn.setText("Export")
            self.export_btn.clicked.connect(exportToExcel)
            self.export_btn.show()

        def add_entry():

            self.entry_widget = QtWidgets.QWidget()
            self.entry_widget.setGeometry(300, 100, 800, 750)
            self.entry_widget.setWindowIcon(QtGui.QIcon("setting.png"))
            self.entry_widget.setWindowTitle("ADD EXTRUDER DATA")
            self.entry_widget.setStyleSheet("background-color : rgb(240,240,240);")
            self.entry_widget.setWindowModality(Qt.ApplicationModal)
            self.entry_widget.show()

            def get_entries():
                # get the data from the tables

                try:

                    # Time Table
                    temp_row = time_table.rowCount()
                    time_start = []
                    time_end = []
                    outputs = []

                    # Getting the data from the Time Table
                    for i in range(self.time_entry):
                        time_start.append(time_table.item(i, 0))  # time start
                        time_end.append(time_table.item(i, 1))  # time end
                        outputs.append(time_table.item(i, 2))

                    # Removing Null Values
                    time_start = [i for i in time_start if i is not None]
                    time_start = [i.text() for i in time_start]
                    time_end = [i for i in time_end if i is not None]
                    time_end = [i.text() for i in time_end]
                    outputs = [i for i in outputs if i is not None]
                    outputs = [i.text() for i in outputs]

                    total_time = timedelta()
                    try:
                        for i in range(len(time_start)):
                            t_start = datetime.strptime(time_start[i], "%m-%d-%Y %H:%M")
                            t_end = datetime.strptime(time_end[i], "%m-%d-%Y %H:%M")
                            total_time = total_time + (t_start - t_end)
                    except Exception as e:
                        print(e)



                    hours = str(int(total_time.total_seconds() // 3600))
                    minutes = str((int(total_time.total_seconds() % 3600) // 60))
                    seconds = str(int(total_time.total_seconds() % 60))

                    total_hours = round(abs(total_time.total_seconds() / 3600), 2)

                    time_start = ', '.join(["'{}'".format(time) for time in time_start])
                    time_end = ', '.join(["'{}'".format(time) for time in time_end])

                    # Getting the Data for temperature
                    temperature = []
                    for i in range(temperature_table.rowCount()):
                        temperature.append(temperature_table.item(i, 0))

                    temperature = [i for i in temperature if i is not None]
                    temperature = [i.text() for i in temperature]

                    # Declare additional variables need here like loss percentage
                    output_percent = round((float(product_output_input.text()) / float(product_input.text())) * 100,
                                           4)  # Round to the 4th decimal
                    loss_percent = round((float(loss_input.text()) / float(product_input.text())) * 100,
                                         4)  # Round to the 4th decimal
                    purge_duration = timedelta()
                    outputPerHour = round(float(product_output_input.text()) / total_hours, 4)

                    try:
                        purge_start = datetime.strptime(purgeStart_input.text(), "%Y-%m-%d %H:%M")
                        purge_end = datetime.strptime(purgeEnd_input.text(), "%Y-%m-%d %H:%M")
                        purge_duration = purge_end - purge_start


                    except:

                        purge_start = datetime.strptime(
                            datetime.today().strftime("%Y-%m-%d") + " " + purgeStart_input.text(), "%Y-%m-%d %H:%M")
                        purge_end = datetime.strptime(
                            datetime.today().strftime("%Y-%m-%d") + " " + purgeEnd_input.text(),
                            "%Y-%m-%d %H:%M")
                        purge_duration = (purge_end - purge_start).total_seconds()

                    purge_duration = purge_duration // 60
                    # SQL command here to insert Items
                    self.cursor.execute(
                        f"SELECT materials FROM production_merge WHERE production_id = '{productionID_input.text()}'")

                    self.total_mats = json.dumps(self.total_mats)
                    self.total_mats = self.total_mats.replace('\\', "")

                    # Convert the list to string
                    temperature = str(temperature).replace("[", "").replace("]", "")
                    outputs = str(outputs).replace("[", "").replace("]", "")
                    self.lot_numberList = str(self.lot_numberList).replace("[", "").replace("]", "")

                    try:

                        self.cursor.execute(f"""INSERT INTO extruder( machine, qty_order, total_output, customer,
                                        formula_id, product_code, order_id, total_time, time_start, time_end, output_percent,
                                        loss, loss_percent, materials, purging, resin, purge_duration, screw_config, feed_rate, 
                                        rpm, screen_size, operator, supervisor, temperature, outputs, output_per_hour, production_id, total_input,
                                        remarks, lot_number) 
                                        VALUES('{machine_input.currentText()}', '{orderedQuantity_input.text()}', '{product_output_input.text()}',
                                        '{customer_input.text().replace("'", "''")}', '{self.formulaID_input.text()}', '{productCode_input.text()}',
                                        '{order_number_input.text()}', '{total_hours}', ARRAY[{time_start}]::timestamp[], ARRAY[{time_end}]::timestamp[], 
                                        '{str(output_percent)}', '{loss_input.text()}', '{loss_percent}', '{self.total_mats}', '{purging_input.text()}',
                                         '{resin_input.text()}', {purge_duration}, '{screwConf_input.text()}', '{feedRate_input.text()}',
                                         '{rpm_input.text()}','{screenSize_input.text()}', '{operator_input.text()}', '{supervisor_input.text()}',
                                         ARRAY[{temperature}]::INTEGER[], ARRAY[{outputs}]::FLOAT[], {outputPerHour}, {productionID_input.text()},
                                         {product_input.text()},'{self.remarks_textBox.toPlainText()}', 
                                         ARRAY[{self.lot_numberList}]::VARCHAR[])

                                                """)
                        print("query successful")
                        self.conn.commit()
                        clear_inputs()
                    except Exception as e:
                        print("Insert Failed")
                        QMessageBox.critical(self.entry_widget, "ERROR", "INVALID ENTRY")
                        print(e)
                        self.conn.rollback()
                except Exception as e:
                    print(e)
                    QMessageBox.critical(self.entry_widget, "ERROR", "INVALID ENTRY")
                    return

            def select_production():

                try:
                    self.table.deleteLater()
                    self.table2.deleteLater()
                    self.table3.deleteLater()

                except:

                    pass

                def show_table():
                    self.table2.clearContents()

                    item = self.table.selectedItems()
                    item = [i.text() for i in item]
                    self.cursor.execute(f"""
                    SELECT production_id, lot_number, t_qtyreq, materials
                    FROM production_merge
                    WHERE production_id = '{item[0]}'
                    """)

                    result = self.cursor.fetchall()
                    result = result[0]

                    material = result[-1]

                    try:

                        for keys in list(material.keys()):
                            key = QTableWidgetItem(str(keys))
                            value = QTableWidgetItem(str(material[keys]))
                            self.table2.setItem(list(material.keys()).index(keys), 0, key)
                            key.setFlags(key.flags() & ~Qt.ItemIsEditable)
                            self.table2.setItem(list(material.keys()).index(keys), 1, value)
                            value.setFlags(value.flags() & ~Qt.ItemIsEditable)
                            self.table2.show()

                        label2.setText(str(result[2]))
                        prod_id = QTableWidgetItem(str(result[0]))
                        prod_id.setTextAlignment(Qt.AlignCenter)  # Align Center
                        lot_num = QTableWidgetItem(str(result[1]))
                        lot_num.setTextAlignment(Qt.AlignCenter)
                        self.table3.setItem(0, 0, prod_id)
                        prod_id.setFlags(prod_id.flags() & ~Qt.ItemIsEditable)
                        self.table3.setItem(0, 1, lot_num)
                        lot_num.setFlags(lot_num.flags() & ~Qt.ItemIsEditable)

                    except Exception as e:
                        print(e)

                self.added_entry = 0  # for counting the lot numbers added
                self.total_mats = {}
                self.total_materialQty = 0
                self.total_quantity_order = 0
                self.total_output = 0
                self.lot_numberList = []
                self.total_outputPercent = 0

                def push_data():
                    item = self.table.selectedItems()
                    item = [i.text() for i in item]

                    self.cursor.execute(f"""
                                                                SELECT * FROM production_merge
                                                                WHERE production_id = '{item[0]}' 

                                                                """)
                    result = self.cursor.fetchall()
                    result = result[0]

                    # Unpack the result
                    prod_id = result[0]
                    customer = result[2]
                    formula_id = result[3]
                    product_code = result[5]
                    product_color = result[6]
                    lot_number = result[9]
                    order_number = result[10]
                    machine_name = result[14]
                    quantity_order = result[15]
                    output_quantity = result[17]
                    remarks = result[18]
                    materials = result[-1]

                    self.added_entry += 1
                    self.lot_numberList.append(lot_number)
                    self.total_output += output_quantity
                    self.total_quantity_order += quantity_order

                    for key in materials.keys():
                        if key in list(self.total_mats.keys()):
                            self.total_mats[key] = self.total_mats[key] + materials[key]
                            self.total_materialQty += materials[key]
                        else:
                            self.total_mats[key] = materials[key]
                            self.total_materialQty += materials[key]

                    # Set the Text to the Extruder Entry Form
                    productionID_input.setText(prod_id)
                    customer_input.setText(customer)
                    productCode_input.setText(product_code)
                    lot_number_input.setText('/'.join(self.lot_numberList))
                    self.formulaID_input.setText(str(formula_id))
                    order_number_input.setText(str(order_number))
                    label2.setText(str(self.total_materialQty))

                    self.cursor.execute(f"""
                    SELECT production_id, lot_number
                    FROM production_merge
                    WHERE product_code = '{product_code}' AND 
                    machine = '{machine_name}' AND t_fid = '{formula_id}';
                    
                    """)
                    query_result = self.cursor.fetchall()
                    self.table.itemSelectionChanged.disconnect(show_table)
                    self.table.clearSelection()
                    self.table.clear()
                    self.table.setRowCount(len(query_result))
                    self.table.setHorizontalHeaderLabels(["Production ID", "Lot Number"])

                    for row in range(len(query_result)):
                        prod = QTableWidgetItem(query_result[row][0])
                        prod.setFlags(prod.flags() & ~Qt.ItemIsEditable)
                        lot = QTableWidgetItem(query_result[row][1])
                        lot.setFlags(lot.flags() & ~Qt.ItemIsEditable)
                        self.table.setItem(row, 0, prod)
                        self.table.setItem(row, 1, lot)
                    self.table.itemSelectionChanged.connect(show_table)
                    self.table.show()

                    self.table2.clearSelection()
                    self.table2.clear()
                    self.table2.setHorizontalHeaderLabels(["Materials", "Quantity"])
                    for keys in list(self.total_mats.keys()):
                        key = QTableWidgetItem(str(keys))
                        value = QTableWidgetItem(str(self.total_mats[keys]))
                        self.table2.setItem(list(self.total_mats.keys()).index(keys), 0, key)
                        key.setFlags(key.flags() & ~Qt.ItemIsEditable)
                        self.table2.setItem(list(self.total_mats.keys()).index(keys), 1, value)
                        value.setFlags(value.flags() & ~Qt.ItemIsEditable)
                    self.table2.show()

                def search():
                    try:
                        self.table.itemSelectionChanged.disconnect(show_table)
                        self.table.clearContents()
                        self.cursor.execute(f"""
                                            SELECT production_id, lot_number
                                            FROM production_merge
                                            WHERE lot_number ILIKE '%{search_bar.text()}%'
                                            """)
                        search_result = self.cursor.fetchall()

                        for i in range(len(search_result)):
                            item_pair = search_result[i]

                            item = QTableWidgetItem(item_pair[0])
                            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                            item2 = QTableWidgetItem(item_pair[1])
                            item2.setFlags(item2.flags() & ~Qt.ItemIsEditable)
                            self.table.setItem(i, 0, item)
                            self.table.setItem(i, 1, item2)

                        self.table.itemSelectionChanged.connect(show_table)
                    except Exception as e:
                        print(e)

                def clear():

                    try:
                        self.cursor.execute("""
                                            SELECT production_id, lot_number
                                            FROM production_merge

                                                                                """)
                        result = self.cursor.fetchall()
                        self.table.setRowCount(len(result))
                        self.table.setColumnCount(2)
                        self.table.setHorizontalHeaderLabels(["Production ID", "Lot Number"])

                        self.table.itemSelectionChanged.disconnect(show_table)
                        self.table.clearSelection()
                        self.table.clear()
                        search_bar.clear()

                        self.table2.clearContents()
                        self.table3.clearContents()
                        self.table.itemSelectionChanged.connect(show_table)

                    except Exception as e:
                        self.table.setHorizontalHeaderLabels(["Production ID", "Lot Number"])
                        self.table.itemSelectionChanged.connect(show_table)

                    try:
                        self.table.setHorizontalHeaderLabels(["Production ID", "Lot Number"])
                        for i in range(len(result)):
                            prod_id = QTableWidgetItem(str(result[i][0]))
                            lot_num = QTableWidgetItem(str(result[i][1]))
                            self.table.setItem(i, 0, prod_id)
                            prod_id.setFlags(prod_id.flags() & ~Qt.ItemIsEditable)
                            self.table.setItem(i, 1, lot_num)
                            lot_num.setFlags(lot_num.flags() & ~Qt.ItemIsEditable)
                            self.table.show()

                    except Exception as e:
                        print(e)

                def close_selection():
                    self.selectProd_widget.close()

                self.cursor.execute("""
                SELECT production_id, lot_number
                FROM production_merge

                """)
                result = self.cursor.fetchall()

                self.selectProd_widget = QtWidgets.QWidget()
                self.selectProd_widget.setGeometry(400, 200, 800, 630)
                self.selectProd_widget.setFixedSize(800, 600)

                search_bar = QtWidgets.QLineEdit(self.selectProd_widget)
                search_bar.setGeometry(40, 25, 170, 25)
                search_bar.setFont(QtGui.QFont("Arial", 10))
                search_bar.setPlaceholderText("Search Lot Number")
                search_bar.show()

                search_btn = QtWidgets.QPushButton(self.selectProd_widget)
                search_btn.setGeometry(210, 25, 70, 27)
                search_btn.setText("Search")
                search_btn.clicked.connect(search)
                search_btn.show()

                clear_btn = QtWidgets.QPushButton(self.selectProd_widget)
                clear_btn.setGeometry(280, 25, 70, 27)
                clear_btn.setText("Clear")
                clear_btn.clicked.connect(clear)
                clear_btn.show()

                # ProductionId and Lot Number Table widget
                self.table = QtWidgets.QTableWidget(self.selectProd_widget)
                self.table.setGeometry(0, 70, 350, 500)
                self.table.setColumnCount(2)
                self.table.setRowCount(len(result))
                self.table.setHorizontalHeaderLabels(["Production ID", "Lot Number"])
                self.table.setColumnWidth(1, 180)
                self.table.setColumnWidth(0, 150)
                self.table.setFont(QtGui.QFont("Arial", 12))
                self.table.setStyleSheet('color: rgb(0,109,189);')
                self.table.verticalHeader().setVisible(False)
                try:
                    for i in range(len(result)):
                        prod_id = QTableWidgetItem(str(result[i][0]))
                        lot_num = QTableWidgetItem(str(result[i][1]))
                        self.table.setItem(i, 0, prod_id)
                        prod_id.setFlags(prod_id.flags() & ~Qt.ItemIsEditable)
                        self.table.setItem(i, 1, lot_num)
                        lot_num.setFlags(lot_num.flags() & ~Qt.ItemIsEditable)
                except Exception as e:
                    print(e)

                self.table.show()
                self.table.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)
                self.table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
                self.table.itemSelectionChanged.connect(show_table)

                # Materials Table
                self.table2 = QtWidgets.QTableWidget(self.selectProd_widget)
                self.table2.setGeometry(350, 70, 450, 500)
                self.table2.setColumnCount(2)
                self.table2.setRowCount(16)
                self.table2.setColumnWidth(0, 205)
                self.table2.setColumnWidth(1, 225)
                self.table2.verticalHeader().setVisible(False)
                self.table2.setHorizontalHeaderLabels(["Materials", "Quantity"])
                self.table2.show()

                # Table 3 For showing Selected ProdID and Lot Number
                self.table3 = QtWidgets.QTableWidget(self.selectProd_widget)
                self.table3.setGeometry(380, 0, 402, 60)
                self.table3.setColumnCount(2)
                self.table3.setRowCount(1)
                self.table3.setHorizontalHeaderLabels(["Production ID", "Lot Number"])
                self.table3.setColumnWidth(0, 200)
                self.table3.setColumnWidth(1, 200)
                self.table3.setRowHeight(0, 35)
                font = QtGui.QFont("Arial", 12)
                font.setBold(True)
                self.table3.setFont(font)
                self.table3.setStyleSheet("color: rgb(0,109,189) ")
                self.table3.horizontalHeader().setStyleSheet("""
                    QHeaderView::section{
                        font-weight: bold;
                        background-color: rgb(0,109,189);
                        color: white;
                    }
                """)

                for i in range(2):
                    item = QTableWidgetItem("")
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)  # Make the cells unable to be edited
                    self.table3.setItem(0, i, item)

                self.table3.verticalHeader().setVisible(False)
                self.table3.show()

                label1 = QtWidgets.QLabel(self.selectProd_widget)
                label1.setGeometry(230, 570, 120, 30)
                label1.setText("Total QTY (Kg)")
                label1.setFont(QtGui.QFont("Arial", 11))
                label1.setAlignment(Qt.AlignCenter)
                label1.setStyleSheet('border: 1px solid black;')
                label1.show()

                label2 = QtWidgets.QLabel(self.selectProd_widget)
                label2.setGeometry(350, 570, 200, 30)
                label2.setFont(QtGui.QFont("Arial", 15))
                label2.setAlignment(Qt.AlignCenter)
                label2.setStyleSheet("background-color: white; color: blue;")
                label2.show()

                # Save Button
                save_prod = QtWidgets.QPushButton(self.selectProd_widget)
                save_prod.setGeometry(660, 570, 70, 30)
                save_prod.setText("Push Data")
                save_prod.clicked.connect(push_data)
                save_prod.show()

                close = QtWidgets.QPushButton(self.selectProd_widget)
                close.setGeometry(730, 570, 70, 30)
                close.setText("Close")
                close.clicked.connect(close_selection)
                close.show()

                self.selectProd_widget.setWindowModality(Qt.ApplicationModal)
                self.selectProd_widget.show()

            self.time_entry = 0

            def add_time():

                product_output_input.setText(str(float(product_output_input.text()) + float(output_lineEdit.text())))

                item1 = QTableWidgetItem(time_start_input.text())
                item2 = QTableWidgetItem(time_end_input.text())
                item3 = QTableWidgetItem(output_lineEdit.text())

                item3.setTextAlignment(Qt.AlignCenter)

                time_table.setItem(self.time_entry, 0, item1)
                time_table.setItem(self.time_entry, 1, item2)
                time_table.setItem(self.time_entry, 2, item3)

                self.time_entry += 1

                output_lineEdit.clear()

            def reset_table():
                time_table.clearContents()
                product_output_input.setText("0.0")
                self.time_entry = 0

            def loss_auto():
                if product_output_input.text() != "":
                    try:
                        loss_input.setText(
                            str(round(float(product_input.text()) - float(product_output_input.text()), 4)))
                    except:
                        loss_input.setText("INVALID")

            def clear_inputs():
                try:

                    productionID_input.clear()
                    machine_input.clear()
                    customer_input.clear()
                    orderedQuantity_input.clear()
                    productCode_input.clear()
                    product_output_input.setText("0.0")
                    self.formulaID_input.clear()
                    lot_number_input.clear()
                    feedRate_input.clear()
                    rpm_input.clear()
                    screenSize_input.clear()
                    screwConf_input.clear()
                    loss_input.clear()
                    purgeStart_input.setText("00:00")
                    purgeEnd_input.setText("00:00")
                    operator_input.clear()
                    supervisor_input.clear()
                    order_number_input.clear()
                    resin_input.clear()
                    purging_input.clear()
                    product_input.clear()
                    time_table.clearContents()
                    temperature_table.clearContents()
                    self.remarks_textBox.clear()
                    self.total_mats = {}

                except:
                    pass

            # Create two new widget for the VBOX Layout
            self.leftInput_side = QtWidgets.QWidget(self.entry_widget)
            self.leftInput_side.setGeometry(0, 0, 400, 450)
            self.leftInput_side.show()

            self.right_side = QtWidgets.QWidget(self.entry_widget)
            self.right_side.setGeometry(400, 0, 400, 450)
            self.right_side.show()

            # Create Vertical Box Layout
            self.left_vbox = QtWidgets.QFormLayout(self.leftInput_side)
            self.left_vbox.setSpacing(20)
            self.right_vbox = QtWidgets.QFormLayout(self.right_side)
            self.right_vbox.setSpacing(20)

            font = QtGui.QFont("Berlin Sans FB", 14)

            productionID_label = QtWidgets.QLabel()
            productionID_label.setText("Production ID")
            productionID_label.setFont(font)

            customer_label = QtWidgets.QLabel()
            customer_label.setText("Customer")
            customer_label.setFont(font)

            machine_label = QtWidgets.QLabel()
            machine_label.setText("Machine No.")
            machine_label.setFont(font)

            productCode_label = QtWidgets.QLabel()
            productCode_label.setText("Product Code")
            productCode_label.setFont(font)

            productOutput_label = QtWidgets.QLabel()
            productOutput_label.setText("Output (kg)")
            productOutput_label.setFont(font)

            formulaID_label = QtWidgets.QLabel()
            formulaID_label.setText("Formula ID")
            formulaID_label.setFont(font)

            lotnumber_label = QtWidgets.QLabel()
            lotnumber_label.setText("Lot Number")
            lotnumber_label.setFont(font)

            orderedQuantity_label = QtWidgets.QLabel()
            orderedQuantity_label.setText("Ordered Qty")
            orderedQuantity_label.setFont(font)

            feedrate_label = QtWidgets.QLabel()
            feedrate_label.setText("Feed Rate")
            feedrate_label.setFont(font)

            rpm_label = QtWidgets.QLabel()
            rpm_label.setText("RPM")
            rpm_label.setFont(font)

            screenSize_label = QtWidgets.QLabel()
            screenSize_label.setText("Screen Size")
            screenSize_label.setFont(font)

            screwConf_label = QtWidgets.QLabel()
            screwConf_label.setText("Screw Config")
            screwConf_label.setFont(font)

            loss_label = QtWidgets.QLabel()
            loss_label.setText("Loss")
            loss_label.setFont(font)

            purgeStart_label = QtWidgets.QLabel()
            purgeStart_label.setText("Purge Start")
            purgeStart_label.setFont(font)

            purgeEnd_label = QtWidgets.QLabel()
            purgeEnd_label.setText("Purge End")
            purgeEnd_label.setFont(font)

            remarks_label = QtWidgets.QLabel()

            operator_label = QtWidgets.QLabel()
            operator_label.setText("Operator")
            operator_label.setFont(font)

            supervisor_label = QtWidgets.QLabel()
            supervisor_label.setText("Supervisor")
            supervisor_label.setFont(font)

            order_number_lbl = QtWidgets.QLabel()
            order_number_lbl.setText("Order Number")
            order_number_lbl.setFont(font)

            resin_label = QtWidgets.QLabel()
            resin_label.setText("Resin")
            resin_label.setFont(font)

            purging_label = QtWidgets.QLabel()
            purging_label.setText("Purging")
            purging_label.setFont(font)

            product_input_label = QtWidgets.QLabel()
            product_input_label.setText("Input")
            product_input_label.setFont(font)

            # QLineEdit Boxes
            productionID_input = QtWidgets.QLineEdit()
            productionID_input.setFixedHeight(25)
            productionID_input.setEnabled(False)
            productionID_input.setAlignment(Qt.AlignCenter)
            productionID_input.setStyleSheet("background-color: white; border: 1px solid black")

            machine_input = QtWidgets.QComboBox()
            machine_input.setFixedHeight(25)
            machine_input.addItem("Extruder 1")
            machine_input.addItem("Extruder 2")
            machine_input.addItem("Extruder 3")
            machine_input.addItem("Extruder 4")
            machine_input.addItem("Extruder 5")
            machine_input.addItem("Extruder 6")
            machine_input.setStyleSheet("background-color: white; border: 1px solid black")

            customer_input = QtWidgets.QLineEdit()
            customer_input.setFixedHeight(25)
            customer_input.setEnabled(False)
            customer_input.setAlignment(Qt.AlignCenter)
            customer_input.setStyleSheet("background-color: white; border: 1px solid black")

            orderedQuantity_input = QtWidgets.QLineEdit()
            orderedQuantity_input.setAlignment(Qt.AlignCenter)
            orderedQuantity_input.setFixedHeight(25)
            orderedQuantity_input.setStyleSheet("background-color: white; border: 1px solid black")

            productCode_input = QtWidgets.QLineEdit()
            productCode_input.setFixedHeight(25)
            productCode_input.setEnabled(False)
            productCode_input.setAlignment(Qt.AlignCenter)
            productCode_input.setStyleSheet("background-color: white; border: 1px solid black")

            product_output_input = QtWidgets.QLineEdit()
            product_output_input.setFixedHeight(25)
            product_output_input.setEnabled(False)
            product_output_input.setText("0.0")
            product_output_input.setAlignment(Qt.AlignCenter)
            product_output_input.setStyleSheet("background-color: white; border: 1px solid black")

            self.formulaID_input = QtWidgets.QLineEdit()
            self.formulaID_input.setAlignment(Qt.AlignCenter)
            self.formulaID_input.setFixedHeight(25)
            self.formulaID_input.setEnabled(False)
            self.formulaID_input.setStyleSheet("background-color: white; border: 1px solid black")

            lot_number_input = QtWidgets.QLineEdit()
            lot_number_input.setAlignment(Qt.AlignCenter)
            lot_number_input.setFixedHeight(25)
            lot_number_input.setEnabled(False)
            lot_number_input.setStyleSheet("background-color: white; border: 1px solid black")

            feedRate_input = QtWidgets.QLineEdit()
            feedRate_input.setFixedHeight(25)
            feedRate_input.setStyleSheet("background-color: white; border: 1px solid black")

            rpm_input = QtWidgets.QLineEdit()
            rpm_input.setFixedHeight(25)
            rpm_input.setStyleSheet("background-color: white; border: 1px solid black")

            screenSize_input = QtWidgets.QLineEdit()
            screenSize_input.setFixedHeight(25)
            screenSize_input.setStyleSheet("background-color: white; border: 1px solid black")

            screwConf_input = QtWidgets.QLineEdit()
            screwConf_input.setFixedHeight(25)
            screwConf_input.setStyleSheet("background-color: white; border: 1px solid black ")

            loss_input = QtWidgets.QLineEdit()
            loss_input.setFixedHeight(25)
            loss_input.setEnabled(False)
            loss_input.setAlignment(Qt.AlignCenter)
            loss_input.setStyleSheet("background-color: white; border: 1px solid black")

            purgeStart_input = QtWidgets.QLineEdit()
            purgeStart_input.setFixedHeight(25)
            purgeStart_input.setAlignment(Qt.AlignCenter)
            purgeStart_input.setStyleSheet("background-color: white; border: 1px solid black")
            purgeStart_input.setText("00:00")

            purgeEnd_input = QtWidgets.QLineEdit()
            purgeEnd_input.setFixedHeight(25)
            purgeEnd_input.setAlignment(Qt.AlignCenter)
            purgeEnd_input.setStyleSheet("background-color: white; border: 1px solid black")
            purgeEnd_input.setText("00:00")

            operator_input = QtWidgets.QLineEdit()
            operator_input.setFixedHeight(25)
            operator_input.setAlignment(Qt.AlignCenter)
            operator_input.setStyleSheet("background-color: white; border: 1px solid black")

            supervisor_input = QtWidgets.QLineEdit()
            supervisor_input.setFixedHeight(25)
            supervisor_input.setAlignment(Qt.AlignCenter)
            supervisor_input.setStyleSheet("background-color: white; border: 1px solid black")

            order_number_input = QtWidgets.QLineEdit()
            order_number_input.setFixedHeight(25)
            order_number_input.setEnabled(False)
            order_number_input.setAlignment(Qt.AlignCenter)
            order_number_input.setStyleSheet("background-color: white; border: 1px solid black")

            resin_input = QtWidgets.QLineEdit()
            resin_input.setFixedHeight(25)
            resin_input.setAlignment(Qt.AlignCenter)
            resin_input.setStyleSheet("background-color: white; border: 1px solid black")

            purging_input = QtWidgets.QLineEdit()
            purging_input.setFixedHeight(25)
            purging_input.setAlignment(Qt.AlignCenter)
            purging_input.setStyleSheet("background-color: white; border: 1px solid black")

            product_input = QtWidgets.QLineEdit()
            product_input.setFixedHeight(25)
            product_input.setAlignment(Qt.AlignCenter)
            product_input.setStyleSheet("background-color: white; border: 1px solid black")
            product_input.textChanged.connect(loss_auto)

            self.groupBoxRemarks = QtWidgets.QGroupBox(self.entry_widget)
            self.groupBoxRemarks.setGeometry(600, 500, 200, 150)
            self.groupBoxRemarks.setTitle("Remarks")
            self.groupBoxRemarks.show()

            self.remarks_textBox = QtWidgets.QTextEdit(self.groupBoxRemarks)
            self.remarks_textBox.setGeometry(0, 20, 200, 130)
            self.remarks_textBox.show()

            # Left Side of Vertical Box
            self.left_vbox.addRow(productionID_label, productionID_input)
            self.left_vbox.addRow(productCode_label, productCode_input)
            self.left_vbox.addRow(customer_label, customer_input)
            self.left_vbox.addRow(orderedQuantity_label, orderedQuantity_input)
            self.left_vbox.addRow(lotnumber_label, lot_number_input)
            self.left_vbox.addRow(product_input_label, product_input)
            self.left_vbox.addRow(productOutput_label, product_output_input)
            self.left_vbox.addRow(loss_label, loss_input)
            self.left_vbox.addRow(machine_label, machine_input)
            self.left_vbox.addRow(formulaID_label, self.formulaID_input)
            self.left_vbox.addRow(order_number_lbl, order_number_input)

            # Add widgets to the right Form Box
            self.right_vbox.addRow(feedrate_label, feedRate_input)
            self.right_vbox.addRow(rpm_label, rpm_input)
            self.right_vbox.addRow(screenSize_label, screenSize_input)
            self.right_vbox.addRow(screwConf_label, screwConf_input)
            self.right_vbox.addRow(purging_label, purging_input)
            self.right_vbox.addRow(resin_label, resin_input)
            self.right_vbox.addRow(purgeStart_label, purgeStart_input)
            self.right_vbox.addRow(purgeEnd_label, purgeEnd_input)
            self.right_vbox.addRow(operator_label, operator_input)
            self.right_vbox.addRow(supervisor_label, supervisor_input)

            # Time Table Entry
            time_table = QtWidgets.QTableWidget(self.entry_widget)
            time_table.setGeometry(0, 500, 450, 200)
            time_table.setColumnCount(3)
            time_table.setRowCount(8)
            time_table.setColumnWidth(0, 150)
            time_table.setColumnWidth(1, 150)
            time_table.setStyleSheet("background-color: white;")
            time_table.setFont(QtGui.QFont("Arial", 10))
            time_table.setEnabled(False)

            time_table.setHorizontalHeaderLabels(["Time Start", "Time End", "Output"])
            time_table.show()

            # Temperature Table Entry
            temperature_table = QtWidgets.QTableWidget(self.entry_widget)
            temperature_table.setGeometry(450, 500, 150, 200)
            temperature_table.setColumnCount(1)
            temperature_table.setRowCount(12)
            temperature_table.setStyleSheet("background-color: white;")
            temperature_table.setHorizontalHeaderLabels(["Temperature"])
            temperature_index = ["Z" + str(i + 1) for i in range(12)]  # set the index
            temperature_table.setVerticalHeaderLabels(temperature_index)
            temperature_table.show()

            # Select Production Data Button
            select_prod = QtWidgets.QPushButton(self.entry_widget)
            select_prod.setGeometry(600, 705, 60, 25)
            select_prod.setText("Select")
            select_prod.clicked.connect(select_production)
            select_prod.setCursor(Qt.PointingHandCursor)
            select_prod.show()

            save_btn = QtWidgets.QPushButton(self.entry_widget)
            save_btn.setGeometry(540, 705, 60, 25)
            save_btn.clicked.connect(get_entries)
            save_btn.setText("Save")
            save_btn.setCursor(Qt.PointingHandCursor)
            save_btn.show()

            clear_btn = QtWidgets.QPushButton(self.entry_widget)
            clear_btn.setGeometry(660, 705, 60, 25)
            clear_btn.clicked.connect(clear_inputs)
            clear_btn.setText("Clear")
            clear_btn.show()

            default_date = QtCore.QDateTime(2024, 1, 1, 0, 0)

            time_start_input = QtWidgets.QDateTimeEdit(self.entry_widget)
            time_start_input.setGeometry(30, 475, 120, 25)
            time_start_input.setDisplayFormat("MM-dd-yyyy HH:mm")
            time_start_input.setDateTime(default_date)
            time_start_input.show()

            time_end_input = QtWidgets.QDateTimeEdit(self.entry_widget)
            time_end_input.setGeometry(180, 475, 120, 25)
            time_end_input.setDisplayFormat("MM-dd-yyyy HH:mm")
            time_end_input.setDateTime(default_date)
            time_end_input.show()

            output_lineEdit = QtWidgets.QLineEdit(self.entry_widget)
            output_lineEdit.setGeometry(310, 475, 80, 25)
            output_lineEdit.setAlignment(Qt.AlignCenter)
            output_lineEdit.setStyleSheet("background-color: white; border: 1px solid black")
            output_lineEdit.show()

            self.plus_icon = ClickableLabel(self.entry_widget)
            self.plus_icon.setGeometry(390, 475, 25, 25)
            self.plus_icon.setPixmap(QtGui.QIcon('plus.png').pixmap(25, 25))
            self.plus_icon.setCursor(Qt.PointingHandCursor)
            self.plus_icon.clicked.connect(add_time)
            self.plus_icon.show()

            self.reset_icon = ClickableLabel(self.entry_widget)
            self.reset_icon.setGeometry(425, 475, 25, 25)
            self.reset_icon.setPixmap(QtGui.QIcon('reset.png').pixmap(20, 20))
            self.reset_icon.setCursor(Qt.PointingHandCursor)
            self.reset_icon.clicked.connect(reset_table)
            self.reset_icon.show()

        def update_entry():
            try:
                selected = self.extruder_table.selectedItems()
                selected = [i.text() for i in selected]
                self.cursor.execute(f"SELECT * FROM extruder WHERE process_id = {selected[0]}")
                result = self.cursor.fetchall()
                result = result[0]

            except:
                QMessageBox.critical(self.production_widget, "ERROR", "No Data Selected")
                return

            self.entry_widget = QtWidgets.QWidget()
            self.entry_widget.setGeometry(300, 100, 800, 750)
            self.entry_widget.setStyleSheet("background-color : rgb(240,240,240);")
            self.entry_widget.setWindowModality(Qt.ApplicationModal)
            self.entry_widget.show()

            def update():
                # get the data from the tables

                # Time Table
                temp_row = time_table.rowCount()
                time_start = []
                time_end = []
                outputs = []

                # Getting the data from the Time Table
                for i in range(8):
                    if time_table.item(i, 0) == None:
                        break
                    else:
                        time_start.append(time_table.item(i, 0))  # time start
                        time_end.append(time_table.item(i, 1))  # time end
                        outputs.append(time_table.item(i, 2))

                # Removing Null Values
                time_start = [i for i in time_start if i is not None]
                time_start = [i.text() for i in time_start]
                time_end = [i for i in time_end if i is not None]
                time_end = [i.text() for i in time_end]
                outputs = [i for i in outputs if i is not None]
                outputs = [i.text() for i in outputs]



                total_time = timedelta()

                try:
                    for i in range(len(time_start)):
                        t_start = datetime.strptime(time_start[i], "%Y-%m-%d %H:%M:%S")
                        t_end = datetime.strptime(time_end[i], "%Y-%m-%d %H:%M:%S")
                        total_time = total_time + (t_end - t_start)

                except Exception as e:
                    print(e)

                hours = str(int(total_time.total_seconds() // 3600)).replace('-', '')
                minutes = str((int(total_time.total_seconds() % 3600) // 60))
                seconds = str(int(total_time.total_seconds() % 60))

                total_hours = round(abs(total_time.total_seconds() / 3600), 2)

                time_start = ', '.join(["'{}'".format(time) for time in time_start])
                time_end = ', '.join(["'{}'".format(time) for time in time_end])


                # Getting the Data for temperature
                temperature = []
                for i in range(temperature_table.rowCount()):
                    temperature.append(temperature_table.item(i, 0))

                temperature = [i for i in temperature if i is not None]
                temperature = [i.text() for i in temperature]

                # Declare additional variables need here like loss percentage
                output_percent = round((float(product_output_input.text()) / float(product_input.text())) * 100,
                                       4)  # Round to the 4th decimal
                loss_percent = round((float(loss_input.text()) / float(product_input.text())) * 100,
                                     4)  # Round to the 4th decimal
                outputPerHour = round(float(product_output_input.text()) / total_hours, 4)

                try:
                    purge_duration = timedelta()
                    try:
                        purge_start = datetime.strptime(purgeStart_input.text(), "%Y-%m-%d %H:%M")
                        purge_end = datetime.strptime(purgeEnd_input.text(), "%Y-%m-%d %H:%M")
                        purge_duration = abs(purge_end - purge_start)

                    except:
                        purge_start = datetime.strptime(
                            datetime.today().strftime("%Y-%m-%d") + " " + purgeStart_input.text(), "%Y-%m-%d %H:%M")
                        purge_end = datetime.strptime(
                            datetime.today().strftime("%Y-%m-%d") + " " + purgeEnd_input.text(),
                            "%Y-%m-%d %H:%M")
                        purge_duration = (purge_end - purge_start).total_seconds()

                    purge_duration = purge_duration // 60
                except:
                    purge_duration = result[25]

                # SQL command here to insert Items
                self.cursor.execute(
                    f"SELECT materials FROM production_merge WHERE production_id = '{productionID_input.text()}'")
                material = self.cursor.fetchall()
                material = material[0][0]
                material = json.dumps(material)

                # Convert the list to string
                temperature = str(temperature).replace("[", "").replace("]", "")
                outputs = str(outputs).replace("[", "").replace("]", "")

                try:

                    self.cursor.execute(f"""
                                UPDATE extruder
                                SET  total_time = {total_hours}, machine = '{machine_input.currentText()}', total_input = {product_input.text()}, outputs = ARRAY[{outputs}]::FLOAT[],
                                temperature = ARRAY[{temperature}]::INTEGER[], remarks = '{self.remarks_textBox.toPlainText()}',
                                feed_rate = '{feedRate_input.text()}', rpm = '{rpm_input.text()}', screen_size = '{screenSize_input.text()}',
                                screw_config = '{screwConf_input.text()}', purging = '{purging_input.text()}', resin = '{resin_input.text()}',
                                purge_duration = {purge_duration}, operator = '{operator_input.text()}', supervisor = '{supervisor_input.text()}',
                                time_start = ARRAY[{time_start}]::timestamp[], time_end =  ARRAY[{time_end}]::timestamp[],
                                output_percent = '{str(output_percent)}', loss = '{loss_input.text()}', loss_percent = '{loss_percent}',
                                output_per_hour = '{outputPerHour}', total_output = {product_output_input.text()}
                                WHERE process_id = {selected[0]};
                                ;      
                                """)

                    QMessageBox.information(self.entry_widget, "UPDATE SUCCESSFUL",
                                            f"Successfully Updated \n Form No. {selected[0]}")
                    print("query successful")
                    self.conn.commit()
                    self.entry_widget.close()
                except Exception as e:
                    print("Insert Failed")
                    print(e)
                    self.conn.rollback()

            self.time_entry = len(result[9])

            def add_time():

                item1 = QTableWidgetItem(time_start_input.text())
                item2 = QTableWidgetItem(time_end_input.text())
                item3 = QTableWidgetItem(output_lineEdit.text())

                item3.setTextAlignment(Qt.AlignCenter)

                time_table.setItem(self.time_entry, 0, item1)
                time_table.setItem(self.time_entry, 1, item2)
                time_table.setItem(self.time_entry, 2, item3)

                self.time_entry += 1

                # Auto increment Output
                product_output_input.setText(str(float(product_output_input.text()) + float(output_lineEdit.text())))

                output_lineEdit.clear()

            def reset_table():
                time_table.clearContents()
                product_output_input.setText("0.0")
                self.time_entry = 0

            def loss_auto():
                if product_output_input.text() != "":
                    try:
                        loss_input.setText(
                            str(round(float(product_input.text()) - float(product_output_input.text()), 4)))
                    except:
                        loss_input.setText("INVALID")

            # Create two new widget for the VBOX Layout
            self.leftInput_side = QtWidgets.QWidget(self.entry_widget)
            self.leftInput_side.setGeometry(0, 0, 400, 450)
            self.leftInput_side.show()

            self.right_side = QtWidgets.QWidget(self.entry_widget)
            self.right_side.setGeometry(400, 0, 400, 450)
            self.right_side.show()

            # Create Vertical Box Layout
            self.left_vbox = QtWidgets.QFormLayout(self.leftInput_side)
            self.left_vbox.setSpacing(20)
            self.right_vbox = QtWidgets.QFormLayout(self.right_side)
            self.right_vbox.setSpacing(20)

            font = QtGui.QFont("Berlin Sans FB", 14)

            productionID_label = QtWidgets.QLabel()
            productionID_label.setText("Production ID")
            productionID_label.setFont(font)

            customer_label = QtWidgets.QLabel()
            customer_label.setText("Customer")
            customer_label.setFont(font)

            machine_label = QtWidgets.QLabel()
            machine_label.setText("Machine No.")
            machine_label.setFont(font)

            productCode_label = QtWidgets.QLabel()
            productCode_label.setText("Product Code")
            productCode_label.setFont(font)

            productOutput_label = QtWidgets.QLabel()
            productOutput_label.setText("Output (kg)")
            productOutput_label.setFont(font)

            formulaID_label = QtWidgets.QLabel()
            formulaID_label.setText("Formula ID")
            formulaID_label.setFont(font)

            lotnumber_label = QtWidgets.QLabel()
            lotnumber_label.setText("Lot Number")
            lotnumber_label.setFont(font)

            orderedQuantity_label = QtWidgets.QLabel()
            orderedQuantity_label.setText("Ordered Qty")
            orderedQuantity_label.setFont(font)

            feedrate_label = QtWidgets.QLabel()
            feedrate_label.setText("Feed Rate")
            feedrate_label.setFont(font)

            rpm_label = QtWidgets.QLabel()
            rpm_label.setText("RPM")
            rpm_label.setFont(font)

            screenSize_label = QtWidgets.QLabel()
            screenSize_label.setText("Screen Size")
            screenSize_label.setFont(font)

            screwConf_label = QtWidgets.QLabel()
            screwConf_label.setText("Screw Config")
            screwConf_label.setFont(font)

            loss_label = QtWidgets.QLabel()
            loss_label.setText("Loss")
            loss_label.setFont(font)

            purgeStart_label = QtWidgets.QLabel()
            purgeStart_label.setText("Purge Start")
            purgeStart_label.setFont(font)

            purgeEnd_label = QtWidgets.QLabel()
            purgeEnd_label.setText("Purge End")
            purgeEnd_label.setFont(font)

            remarks_label = QtWidgets.QLabel()

            operator_label = QtWidgets.QLabel()
            operator_label.setText("Operator")
            operator_label.setFont(font)

            supervisor_label = QtWidgets.QLabel()
            supervisor_label.setText("Supervisor")
            supervisor_label.setFont(font)

            order_number_lbl = QtWidgets.QLabel()
            order_number_lbl.setText("Order Number")
            order_number_lbl.setFont(font)

            resin_label = QtWidgets.QLabel()
            resin_label.setText("Resin")
            resin_label.setFont(font)

            purging_label = QtWidgets.QLabel()
            purging_label.setText("Purging")
            purging_label.setFont(font)

            product_input_label = QtWidgets.QLabel()
            product_input_label.setText("Input")
            product_input_label.setFont(font)

            # QLineEdit Boxes
            productionID_input = QtWidgets.QLineEdit()
            productionID_input.setFixedHeight(25)
            productionID_input.setEnabled(False)
            productionID_input.setAlignment(Qt.AlignCenter)
            productionID_input.setStyleSheet("background-color: white; border: 1px solid black")
            productionID_input.setText(str(result[28]))
            productionID_input.setEnabled(False)

            machine_input = QtWidgets.QComboBox()
            machine_input.setFixedHeight(25)
            machine_input.addItem("Extruder 1")
            machine_input.addItem("Extruder 2")
            machine_input.addItem("Extruder 3")
            machine_input.addItem("Extruder 4")
            machine_input.addItem("Extruder 5")
            machine_input.addItem("Extruder 6")
            machine_input.setStyleSheet("background-color: white; border: 1px solid black")
            machine_input.setText(result[1])


            customer_input = QtWidgets.QLineEdit()
            customer_input.setFixedHeight(25)
            customer_input.setEnabled(False)
            customer_input.setAlignment(Qt.AlignCenter)
            customer_input.setStyleSheet("background-color: white; border: 1px solid black")
            customer_input.setText(result[4])
            customer_input.setEnabled(False)

            orderedQuantity_input = QtWidgets.QLineEdit()
            orderedQuantity_input.setAlignment(Qt.AlignCenter)
            orderedQuantity_input.setFixedHeight(25)
            orderedQuantity_input.setStyleSheet("background-color: white; border: 1px solid black")
            orderedQuantity_input.setText(str(result[2]))

            productCode_input = QtWidgets.QLineEdit()
            productCode_input.setFixedHeight(25)
            productCode_input.setEnabled(False)
            productCode_input.setAlignment(Qt.AlignCenter)
            productCode_input.setStyleSheet("background-color: white; border: 1px solid black")
            productCode_input.setText(result[6])

            product_output_input = QtWidgets.QLineEdit()
            product_output_input.setFixedHeight(25)
            product_output_input.setAlignment(Qt.AlignCenter)
            product_output_input.setStyleSheet("background-color: white; border: 1px solid black")
            product_output_input.setText(str(result[3]))

            self.formulaID_input = QtWidgets.QLineEdit()
            self.formulaID_input.setAlignment(Qt.AlignCenter)
            self.formulaID_input.setFixedHeight(25)
            self.formulaID_input.setEnabled(False)
            self.formulaID_input.setStyleSheet("background-color: white; border: 1px solid black")
            self.formulaID_input.setText(str(result[5]))

            lot_number_input = QtWidgets.QLineEdit()
            lot_number_input.setAlignment(Qt.AlignCenter)
            lot_number_input.setFixedHeight(25)
            lot_number_input.setEnabled(False)
            lot_number_input.setStyleSheet("background-color: white; border: 1px solid black")

            try:
                lot_number_input.setText('/'.join(result[30]))
            except:
                lot_number_input.setText(None)

            feedRate_input = QtWidgets.QLineEdit()
            feedRate_input.setFixedHeight(25)
            feedRate_input.setStyleSheet("background-color: white; border: 1px solid black")
            feedRate_input.setText(str(result[18]))

            rpm_input = QtWidgets.QLineEdit()
            rpm_input.setFixedHeight(25)
            rpm_input.setStyleSheet("background-color: white; border: 1px solid black")
            rpm_input.setText(str(result[19]))

            screenSize_input = QtWidgets.QLineEdit()
            screenSize_input.setFixedHeight(25)
            screenSize_input.setStyleSheet("background-color: white; border: 1px solid black")
            screenSize_input.setText(result[20])

            screwConf_input = QtWidgets.QLineEdit()
            screwConf_input.setFixedHeight(25)
            screwConf_input.setStyleSheet("background-color: white; border: 1px solid black ")
            screwConf_input.setText(result[17])

            loss_input = QtWidgets.QLineEdit()
            loss_input.setFixedHeight(25)
            loss_input.setEnabled(False)
            loss_input.setAlignment(Qt.AlignCenter)
            loss_input.setStyleSheet("background-color: white; border: 1px solid black")
            loss_input.setText(str(result[12]))

            purgeStart_input = QtWidgets.QLineEdit()
            purgeStart_input.setFixedHeight(25)
            purgeStart_input.setAlignment(Qt.AlignCenter)
            purgeStart_input.setStyleSheet("background-color: white; border: 1px solid black")

            purgeEnd_input = QtWidgets.QLineEdit()
            purgeEnd_input.setFixedHeight(25)
            purgeEnd_input.setAlignment(Qt.AlignCenter)
            purgeEnd_input.setStyleSheet("background-color: white; border: 1px solid black")

            remarks = QtWidgets.QTextEdit()

            operator_input = QtWidgets.QLineEdit()
            operator_input.setFixedHeight(25)
            operator_input.setAlignment(Qt.AlignCenter)
            operator_input.setStyleSheet("background-color: white; border: 1px solid black")
            operator_input.setText(result[21])

            supervisor_input = QtWidgets.QLineEdit()
            supervisor_input.setFixedHeight(25)
            supervisor_input.setAlignment(Qt.AlignCenter)
            supervisor_input.setStyleSheet("background-color: white; border: 1px solid black")
            supervisor_input.setText(result[22])

            order_number_input = QtWidgets.QLineEdit()
            order_number_input.setFixedHeight(25)
            order_number_input.setEnabled(False)
            order_number_input.setAlignment(Qt.AlignCenter)
            order_number_input.setStyleSheet("background-color: white; border: 1px solid black")
            order_number_input.setText(str(result[7]))

            resin_input = QtWidgets.QLineEdit()
            resin_input.setFixedHeight(25)
            resin_input.setAlignment(Qt.AlignCenter)
            resin_input.setStyleSheet("background-color: white; border: 1px solid black")
            resin_input.setText(result[15])

            purging_input = QtWidgets.QLineEdit()
            purging_input.setFixedHeight(25)
            purging_input.setAlignment(Qt.AlignCenter)
            purging_input.setStyleSheet("background-color: white; border: 1px solid black")
            purging_input.setText(result[14])

            product_input = QtWidgets.QLineEdit()
            product_input.setFixedHeight(25)
            product_input.setAlignment(Qt.AlignCenter)
            product_input.setStyleSheet("background-color: white; border: 1px solid black")
            product_input.textChanged.connect(loss_auto)
            product_input.setText(str(result[29]))

            self.groupBoxRemarks = QtWidgets.QGroupBox(self.entry_widget)
            self.groupBoxRemarks.setGeometry(600, 500, 200, 150)
            self.groupBoxRemarks.setTitle("Remarks")
            self.groupBoxRemarks.show()

            self.remarks_textBox = QtWidgets.QTextEdit(self.groupBoxRemarks)
            self.remarks_textBox.setGeometry(0, 20, 200, 130)
            self.remarks_textBox.setText(result[16])
            self.remarks_textBox.show()

            # Left Side of Vertical Box
            self.left_vbox.addRow(productionID_label, productionID_input)
            self.left_vbox.addRow(productCode_label, productCode_input)
            self.left_vbox.addRow(customer_label, customer_input)
            self.left_vbox.addRow(orderedQuantity_label, orderedQuantity_input)
            self.left_vbox.addRow(lotnumber_label, lot_number_input)
            self.left_vbox.addRow(product_input_label, product_input)
            self.left_vbox.addRow(productOutput_label, product_output_input)
            self.left_vbox.addRow(loss_label, loss_input)
            self.left_vbox.addRow(machine_label, machine_input)
            self.left_vbox.addRow(formulaID_label, self.formulaID_input)
            self.left_vbox.addRow(order_number_lbl, order_number_input)

            # Add widgets to the right Form Box
            self.right_vbox.addRow(feedrate_label, feedRate_input)
            self.right_vbox.addRow(rpm_label, rpm_input)
            self.right_vbox.addRow(screenSize_label, screenSize_input)
            self.right_vbox.addRow(screwConf_label, screwConf_input)
            self.right_vbox.addRow(purging_label, purging_input)
            self.right_vbox.addRow(resin_label, resin_input)
            self.right_vbox.addRow(purgeStart_label, purgeStart_input)
            self.right_vbox.addRow(purgeEnd_label, purgeEnd_input)
            self.right_vbox.addRow(operator_label, operator_input)
            self.right_vbox.addRow(supervisor_label, supervisor_input)

            # Time Table Entry
            time_table = QtWidgets.QTableWidget(self.entry_widget)
            time_table.setGeometry(0, 500, 450, 200)
            time_table.setColumnCount(3)
            time_table.setRowCount(8)
            time_table.setColumnWidth(0, 150)
            time_table.setColumnWidth(1, 150)
            time_table.setStyleSheet("background-color: white;")
            time_table.setFont(QtGui.QFont("Arial", 10))
            time_table.setHorizontalHeaderLabels(["Time Start", "Time End", "Output"])

            # Populate the Table
            for i in range(len(result[9])):
                item = QTableWidgetItem(str(result[9][i]))
                item2 = QTableWidgetItem(str(result[10][i]))
                item3 = QTableWidgetItem(str(result[26][i]))
                time_table.setItem(i, 0, item)
                time_table.setItem(i, 1, item2)
                time_table.setItem(i, 2, item3)

            time_table.show()

            # Temperature Table Entry
            temperature_table = QtWidgets.QTableWidget(self.entry_widget)
            temperature_table.setGeometry(450, 500, 150, 200)
            temperature_table.setColumnCount(1)
            temperature_table.setRowCount(12)
            temperature_table.setStyleSheet("background-color: white;")
            temperature_table.setHorizontalHeaderLabels(["Temperature"])
            temperature_index = ["Z" + str(i + 1) for i in range(12)]  # set the index
            temperature_table.setVerticalHeaderLabels(temperature_index)

            # Populate the table
            for i in range(len(result[24])):
                item = QTableWidgetItem(str(result[24][i]))
                temperature_table.setItem(i, 0, item)

            temperature_table.show()

            save_btn = QtWidgets.QPushButton(self.entry_widget)
            save_btn.setGeometry(540, 705, 60, 25)
            save_btn.clicked.connect(update)
            save_btn.setText("Save")
            save_btn.setCursor(Qt.PointingHandCursor)
            save_btn.show()

            default_date = QtCore.QDateTime(2024, 1, 1, 0, 0)

            time_start_input = QtWidgets.QDateTimeEdit(self.entry_widget)
            time_start_input.setGeometry(30, 475, 120, 25)
            time_start_input.setDisplayFormat("yyyy-MM-dd HH:mm")
            time_start_input.setDateTime(default_date)
            time_start_input.show()

            time_end_input = QtWidgets.QDateTimeEdit(self.entry_widget)
            time_end_input.setGeometry(180, 475, 120, 25)
            time_end_input.setDisplayFormat("yyyy-MM-dd HH:mm")
            time_end_input.setDateTime(default_date)
            time_end_input.show()

            output_lineEdit = QtWidgets.QLineEdit(self.entry_widget)
            output_lineEdit.setGeometry(310, 475, 80, 25)
            output_lineEdit.setAlignment(Qt.AlignCenter)
            output_lineEdit.setStyleSheet("background-color: white; border: 1px solid black")
            output_lineEdit.show()

            self.plus_icon = ClickableLabel(self.entry_widget)
            self.plus_icon.setGeometry(390, 475, 25, 25)
            self.plus_icon.setPixmap(QtGui.QIcon('plus.png').pixmap(25, 25))
            self.plus_icon.setCursor(Qt.PointingHandCursor)
            self.plus_icon.clicked.connect(add_time)
            self.plus_icon.show()

            self.reset_icon = ClickableLabel(self.entry_widget)
            self.reset_icon.setGeometry(425, 475, 25, 25)
            self.reset_icon.setPixmap(QtGui.QIcon('reset.png').pixmap(20, 20))
            self.reset_icon.setCursor(Qt.PointingHandCursor)
            self.reset_icon.clicked.connect(reset_table)
            self.reset_icon.show()

        def print_file():

            try:
                from openpyxl.styles import Font
                from openpyxl.styles import Alignment

                selected = self.extruder_table.selectedItems()

                process_id = selected[0].text()

                self.cursor.execute(f"""
                            SELECT * FROM extruder 
                            WHERE process_id = '{process_id}';

                            """)
                items = self.cursor.fetchall()[0]
                # Unpack the Items
                machine_number = items[1]
                quantity_order = items[2]
                customer = items[4]
                code = items[6]
                total_input = items[29]
                total_time = items[8]
                time_start = items[9]
                time_end = items[10]
                outputPerHour = items[27]
                total_output = items[3]
                outputPercent = items[11]
                loss = items[12]
                lossPercent = items[13]
                purging = items[14]
                resin = items[15]
                remarks = items[16]
                operator = items[21]
                supervisor = items[22]
                outputs = items[26]
                materials = items[23]
                lot_number = items[30]
                print(lot_number)

                wb = load_workbook(r"\\mbpi-server-01\IT\AMIEL\Extruder System\dist\Extruder Template.xlsx")
                worksheet = wb.active

                font = Font(size=8, bold=True, name='Arial')
                center_Alignment = Alignment(horizontal='center', vertical='center')

                worksheet["F5"] = "Extruder Machine No. " + machine_number[-1]
                worksheet["A8"] = machine_number[-1]
                worksheet["B8"] = quantity_order  # quantity order
                worksheet["C8"].font = font
                worksheet["C8"].alignment = center_Alignment
                worksheet["C8"] = customer  # customer
                worksheet["F8"] = code  # product code
                worksheet["G9"] = total_input  # total input
                worksheet["H9"] = total_time  # total time used
                worksheet["I9"] = outputPerHour  # output Per Hour
                worksheet["K9"] = total_output  # total Output
                worksheet["L9"] = outputPercent  # Total Output Percentage
                worksheet["M9"] = loss
                worksheet["N9"] = lossPercent

                total_sec = timedelta()
                for row in range(len(time_start)):
                    worksheet["A" + str(12 + row)] = time_start[row].strftime("%d-%b-%Y %H:%M")
                    worksheet["D" + str(12 + row)] = time_end[row].strftime("%d-%b-%Y %H:%M")
                    worksheet["F" + str(12 + row)] = time_end[row] - time_start[row]
                    worksheet["G" + str(12 + row)] = outputs[row]
                    total_sec = total_sec + (time_end[row] - time_start[row])

                try:
                    hour = str(int(total_sec.total_seconds() // 3600))
                    minute = str((int(total_sec.total_seconds() % 3600) // 60))
                    total_time_used = time(int(hour), int(minute))

                    worksheet["F25"] = total_time_used
                except ValueError:
                    worksheet["F25"] = hour + ":" + minute

                for key in list(materials.keys()):
                    worksheet["I" + str(12 + list(materials.keys()).index(key))] = key
                    worksheet["K" + str(12 + list(materials.keys()).index(key))] = materials[key]

                for ln in range(len(lot_number)):
                    worksheet["M" + str(12 + ln)] = lot_number[ln]

                worksheet["B27"] = purging
                worksheet["B29"] = resin
                worksheet["G26"] = remarks
                worksheet["M28"] = operator
                worksheet["M29"] = supervisor

                wb.save(r"\\mbpi-server-01\IT\AMIEL\Extruder System\dist\text.xlsx")
                QMessageBox.information(self.production_widget, "Export Success", "File Successfully Exported!")
            except Exception as e:
                print(e)
                QMessageBox.information(self.production_widget, "ERROR", "No Selected Items")

        def filter_table():

            queryConList = []

            if self.machine_combo.currentText() != "-":
                queryConList.append(f"machine = '{self.machine_combo.currentText()}'")
            if self.company_combo.currentText() != "-":
                queryConList.append(f"customer = '{self.company_combo.currentText().replace("'", "''")}'")
            if self.formula_combo.currentText() != "":
                queryConList.append(f"formula_id = '{self.formula_combo.currentText()}'")
            if self.productCode_combo.currentText() != "":
                queryConList.append(f"product_code = '{self.productCode_combo.currentText()}'")

            query = f"""
                    SELECT 
                    process_id, machine, customer, qty_order, total_output, formula_id, product_code, total_time
                    FROM extruder WHERE
                                    """

            if len(queryConList) == 0:
                query = query.replace("WHERE", "")

            for condition in queryConList:
                if condition != queryConList[-1]:
                    query += condition + " AND "
                else:
                    query += condition

            query += "ORDER BY process_id"

            self.cursor.execute(query)
            result = self.cursor.fetchall()

            self.extruder_table.clearContents()
            for i in range(len(result)):
                for j in range(len(column_names)):
                    item = QtWidgets.QTableWidgetItem(str(result[i][j]))  # Convert to string
                    # Set Alignment for specific columns
                    if j == 2 or j == 6 or j == 3 or j == 4 or j == 7:
                        item.setTextAlignment(Qt.AlignCenter)
                    else:
                        pass
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)  # Make the cells unable to be edited
                    self.extruder_table.setItem(i, j, item)

            self.extruder_table.show()
            pass

        def show_tables():

            try:
                main_time_table.clear()
                material_table.clear()
                lotNumber_table.clear()
            except Exception as e:
                print(e)
            selected = self.extruder_table.selectedItems()
            if selected:
                items = [item.text() for item in selected]
                process_id = items[0]

            try:
                self.cursor.execute(f"""
                            SELECT time_start, time_end, outputs, materials, lot_number
                            FROM extruder
                            WHERE process_id = {process_id};

                            """)
                result = self.cursor.fetchall()
            except:
                QMessageBox.information(self.production_widget, "ERROR", "No Selected Item")
                self.extruder_table.itemSelectionChanged.connect(show_tables)
                return

            t_start = result[0][0]
            t_end = result[0][1]
            outputs = result[0][2]
            materials = result[0][3]
            lotNumber = result[0][4]


            # Set the Header Labels again
            main_time_table.setHorizontalHeaderLabels(["Time Start", "Time End", "Output"])
            material_table.setHorizontalHeaderLabels(["Materials", "Quantity(kg)"])
            lotNumber_table.setHorizontalHeaderLabels(["Lot Number"])

            # Populating Time Table
            for i in range(len(t_start)):
                item1 = QTableWidgetItem(str(t_start[i].strftime("%d-%b-%Y %H:%M")))
                item1.setTextAlignment(Qt.AlignCenter)
                item1.setFlags(item1.flags() & ~Qt.ItemIsEditable)
                item2 = QTableWidgetItem(str(t_end[i].strftime("%d-%b-%Y %H:%M")))
                item2.setTextAlignment(Qt.AlignCenter)
                item2.setFlags(item2.flags() & ~Qt.ItemIsEditable)
                item3 = QTableWidgetItem(str(outputs[i]))
                item3.setTextAlignment(Qt.AlignCenter)
                item3.setFlags(item3.flags() & ~Qt.ItemIsEditable)
                main_time_table.setItem(i, 0, item1)
                main_time_table.setItem(i, 1, item2)
                main_time_table.setItem(i, 2, item3)



            if len(materials) > material_table.rowCount():
                material_table.setRowCount(len(materials))

            for i in list(materials.keys()):
                key = QTableWidgetItem(str(i))
                key.setFlags(key.flags() & ~Qt.ItemIsEditable)
                value = QTableWidgetItem(str(round((materials[i]), 4)))
                value.setFlags(value.flags() & ~Qt.ItemIsEditable)
                material_table.setItem(list(materials.keys()).index(i), 0, key)
                material_table.setItem(list(materials.keys()).index(i), 1, value)


            for i in range(len(lotNumber)):
                item = QTableWidgetItem(str(lotNumber[i]))
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                lotNumber_table.setItem(i, 0, item)
        
        self.production_widget = QtWidgets.QWidget(self.main_widget)
        self.production_widget.setGeometry(0, 0, 991, 751)
        self.production_widget.setStyleSheet("background-color: rgb(240,240,240);")
        self.production_widget.show()
        
        self.extruder_table = QtWidgets.QTableWidget(self.production_widget)
        self.extruder_table.setGeometry(QtCore.QRect(20, 80, 900, 375))
        self.extruder_table.verticalHeader().setVisible(False)
        self.extruder_table.setSortingEnabled(True)

        self.cursor.execute("""
        SELECT column_name FROM information_schema.columns
        WHERE TABLE_NAME = 'extruder';
        """)
        # column_names = self.cursor.fetchall()
        # column_names = [i[0] for i in column_names]

        column_names = ["process_id", "machine", "customer", "qty_order", "total_output", "formula_id", "product_code",
                        "total time(hr)"]

        try:
            self.cursor.execute("""SELECT 
                        process_id, machine, customer, qty_order, total_output, formula_id, product_code, total_time
                        FROM extruder
                        ORDER BY process_id DESC;
                        """)
            result = self.cursor.fetchall()
        except Exception as e:
            self.cursor.execute("""
                        SELECT 
                        process_id, machine, customer, qty_order, total_output, formula_id, product_code, total_time
                        FROM extruder
                        ORDER BY process_id DESC
                        ; 

                        """)
            result = self.cursor.fetchall()

        # Set Column Count
        self.extruder_table.setColumnCount(len(column_names))
        # Set Row Count
        self.extruder_table.setRowCount(len(result))

        self.extruder_table.setStyleSheet("""
        gridline-color: black; 
        color : black;
        """)

        # Populate table with data
        for i in range(len(result)):
            for j in range(len(column_names)):
                item = QtWidgets.QTableWidgetItem(str(result[i][j]))  # Convert to string
                # Set Alignment for specific columns
                if j == 2 or j == 6 or j == 3 or j == 4 or j == 7 or j == 5:
                    item.setTextAlignment(Qt.AlignCenter)
                else:
                    pass
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)  # Make the cells unable to be edited
                self.extruder_table.setItem(i, j, item)

        bold_font = QtGui.QFont()
        bold_font.setBold(True)
        self.extruder_table.horizontalHeader().setFont(bold_font)
        self.extruder_table.horizontalHeader().setStyleSheet("""
        QHeaderView::section{
        font-weight: bold;
        background-color: black;
        color: white;
        }

        """)

        # Set Column Width
        self.extruder_table.setColumnWidth(2, 198)

        self.extruder_table.setHorizontalHeaderLabels([col.upper() for col in column_names])  # Set column names
        # Set selection mode to select entire rows and disable single item selection
        self.extruder_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.extruder_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.extruder_table.itemSelectionChanged.connect(show_tables)
        self.extruder_table.show()

        #Filters
        self.machine_combo = QtWidgets.QComboBox(self.production_widget)
        self.machine_combo.setGeometry(120, 45, 100, 25)
        self.cursor.execute("""
                    SELECT DISTINCT(machine) FROM extruder;
                """)
        machine = self.cursor.fetchall()
        self.machine_combo.addItem("-")
        for i in machine:
            self.machine_combo.addItem(i[0])
        self.machine_combo.currentIndexChanged.connect(filter_table)
        self.machine_combo.show()

        self.company_combo = QtWidgets.QComboBox(self.production_widget)
        self.company_combo.setGeometry(220, 45, 200, 25)
        self.cursor.execute("""
            SELECT DISTINCT(customer) FROM extruder;
        """)
        customers = self.cursor.fetchall()
        self.company_combo.addItem("-")
        for i in customers:
            self.company_combo.addItem(i[0])
        self.company_combo.setEditable(True)
        self.company_combo.currentIndexChanged.connect(filter_table)
        self.company_combo.show()

        self.formula_combo = QtWidgets.QComboBox(self.production_widget)
        self.formula_combo.setGeometry(620, 45, 100, 25)
        self.formula_combo.setEditable(True)
        self.formula_combo.currentTextChanged.connect(filter_table)
        self.formula_combo.show()

        self.productCode_combo = QtWidgets.QComboBox(self.production_widget)
        self.productCode_combo.setGeometry(720, 45, 100, 25)
        self.productCode_combo.currentTextChanged.connect(filter_table)
        self.productCode_combo.setEditable(True)
        self.productCode_combo.show()

        self.view_btn = QtWidgets.QPushButton(self.production_widget)
        self.view_btn.setGeometry(600, 700, 80, 30)
        self.view_btn.setText("View")
        self.view_btn.setStyleSheet("background-color : rgb(240,240,240);")
        self.view_btn.clicked.connect(show_form)
        self.view_btn.setCursor(Qt.PointingHandCursor)
        self.view_btn.show()

        self.add_btn = QtWidgets.QPushButton(self.production_widget)
        self.add_btn.setGeometry(681, 700, 80, 30)
        self.add_btn.setText("Add Entry")
        self.add_btn.setStyleSheet("background-color : rgb(240,240,240);")
        self.add_btn.clicked.connect(add_entry)
        self.add_btn.setCursor(Qt.PointingHandCursor)
        self.add_btn.show()

        self.update_btn = QtWidgets.QPushButton(self.production_widget)
        self.update_btn.setGeometry(762, 700, 80, 30)
        self.update_btn.setText("Update")
        self.update_btn.setStyleSheet("background-color : rgb(240,240,240);")
        self.update_btn.clicked.connect(update_entry)
        self.update_btn.setCursor(Qt.PointingHandCursor)
        self.update_btn.show()

        self.print_btn = QtWidgets.QPushButton(self.production_widget)
        self.print_btn.setGeometry(843, 700, 80, 30)
        self.print_btn.setText("Print")
        self.print_btn.setStyleSheet("background-color: rgb(240,240,240);")
        self.print_btn.clicked.connect(print_file)
        self.print_btn.setCursor(Qt.PointingHandCursor)
        self.print_btn.show()

        main_time_table = QtWidgets.QTableWidget(self.production_widget)
        main_time_table.setGeometry(20, 455, 475, 225)
        main_time_table.setColumnCount(3)
        main_time_table.setRowCount(6)
        main_time_table.setHorizontalHeaderLabels(["Time Start", "Time End", "Output"])
        main_time_table.setColumnWidth(0, 180)
        main_time_table.setColumnWidth(1, 180)
        main_time_table.setColumnWidth(2, 98)
        main_time_table.show()

        material_table = QtWidgets.QTableWidget(self.production_widget)
        material_table.setGeometry(495, 455, 263, 225)
        material_table.setColumnCount(2)
        material_table.setRowCount(13)
        material_table.setHorizontalHeaderLabels(["Material", "Value(Kg)"])
        material_table.setColumnWidth(0, 130)
        material_table.setColumnWidth(1, 90)
        material_table.show()

        lotNumber_table = QtWidgets.QTableWidget(self.production_widget)
        lotNumber_table.setGeometry(758, 455, 162, 225)
        lotNumber_table.setColumnCount(1)
        lotNumber_table.setRowCount(6)
        lotNumber_table.setHorizontalHeaderLabels(["Lot Number"])
        lotNumber_table.setColumnWidth(0, 162)
        lotNumber_table.show()

    def quality_control(self):
        try:
            self.production_widget.deleteLater()
            body_widget.deleteLater()

        except Exception as e:
            print(e)

        def exportBtn_clicked():
            date_from = date1.text()
            date_to = date2.text()
            self.cursor.execute(f"""
            SELECT * FROM quality_control
            WHERE evaluated_on::DATE BETWEEN '{date_from}' AND '{date_to}'
            ORDER BY id 
            
            """)

            result = self.cursor.fetchall()

            df = pd.DataFrame(result)

            # Get the Column Names
            self.cursor.execute("""
            SELECT column_name FROM information_schema.columns
            WHERE TABLE_NAME = 'quality_control'
            
            """)
            result = self.cursor.fetchall()
            column_names = ['id', 'lot_number', 'product_code', 'customer', 'status', 'remarks', 'action',
                            'original_lot', 'evaluated_by', 'evaluated_on', 'encoded_on', 'updated_by',
                            'updated_on', 'time_endorsed', 'edited', 'qc_type', 'formula_id']

            try:
                df.to_excel(excel_writer=r'\\mbpi-server-01\IT\QC Data Imports\blank.xlsx', index=False,
                            header=column_names)
                QMessageBox.information(self.qc_widget, "File Imported", "Successfully Imported Data")
            except PermissionError:
                QMessageBox.critical(self.qc_widget, "Permission Error", "Unable to Export the File. \n "
                                                                         "Someone is using blank.xlsx")

        def edited_checkbox_changed():
            # Select Only Edited Data
            if edited_checkbox.isChecked():
                self.qc_table.itemSelectionChanged.disconnect(show_items)
                self.qc_table.clearSelection()
                self.qc_table.clear()
                self.qc_table.setHorizontalHeaderLabels(['ID', 'Lot Number', 'Customer','Product Code',
                                                         'Status', 'Remarks', 'Action Taken'])
                self.cursor.execute("""
                SELECT id, lot_number, customer, product_code, status, remarks, action
                FROM quality_control
                WHERE edited = 'True'
                
                """)

                result = self.cursor.fetchall()

                # Populate the table
                for i in range(len(result)):
                    for j in range(len(result[i])):
                        item = QTableWidgetItem(str(result[i][j]))
                        item.setTextAlignment(Qt.AlignCenter)
                        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                        self.qc_table.setItem(i, j, item)
                self.qc_table.itemSelectionChanged.connect(show_items)
                self.qc_table.show()
            else:
                # Select ALL
                self.qc_table.itemSelectionChanged.disconnect(show_items)
                self.qc_table.clearSelection()
                self.qc_table.clear()
                self.qc_table.setHorizontalHeaderLabels(['ID', 'Lot Number', 'Customer', 'Product Code',
                                                         'Status', 'Remarks', 'Action Taken'])
                self.cursor.execute("""
                                SELECT id, lot_number, customer, product_code, status, remarks, action 
                                FROM quality_control

                                """)

                result = self.cursor.fetchall()

                # Populate the table
                for i in range(len(result)):
                    for j in range(len(result[i])):
                        item = QTableWidgetItem(str(result[i][j]))
                        item.setTextAlignment(Qt.AlignCenter)
                        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                        self.qc_table.setItem(i, j, item)

                self.qc_table.itemSelectionChanged.connect(show_items)
                self.qc_table.show()

        def lotNumber_search():
            if search_bar.text() != '':
                self.qc_table.itemSelectionChanged.disconnect(show_items)
                self.qc_table.clearSelection()
                self.qc_table.clear()
                self.qc_table.setHorizontalHeaderLabels(['ID', 'Lot Number', 'Customer', 'Product Code',
                                                         'Status', 'Remarks', 'Action Taken'])

                self.cursor.execute(f"""
                            SELECT id, lot_number, customer, product_code, status, remarks, action
                            FROM quality_control
                            WHERE lot_number ILIKE '%{search_bar.text()}%'

                            """)
                result = self.cursor.fetchall()

                # Populate the table
                for i in range(len(result)):
                    for j in range(len(result[i])):
                        item = QTableWidgetItem(str(result[i][j]))
                        item.setTextAlignment(Qt.AlignCenter)
                        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                        self.qc_table.setItem(i, j, item)

                self.qc_table.itemSelectionChanged.connect(show_items)
                self.qc_table.show()

        def evaluation_entry():

            # Getting every single Lot Number in a Multiple Lot Number
            new_lot_list = []
            old_lot_list = []

            # Getting the new Lot List
            def multiple_lotNumber():
                try:
                    new_lot_list.clear()
                    lotNumbers_board.clear()
                    if len(lotNumber_input.text().split("-")) == 2:
                        start_lot = lotNumber_input.text().split("-")[0][:4]
                        end_lot = lotNumber_input.text().split("-")[1][:4]
                        string_code = lotNumber_input.text().split("-")[0][4:6]

                        for i in range(int(start_lot), int(end_lot) + 1):
                            while len(str(i)) != 4:
                                i = '0' + str(i)
                            new_lot_list.append(str(i) + string_code)

                        lotNumbers_board.setText("\n".join(new_lot_list))
                    else:
                        lotNumbers_board.clear()
                except:
                    print("INVALID")

           # For getting multiple old Lot Numbers from correction input
            def get_old_lotNumbers():

                try:
                    if len(correction_input.text().split("-")) == 2:
                        start_lot = correction_input.text().split("-")[0][:4]
                        end_lot = correction_input.text().split("-")[1][:4]
                        string_code = correction_input.text().split("-")[0][4:6]

                        for i in range(int(start_lot), int(end_lot) + 1):
                            old_lot_list.append(str(i) + string_code)


                except Exception as e:
                    print(e)

            def saveBtn_clicked():

                def clear_entries():
                    lotNumber_input.clear()
                    productCode_dropdown.setCurrentIndex(0)
                    customer_dropdown.setCurrentIndex(0)
                    formulaID_input.clear()
                    evaluatedBy_dropdown.setCurrentIndex(0)
                    result_dropdown.setCurrentIndex(0)
                    remarks_box.clear()

                try:
                    # For saving Multiple Lot Number in quality_control_tbl2

                    self.cursor.execute("SELECT MAX(id) FROM quality_control")
                    qc_ID = self.cursor.fetchone()[0]


                    if qcType_dropdown.currentText() == "NEW":

                        self.cursor.execute(f"""
                        SELECT lot_number 
                        FROM quality_control
                        WHERE lot_number = '{lotNumber_input.text()}'
                        
                        """)
                        result = self.cursor.fetchone()


                        # This if statement break the cancel the saving if the lot number is already in the Database
                        if lotNumber_input.text() in result:
                            QMessageBox.information(self.qc_widget, "Data Exist", "Data is already Entered.")
                            return


                        self.cursor.execute(f"""
                                           INSERT INTO quality_control
                                           (lot_number, product_code, customer, status, remarks, action, original_lot, evaluated_by,
                                           evaluated_on, encoded_on, time_endorsed, qc_type, formula_id)
                                           VALUES('{lotNumber_input.text().strip()}', '{productCode_dropdown.currentText().strip()}', '{customer_dropdown.currentText()}',
                                           '{result_dropdown.currentText()}', '{remarks_box.toPlainText()}', '{actionTake_box.toPlainText()}',
                                           '{lotNumber_input.text()}', '{evaluatedBy_dropdown.currentText()}', '{date_started_input.text()}', '{datetime.now().strftime("%Y-%m-%d %H:%M")}',
                                           '{time_endorsed_input.text()}', '{qcType_dropdown.currentText()}', '{formulaID_input.text()}' )
                                           """)
                        self.conn.commit()

                        # For saving Multiple Lot Number in quality_control_tbl2

                        self.cursor.execute("SELECT MAX(id) FROM quality_control")
                        qc_ID = self.cursor.fetchone()[0]

                        if "-" in lotNumber_input.text():
                            for lot in new_lot_list:
                                self.cursor.execute(f"""
                                        INSERT INTO  quality_control_tbl2(id, lot_number, evaluation_date, original_lot,
                                         status, product_code, qc_type, formula_id)
                                        VALUES('{qc_ID}','{lot.strip()}', '{date_started_input.text()}', '{lot.strip()}',
                                        '{result_dropdown.currentText()}', '{productCode_dropdown.currentText().strip()}', 
                                        '{qcType_dropdown.currentText()}', '{formulaID_input.text()}' )

                                                    """)
                                self.conn.commit()
                            print("query successful")
                            clear_entries() # Clear the entries after successful entry
                        else:
                            self.cursor.execute(f"""
                                    INSERT INTO quality_control_tbl2(id, lot_number, evaluation_date, original_lot, status, 
                                    product_code, qc_type, formula_id)
                                    VALUES('{qc_ID}', '{lotNumber_input.text().strip()}', '{date_started_input.text()}', '{lotNumber_input.text()}',
                                    '{result_dropdown.currentText()}', '{productCode_dropdown.currentText().strip()}', '{qcType_dropdown.currentText()}', 
                                    '{formulaID_input.text()}')    """)
                            self.conn.commit()
                            clear_entries()

                        QMessageBox.information(self.body_widget.setStyleSheet("border: none;"), "Query Success", "QC Entry Added")
                        new_lot_list.clear()

                    else: # CORRECTION

                        # SAVE TO THE FIRST QC TABLE 1
                        self.cursor.execute(f""" SELECT original_lot FROM quality_control
                                            WHERE original_lot = '{correction_input.text().strip()}';
                                       """)

                        result = self.cursor.fetchall()
                        try:
                            orig_lot = result[0][0]
                        except:
                            orig_lot = correction_input.text().strip()


                        self.cursor.execute(f"""
                                       INSERT INTO quality_control
                                           (lot_number, product_code, customer, status, remarks, action, original_lot, evaluated_by,
                                           evaluated_on, encoded_on, time_endorsed, qc_type, updated_by, updated_on, formula_id)
                                           VALUES('{lotNumber_input.text().strip()}', '{productCode_dropdown.currentText().strip()}', '{customer_dropdown.currentText()}',
                                           '{result_dropdown.currentText()}', '{remarks_box.toPlainText()}', '{actionTake_box.toPlainText()}',
                                           '{orig_lot.strip()}', '{evaluatedBy_dropdown.currentText()}', '{date_started_input.text()}', '{datetime.now().strftime("%Y-%m-%d %H:%M")}',
                                          ' {time_endorsed_input.text()}', '{qcType_dropdown.currentText()}', 
                                          '{updatedBy_input.currentText()}', '{datetime.now().strftime("%Y-%m-%d %H:%M")}',
                                          '{formulaID_input.text()}')

                                       """)

                        self.conn.commit()

                        # For saving Multiple Lot Number in quality_control_tbl2

                        self.cursor.execute("SELECT MAX(id) FROM quality_control")
                        qc_ID = self.cursor.fetchone()[0]

                        # Save To quality_control_tbl2 DB

                        if "-" in lotNumber_input.text():
                            if len(old_lot_list) == len(new_lot_list):
                                pass
                            else:
                                QMessageBox.critical(self.body_widget.setStyleSheet("border: none;"),
                                                        "ERROR",
                                                        "CORRECTION AND LOT NUMBER SHOULD BE EQUAL RANGE")
                                return

                            for i in range(len(old_lot_list)):
                                print("Run " + str(i))
                                self.cursor.execute(f"""
                                SELECT original_lot FROM quality_control_tbl2
                                WHERE lot_number = '{old_lot_list[i]}'

                                """)

                                result = self.cursor.fetchall()
                                orig_lot = result[0][0]

                                self.cursor.execute(f"""
                                                        INSERT INTO quality_control_tbl2(id, lot_number, evaluation_date, original_lot, status,
                                                        product_code, qc_type, formula_id)
                                                        VALUES({qc_ID}, '{new_lot_list[i].strip()}', '{date_started_input.text()}', '{orig_lot.strip()}',
                                                        '{result_dropdown.currentText()}', '{productCode_dropdown.currentText().strip()}',
                                                         '{qcType_dropdown.currentText()}', '{formulaID_input.text()}')

                                                        """)
                                self.conn.commit()
                            clear_entries()

                        else:
                            # Getting the original Lot
                            self.cursor.execute(f"""
                                                            SELECT original_lot FROM quality_control_tbl2
                                                            WHERE lot_number = '{correction_input.text()}'

                                                            """)

                            result = self.cursor.fetchall()

                            orig_lot = result[0][0]

                            self.cursor.execute(f"""
                                    INSERT INTO quality_control_tbl2(id, lot_number, evaluation_date, original_lot, status,
                                    product_code, qc_type, formula_id)
                                    VALUES('{qc_ID}, {lotNumber_input.text().strip()}', '{date_started_input.text()}', '{orig_lot.strip()}',
                                    '{result_dropdown.currentText()}', '{productCode_dropdown.currentText().strip()}',
                                    '{qcType_dropdown.currentText()}', '{formulaID_input.text()}')
                                                """)
                            self.conn.commit()
                        QMessageBox.information(self.body_widget.setStyleSheet("border: none;"), "Query Success",
                                                "QC Entry Added")
                        new_lot_list.clear()
                        clear_entries()

                except Exception as e:
                    print(e)
                    QMessageBox.critical(self.body_widget, "ERROR", "test")
                    self.conn.rollback()

            def correction_enabled():
                if qcType_dropdown.currentText() == "CORRECTION":
                    updatedBy_input.setEnabled(True)
                    correction_input.setEnabled(True)

                    updatedBy_input.setStyleSheet(
                        "background-color: rgb(255, 255, 0); border: 1px solid rgb(171, 173, 179);")
                    correction_input.setStyleSheet(
                        "background-color: rgb(255, 255, 0); border: 1px solid rgb(171, 173, 179);")
                else:
                    updatedBy_input.setEnabled(False)
                    correction_input.setEnabled(False)

                    correction_input.setStyleSheet(
                        "background-color: rgb(240, 240, 240); border: 1px solid rgb(171, 173, 179);")
                    updatedBy_input.setStyleSheet(
                        "background-color: rgb(240, 240, 240); border: 1px solid rgb(171, 173, 179);")
                    correction_input.clear()
                    updatedBy_input.clear()

            def correction_auto_input():

                self.cursor.execute(f"""
                SELECT product_code, customer, evaluated_by FROM quality_control
                WHERE original_lot = '{correction_input.text()}'
                
                
                """)
                result = self.cursor.fetchone()
                if result == None:
                    print("no result")
                else:
                    customer_dropdown.setCurrentText(result[1])
                    productCode_dropdown.setCurrentText(result[0])
                    evaluatedBy_dropdown.setCurrentText(result[2])

            self.qc_widget.deleteLater()

            self.qc_widget = QtWidgets.QWidget(self.main_widget)
            self.qc_widget.setGeometry(0, 0, 991, 751)
            self.qc_widget.setStyleSheet("background-color: rgb(240,240,240);")
            self.qc_widget.show()

            self.qcBtn_topBorder = QtWidgets.QWidget(self.qc_widget)
            self.qcBtn_topBorder.setGeometry(0, 0, 991, 30)
            self.qcBtn_topBorder.show()

            self.qc_TableBtn = QtWidgets.QPushButton(self.qcBtn_topBorder)
            self.qc_TableBtn.setGeometry(0, 0, 150, 30)
            self.qc_TableBtn.setText("Evaluated Products")
            self.qc_TableBtn.setCursor(Qt.PointingHandCursor)
            self.qc_TableBtn.setFont(QtGui.QFont("Arial", 11))
            self.qc_TableBtn.setStyleSheet("border: 1px solid rgb(160, 160, 160); text-align: top;")
            self.qc_TableBtn.clicked.connect(self.quality_control)
            self.qc_TableBtn.setShortcut("F1")
            self.qc_TableBtn.show()

            self.qc_addEntryBtn = QtWidgets.QPushButton(self.qcBtn_topBorder)
            self.qc_addEntryBtn.setGeometry(150, 0, 150, 30)
            self.qc_addEntryBtn.setText("Evaluation Entry")
            self.qc_addEntryBtn.setCursor(Qt.PointingHandCursor)
            self.qc_addEntryBtn.setStyleSheet("border: 1px solid rgb(160, 160, 160); color: rgb(0,109,189);")
            self.qc_addEntryBtn.clicked.connect(evaluation_entry)
            self.qc_addEntryBtn.setFont(QtGui.QFont("Arial", 11))
            self.qc_addEntryBtn.setShortcut("F2")
            self.qc_addEntryBtn.show()

            self.qc_dataBtn = QtWidgets.QPushButton(self.qcBtn_topBorder)
            self.qc_dataBtn.setGeometry(300, 0, 150, 30)
            self.qc_dataBtn.setText("QC Data")
            self.qc_dataBtn.setCursor(Qt.PointingHandCursor)
            self.qc_dataBtn.setStyleSheet("border: 1px solid rgb(160, 160, 160); text-align: top;")
            self.qc_dataBtn.clicked.connect(show_qc_data)
            self.qc_dataBtn.setFont(QtGui.QFont("Arial", 11))
            self.qc_dataBtn.setShortcut("F3")
            self.qc_dataBtn.show()

            self.dashboardBtn = QtWidgets.QPushButton(self.qcBtn_topBorder)
            self.dashboardBtn.setGeometry(450, 0, 150, 30)
            self.dashboardBtn.setText("Dashboard")
            self.dashboardBtn.setCursor(Qt.PointingHandCursor)
            self.dashboardBtn.setStyleSheet("border: 1px solid rgb(160, 160, 160); text-align: top;")
            self.dashboardBtn.clicked.connect(show_dashboards)
            self.dashboardBtn.setFont(QtGui.QFont("Arial", 11))
            self.dashboardBtn.setShortcut("F4")
            self.dashboardBtn.show()

            self.body_widget = QtWidgets.QWidget(self.qc_widget)
            self.body_widget.setGeometry(0, 30, 991, 721)
            self.body_widget.setStyleSheet(
                "background-color: rgb(239, 243, 254); border-top : 1px solid rgb(160, 160, 160);")
            self.body_widget.show()

            self.qc_returns = QtWidgets.QPushButton(self.qcBtn_topBorder)
            self.qc_returns.setGeometry(600, 0, 150, 30)
            self.qc_returns.setText("Returns")
            self.qc_returns.setCursor(Qt.PointingHandCursor)
            self.qc_returns.setStyleSheet("border: 1px solid rgb(160, 160, 160);")
            self.qc_returns.setFont(QtGui.QFont("Arial", 11))
            self.qc_returns.clicked.connect(qc_returns)
            self.qc_returns.setShortcut("F5")
            self.qc_returns.show()

            label1 = QtWidgets.QLabel(self.body_widget)
            label1.setGeometry(0, 0, 300, 15)
            label1.setText("QC EVALUATION AND RESULT")
            label1.setStyleSheet("border-top: none")
            label1.setFont(QtGui.QFont("Arial", 9))
            label1.show()

            label2 = QtWidgets.QLabel(self.body_widget)
            label2.setGeometry(0, 15, 250, 13)
            label2.setText("USER CAN VIEW, ADD AND UPDATE QC RECORDS")
            label2.setFont(QtGui.QFont("Arial", 8))
            label2.setStyleSheet("border-top: none")
            label2.show()

            widget1 = QtWidgets.QWidget(self.body_widget)
            widget1.setGeometry(0, 33, 440, 395)
            widget1.setStyleSheet("background-color: rgb(239, 243, 254); border: none;")

            font = QtGui.QFont("Arial", 10)

            # Create widgets to put in widget1
            qcType_label = QLabel()
            qcType_label.setText("QC TYPE")
            qcType_label.setFixedWidth(110)
            qcType_label.setFixedHeight(20)
            qcType_label.setFont(font)
            qcType_label.setStyleSheet("border: none;")

            qcControl_label = QLabel()
            qcControl_label.setText("QC Control")
            qcControl_label.setFont(font)
            qcControl_label.setStyleSheet("border: none;")

            customer_label = QLabel()
            customer_label.setText("CUSTOMER")
            customer_label.setFixedWidth(110)
            customer_label.setFixedHeight(20)
            customer_label.setFont(font)
            customer_label.setStyleSheet("border: none;")

            productCode_label = QLabel()
            productCode_label.setText("PRODUCT CODE")
            productCode_label.setFixedWidth(110)
            productCode_label.setFixedHeight(20)
            productCode_label.setFont(font)
            productCode_label.setStyleSheet("border: none;")

            evaluatedBy_label = QLabel()
            evaluatedBy_label.setText("EVALUATED BY")
            evaluatedBy_label.setFixedWidth(110)
            evaluatedBy_label.setFixedHeight(20)
            evaluatedBy_label.setFont(font)
            evaluatedBy_label.setStyleSheet("border: none;")

            date_started_label = QLabel()
            date_started_label.setText("DATE STARTED")
            date_started_label.setFixedWidth(110)
            date_started_label.setFixedHeight(20)
            date_started_label.setFont(font)
            date_started_label.setStyleSheet("border: none;")

            lotNumber_label = QLabel()
            lotNumber_label.setText("LOT NUMBER")
            lotNumber_label.setFixedWidth(110)
            lotNumber_label.setFixedHeight(20)
            lotNumber_label.setFont(font)
            lotNumber_label.setStyleSheet("border: none;")

            time_started_label = QLabel()
            time_started_label.setText("TIME STARTED")
            time_started_label.setFixedWidth(110)
            time_started_label.setFixedHeight(20)
            time_started_label.setFont(font)
            time_started_label.setStyleSheet("border: none;")

            time_endorsed_label = QLabel()
            time_endorsed_label.setText("TIME ENDORSED")
            time_endorsed_label.setFixedWidth(110)
            time_endorsed_label.setFixedHeight(20)
            time_endorsed_label.setFont(font)
            time_endorsed_label.setStyleSheet("border: none;")

            result_label = QLabel()
            result_label.setText("TEST RESULT")
            result_label.setFixedWidth(110)
            result_label.setFixedHeight(20)
            result_label.setFont(font)
            result_label.setStyleSheet("border: none;")

            updatedBy_label = QLabel()
            updatedBy_label.setText("UPDATED BY")
            updatedBy_label.setFixedWidth(110)
            updatedBy_label.setFixedHeight(20)
            updatedBy_label.setFont(font)
            updatedBy_label.setStyleSheet("border: none;")

            # Right Side Widgets
            qcType_dropdown = QtWidgets.QComboBox()
            qcType_dropdown.setStyleSheet("border: 1px solid rgb(171, 173, 179); background-color: yellow;")
            qcType_dropdown.addItem("NEW")
            qcType_dropdown.addItem("CORRECTION")
            qcType_dropdown.setFixedHeight(25)
            qcType_dropdown.setFixedWidth(296)
            qcType_dropdown.currentTextChanged.connect(correction_enabled)

            qcControl_input = QtWidgets.QLineEdit()
            qcControl_input.setStyleSheet("background-color: rgb(238, 238, 238); border: 1px solid rgb(171, 173, 179);")
            qcControl_input.setFixedHeight(35)
            qcControl_input.setEnabled(False)
            qcControl_input.setFixedWidth(296)

            customer_dropdown = QtWidgets.QComboBox()
            customer_dropdown.setStyleSheet("border: 1px solid rgb(171, 173, 179); background-color: yellow;")
            customer_dropdown.setFixedHeight(25)
            customer_dropdown.setFixedWidth(296)
            customer_dropdown.setEditable(False)

            # ADD customer to dropdown Menu
            self.cursor.execute("""
            SELECT customers FROM customer;
            
            """)
            result = self.cursor.fetchall()
            result = sorted(result)
            result = [i[0] for i in result]
            for i in range(len(result)):
                customer_dropdown.addItem(result[i])

            productCode_dropdown = QtWidgets.QComboBox()
            productCode_dropdown.setStyleSheet("border: 1px solid rgb(171, 173, 179); background-color: yellow; ")
            productCode_dropdown.setFixedHeight(25)
            productCode_dropdown.setFixedWidth(296)
            productCode_dropdown.setEditable(True)

            evaluatedBy_dropdown = QtWidgets.QComboBox()
            evaluatedBy_dropdown.setStyleSheet("border: 1px solid rgb(171, 173, 179); background-color: yellow;")
            evaluatedBy_dropdown.setFixedHeight(25)
            evaluatedBy_dropdown.setFixedWidth(296)
            evaluatedBy_dropdown.addItem("Linzy Jam")
            evaluatedBy_dropdown.addItem("Jinky")
            evaluatedBy_dropdown.addItem("Ana")


            date_started_input = QDateTimeEdit()
            date_started_input.setStyleSheet("border: 1px solid rgb(171, 173, 179); background-color: yellow;")
            date_started_input.setFixedHeight(25)
            date_started_input.setFont(font)
            date_started_input.setFixedWidth(296)
            date_now = datetime.now()
            date_started_input.setDateTime(date_now)
            date_started_input.setDisplayFormat("MM-dd-yyyy HH:mm")


            lotNumber_input = QtWidgets.QLineEdit()
            lotNumber_input.setStyleSheet("border: 1px solid rgb(171, 173, 179); background-color: yellow;")
            lotNumber_input.setFixedHeight(25)
            lotNumber_input.setFixedWidth(296)
            lotNumber_input.editingFinished.connect(multiple_lotNumber)

            time_started_input = QTimeEdit()
            time_started_input.setStyleSheet("border: 1px solid rgb(171, 173, 179);")
            time_started_input.setFixedHeight(25)
            time_started_input.setFixedWidth(120)
            time_started_input.setDisplayFormat("HH:mm")
            time_started_input.setFixedWidth(296)

            time_endorsed_input = QDateTimeEdit()
            time_endorsed_input.setStyleSheet("border: 1px solid rgb(171, 173, 179); background-color: rgb(255, 255, 0)")
            time_endorsed_input.setFixedHeight(25)
            time_endorsed_input.setFixedWidth(296)
            time_endorsed_input.setFont(font)
            time_endorsed_input.setDate(date_now)
            time_endorsed_input.setDisplayFormat("MM-dd-yyyy HH:mm")

            remarks_label = QLabel(self.body_widget)
            remarks_label.setGeometry(0, 418, 100, 25)
            remarks_label.setText("   REMARKS")
            remarks_label.setFont(QtGui.QFont("Arial", 9))
            remarks_label.setStyleSheet("border: none;")
            remarks_label.show()

            remarks_box = QtWidgets.QTextEdit(self.body_widget)
            remarks_box.setGeometry(125, 418, 595, 104)
            remarks_box.setStyleSheet("background-color: rgb(255, 255, 255); border: 1px solid rgb(171, 173, 179)")
            remarks_box.show()

            result_dropdown = QComboBox()
            result_dropdown.setFixedWidth(296)
            result_dropdown.setFixedHeight(25)
            result_dropdown.setStyleSheet("background-color: rgb(255, 255, 0); border: 1px solid rgb(171, 173, 179)")
            result_dropdown.addItem("Passed")
            result_dropdown.addItem("Failed")

            correction_label = QLabel()
            correction_label.setText("CORRECTION")
            correction_label.setFixedWidth(110)
            correction_label.setFixedHeight(20)
            correction_label.setFont(font)
            correction_label.setStyleSheet("border: none;")

            updatedBy_input = QComboBox()
            updatedBy_input.setStyleSheet(
                "background-color: rgb(240, 240, 240); border: 1px solid rgb(171, 173, 179);")
            updatedBy_input.setFixedHeight(25)
            updatedBy_input.setEnabled(False)
            updatedBy_input.setFixedWidth(296)
            updatedBy_input.addItem("")
            updatedBy_input.addItem("Linzy Jam")
            updatedBy_input.addItem("Ana")
            updatedBy_input.addItem("Jinky")

            correction_input = QLineEdit()
            correction_input.setStyleSheet("background-color: rgb(240, 240, 240); border: 1px solid rgb(171, 173, 179);")
            correction_input.setFixedHeight(25)
            correction_input.setEnabled(False)
            correction_input.setFixedWidth(296)
            correction_input.editingFinished.connect(correction_auto_input)
            correction_input.editingFinished.connect(get_old_lotNumbers)

            formulaID_label = QLabel()
            formulaID_label.setText("FORMULA ID")
            formulaID_label.setFixedWidth(110)
            formulaID_label.setFixedHeight(20)
            formulaID_label.setFont(font)
            formulaID_label.setStyleSheet("border: none;")

            formulaID_input = QLineEdit()
            formulaID_input.setStyleSheet("border: 1px solid rgb(171, 173, 179); background-color: yellow;")
            formulaID_input.setFixedHeight(25)
            formulaID_input.setFixedWidth(296)


            actionTaken_label = QLabel(self.body_widget)
            actionTaken_label.setGeometry(0, 525, 100, 25)
            actionTaken_label.setText("   ACTION TAKEN")
            actionTaken_label.setFont(QtGui.QFont("Arial", 9))
            actionTaken_label.setStyleSheet("border: none;")
            actionTaken_label.show()

            actionTake_box = QTextEdit(self.body_widget)
            actionTake_box.setGeometry(125, 525, 595, 104)
            actionTake_box.setStyleSheet("background-color: rgb(255, 255, 255); border: 1px solid rgb(171, 173, 179)")
            actionTake_box.show()

            lotNumbers_board = QtWidgets.QTextEdit(self.body_widget)
            lotNumbers_board.setGeometry(470, 200, 250, 180)
            lotNumbers_board.setStyleSheet("border:none; background-color: rgb(240, 240, 240)")
            lotNumbers_board.setEnabled(False)
            lotNumbers_board.setFont(font)
            lotNumbers_board.show()

            label3 = QLabel(self.body_widget)
            label3.setText(" WORKSTATION AND USERNAME")
            label3.setGeometry(0, 650, 194, 25)
            label3.setFont(QtGui.QFont("Arial", 9))
            label3.setStyleSheet("border: none;")
            label3.show()

            user_input = QLineEdit(self.body_widget)
            user_input.setGeometry(194, 650, 221, 25)
            user_input.setEnabled(False)
            user_input.setText(platform.node())
            user_input.setStyleSheet("border: 1px solid rgb(177, 206, 237); background-color: rgb(192, 192, 192); ")
            user_input.show()

            save_btn = QPushButton(self.body_widget)
            save_btn.setGeometry(804, 648, 60, 27)
            save_btn.setStyleSheet("border-radius: 2px; border: 1px solid rgb(171, 173, 179);")
            save_btn.setText("SAVE")
            save_btn.clicked.connect(saveBtn_clicked)
            save_btn.setShortcut("Return")
            save_btn.show()

            clear_btn = QPushButton(self.body_widget)
            clear_btn.setGeometry(866, 648, 60, 27)
            clear_btn.setStyleSheet("border-radius: 2px; border: 1px solid rgb(171, 173, 179);")
            clear_btn.setText("CLEAR")
            clear_btn.show()

            close_btn = QPushButton(self.body_widget)
            close_btn.setGeometry(928, 648, 60, 27)
            close_btn.setStyleSheet("border-radius: 2px; border: 1px solid rgb(171, 173, 179);")
            close_btn.setText("CLOSE")
            close_btn.show()

            topFormLayout = QFormLayout(widget1)
            topFormLayout.addRow(qcType_label, qcType_dropdown)
            topFormLayout.addRow(qcControl_label, qcControl_input)
            topFormLayout.addRow(correction_label, correction_input)
            topFormLayout.addRow(customer_label, customer_dropdown)
            topFormLayout.addRow(lotNumber_label, lotNumber_input)
            topFormLayout.addRow(productCode_label, productCode_dropdown)
            topFormLayout.addRow(formulaID_label, formulaID_input)
            topFormLayout.addRow(evaluatedBy_label, evaluatedBy_dropdown)
            topFormLayout.addRow(date_started_label, date_started_input)
            topFormLayout.addRow(time_endorsed_label, time_endorsed_input)
            topFormLayout.addRow(result_label, result_dropdown)
            topFormLayout.addRow(updatedBy_label, updatedBy_input)


            widget1.show()

        def show_qc_data():
            self.qc_widget.deleteLater()

            self.qc_widget = QtWidgets.QWidget(self.main_widget)
            self.qc_widget.setGeometry(0, 0, 991, 751)
            self.qc_widget.setStyleSheet("background-color: rgb(240,240,240);")
            self.qc_widget.show()

            self.qcBtn_topBorder = QtWidgets.QWidget(self.qc_widget)
            self.qcBtn_topBorder.setGeometry(0, 0, 991, 30)
            self.qcBtn_topBorder.show()

            self.qc_TableBtn = QtWidgets.QPushButton(self.qcBtn_topBorder)
            self.qc_TableBtn.setGeometry(0, 0, 150, 30)
            self.qc_TableBtn.setText("Evaluated Products")
            self.qc_TableBtn.setCursor(Qt.PointingHandCursor)
            self.qc_TableBtn.setFont(QtGui.QFont("Arial", 11))
            self.qc_TableBtn.setStyleSheet("border: 1px solid rgb(160, 160, 160); text-align: top;")
            self.qc_TableBtn.clicked.connect(self.quality_control)
            self.qc_TableBtn.setShortcut("F1")
            self.qc_TableBtn.show()

            self.qc_addEntryBtn = QtWidgets.QPushButton(self.qcBtn_topBorder)
            self.qc_addEntryBtn.setGeometry(150, 0, 150, 30)
            self.qc_addEntryBtn.setText("Evaluation Entry")
            self.qc_addEntryBtn.setCursor(Qt.PointingHandCursor)
            self.qc_addEntryBtn.setStyleSheet("border: 1px solid rgb(160, 160, 160); text-align: top;")
            self.qc_addEntryBtn.clicked.connect(evaluation_entry)
            self.qc_addEntryBtn.setFont(QtGui.QFont("Arial", 11))
            self.qc_addEntryBtn.setShortcut("F2")
            self.qc_addEntryBtn.show()

            self.qc_dataBtn = QtWidgets.QPushButton(self.qcBtn_topBorder)
            self.qc_dataBtn.setGeometry(300, 0, 150, 30)
            self.qc_dataBtn.setText("QC Data")
            self.qc_dataBtn.setCursor(Qt.PointingHandCursor)
            self.qc_dataBtn.setStyleSheet("border: 1px solid rgb(160, 160, 160); color: rgb(0,109,189);")
            self.qc_dataBtn.clicked.connect(show_qc_data)
            self.qc_dataBtn.setFont(QtGui.QFont("Arial", 11))
            self.qc_dataBtn.setShortcut("F3")
            self.qc_dataBtn.show()

            self.dashboardBtn = QtWidgets.QPushButton(self.qcBtn_topBorder)
            self.dashboardBtn.setGeometry(450, 0, 150, 30)
            self.dashboardBtn.setText("Dashboard")
            self.dashboardBtn.setCursor(Qt.PointingHandCursor)
            self.dashboardBtn.setStyleSheet("border: 1px solid rgb(160, 160, 160); text-align: top;")
            self.dashboardBtn.clicked.connect(show_dashboards)
            self.dashboardBtn.setFont(QtGui.QFont("Arial", 11))
            self.dashboardBtn.setShortcut("F4")
            self.dashboardBtn.show()

            self.qc_returns = QtWidgets.QPushButton(self.qcBtn_topBorder)
            self.qc_returns.setGeometry(600, 0, 150, 30)
            self.qc_returns.setText("Returns")
            self.qc_returns.setCursor(Qt.PointingHandCursor)
            self.qc_returns.setStyleSheet("border: 1px solid rgb(160, 160, 160); text-align: top;")
            self.qc_returns.setFont(QtGui.QFont("Arial", 11))
            self.qc_returns.clicked.connect(qc_returns)
            self.qc_returns.setShortcut("F5")
            self.qc_returns.show()

            self.body_widget = QtWidgets.QWidget(self.qc_widget)
            self.body_widget.setGeometry(0, 30, 991, 721)
            self.body_widget.setStyleSheet(
                "background-color: rgb(239, 243, 254); border-top : 1px solid rgb(160, 160, 160);")
            self.body_widget.show()

            qc_data_table = QTableWidget(self.body_widget)
            qc_data_table.setGeometry(50, 20, 890, 340)
            qc_data_table.setStyleSheet("border: 1px solid black; ")
            qc_data_table.horizontalHeader().setStyleSheet("""
                    QHeaderView::section{
                    font-weight: bold;
                    background-color: rgb(0, 109, 189);
                    color: white;
                    }

                    """)

            ph_holiday = hd.country_holidays('PH')
            self.cursor.execute("""
            SELECT original_lot, MIN(evaluation_date)::DATE as min_date, MAX(evaluation_date)::DATE as max_date
            FROM qc_num_days  
            GROUP BY original_lot

            """)
            result = self.cursor.fetchall()

            dayoff = []
            original_lot = []
            for entry in result:

                prod_code = entry[0]
                min_date = entry[1]
                max_date = entry[2]

                date_range = pd.date_range(start=f'{min_date}', end=f'{max_date}').strftime('%m-%d-%Y')

                holidays = []
                sundays = []
                for i in list(date_range):
                    if i in ph_holiday:
                        holidays.append(i)
                    if datetime.strptime(i, '%m-%d-%Y').weekday() == 6:
                        sundays.append(i)
                no_operation = holidays + sundays
                no_operation = len(set(no_operation))
                dayoff.append(no_operation)
                original_lot.append(prod_code)


            data = [(x, y) for x, y in zip(original_lot, dayoff)]

            # DELETE THE TABLE TO CLEAR THE CONTENT FOR UPDATE
            self.cursor.execute("""
            DELETE FROM qc_dayoff

            """)
            self.conn.commit()

            insert_query = sql.SQL("""
                        INSERT INTO qc_dayoff
                        VALUES(%s, %s)
            """)

            self.cursor.executemany(insert_query, data)
            self.conn.commit()



            # get the data from the Database
            self.cursor.execute("""
            SELECT t1.qc_id, lot_number, evaluation_date, t1.original_lot, status, product_code, t1.qc_days - (t2.dayoff || ' day')::interval AS adjusted_qc_days
            FROM qc_num_days AS t1
            JOIN qc_dayoff AS t2 ON t1.original_lot = t2.original_lot
            ORDER BY t1.qc_id DESC
            
            ;
            """)

            result = self.cursor.fetchall()

            # Set Row Count
            qc_data_table.setRowCount(len(result))
            qc_data_table.setColumnCount(7)


            for i in range(len(result)):
                if result[i][4] == "Failed":
                    for j in range(len(result[i])):
                        item = QTableWidgetItem(str(result[i][j]))
                        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                        item.setTextAlignment(Qt.AlignCenter)
                        item.setBackground(QtGui.QColor(252, 3, 28))
                        item.setForeground(QtGui.QColor(0, 0, 0))
                        qc_data_table.setItem(i, j, item)
                else:
                    for j in range(len(result[i])):
                        item = QTableWidgetItem(str(result[i][j]))
                        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                        item.setTextAlignment(Qt.AlignCenter)
                        qc_data_table.setItem(i, j, item)

            qc_data_table.setColumnWidth(1, 150)
            qc_data_table.setColumnWidth(2, 150)
            qc_data_table.setColumnWidth(3, 150)
            qc_data_table.setColumnWidth(6, 118)
            qc_data_table.verticalHeader().setVisible(False)
            qc_data_table.setHorizontalHeaderLabels(["QC ID", "LOT NUMBER", "EVALUATION DATE", "ORIGINAL LOT", "STATUS",
                                                     "PRODUCT CODE", "QC DAYS"])
            qc_data_table.show()


            self.cursor.execute("""
            WITH numbered_row AS (SELECT * , ROW_NUMBER() OVER (PARTITION BY original_lot order by evaluation_date) AS rn
            FROM quality_control_tbl2)
            SELECT numbered_row.original_lot, numbered_row.status, product_code, numbered_row.rn 
            FROM numbered_row

            """)

            result = self.cursor.fetchall()
            df = pd.DataFrame(result)
            df.columns = ["original_lot", "status", "product_code", "row_number"]
            passToFail_counter = {}
            firstTry_failed = {}

            # Getting the number of Lot Number with Passed Status and then become Failed Later
            for index, row in df.iterrows():
                original_lot = row['original_lot']
                status = row['status']
                product_code = row['product_code']
                row_number = row['row_number']
                if status == "Passed" and row_number == 1:
                    try:
                        if df.iat[index + 1, 1] == 'Failed' and df.iat[index + 1, 0] == original_lot:

                            if product_code not in passToFail_counter.keys():
                                passToFail_counter[product_code] = 1
                            else:
                                passToFail_counter[product_code] += 1
                    except Exception as e:
                        print(e)

                # For getting the Failed Lot on the First Qc
                elif status == "Failed" and row_number == 1:
                    if product_code not in firstTry_failed.keys():
                        firstTry_failed[product_code] = 1
                    else:
                        firstTry_failed[product_code] += 1


            passToFail_x = []
            passToFail_y = []


            # Query For Getting the total Amount of original_lot per Product Code
            self.cursor.execute("""
            SELECT product_code, COUNT(*) AS total_quantity
            FROM (SELECT DISTINCT ON (product_code, original_lot) *
                  FROM quality_control_tbl2
                  ORDER BY product_code, original_lot, evaluation_date ) AS distinct_lots
            GROUP BY product_code
            ORDER BY product_code;
            
            """)
            result = self.cursor.fetchall()
            total_productcode = {}
            for i in result:
                total_productcode[i[0]] = i[1]

            # Getting the percentage of Pass to Fail of Product Codes
            passToFail_percentage = {}
            for key in passToFail_counter.keys():
                passToFail_percentage[key] = passToFail_counter[key] / total_productcode[key]

            # Get the DISTINCT OF PRODUCT CODE
            self.cursor.execute("""
            SELECT DISTINCT product_code
            FROM quality_control_tbl2
            
            """)
            result = self.cursor.fetchall()
            prod_code_list = [i[0] for i in result] # Parse the data

            # Add the other product code and Set the Other Value to 0
            for i in prod_code_list:
                if i not in passToFail_percentage.keys():
                    passToFail_percentage[i] = 0

            # unpacking the dictionary to list
            for key, value in passToFail_percentage.items():
                passToFail_x.append(key)
                passToFail_y.append(value)

            # Table For Showing Average QC days per Product Code
            aggregated_products_table = QTableWidget(self.body_widget)
            aggregated_products_table.setGeometry(50, 390, 300, 300)
            aggregated_products_table.setColumnCount(3)
            aggregated_products_table.setRowCount(10)
            aggregated_products_table.verticalHeader().setVisible(False)
            aggregated_products_table.setHorizontalHeaderLabels(["Product Code", "Average QC days", "Pass to Fail"])
            aggregated_products_table.show()



        def show_dashboards():

            def get_data():

                # This is for the Other Graphs that cant make it in the First Six
                def change_graph():
                    if graph5_selections.currentText() == 'Supervisor':
                        self.cursor.execute(f"""
                                                SELECT supervisor, COUNT(supervisor)
                                                FROM
                                                (SELECT t1.lot_number, t2.operator, t2.supervisor
                                                FROM returns t1
                                                JOIN extruder t2 ON t1.origin_lot = ANY(t2.lot_number)
                                                WHERE t1.return_date BETWEEN '2024-{date1}-01' AND '2024-{date2}-{calendar.monthrange(2024, date2)[1]}')
                                                GROUP BY supervisor

                                                                                """)
                        result = self.cursor.fetchall()
                        x = []
                        y = []
                        for i, j in result:
                            x.append(i)
                            y.append(j)
                        self.graph5.clear()
                        self.graph5.bar(x, y)
                        self.graph5.set_yticks(np.arange(0, 110, 10))
                        self.graph5.set_ylim(0,110)
                        self.graph5.set_xticklabels(x, rotation=30)
                        self.graph5.set_title("Highest AVG Returns By Supervisor", fontsize=15)
                        self.canvas.draw()
                    elif graph5_selections.currentText() == 'Operator':

                        self.cursor.execute(f"""
                                            SELECT operator, COUNT(operator)
                                            FROM
                                            (SELECT t1.lot_number, t2.operator, t2.supervisor
                                            FROM returns t1
                                            JOIN extruder t2 ON t1.origin_lot = ANY(t2.lot_number)
                                            WHERE t1.return_date BETWEEN '2024-{date1}-01' AND '2024-{date2}-{calendar.monthrange(2024, date2)[1]}')
                                            GROUP BY operator

                                        """)
                        result = self.cursor.fetchall()
                        x = []
                        y = []
                        for i, j in result:
                            x.append(i)
                            y.append(j)

                        self.graph5.clear()
                        self.graph5.bar(x, y)
                        self.graph5.set_yticks(np.arange(0, 110, 10))
                        self.graph5.set_ylim(0, 110)
                        self.graph5.set_xticklabels(x, rotation=30)
                        self.graph5.set_title("Highest AVG Returns By Operator", fontsize=15)
                        self.graph5.set_position([0.4, 0.08, 0.228, 0.35])

                        self.canvas.draw()

                    elif graph5_selections.currentText() == 'QC Analyst':

                        self.cursor.execute(f"""
                        SELECT evaluated_by, COUNT(evaluated_by)
                        FROM (SELECT t1.*, t2.evaluated_by
                        FROM returns t1
                        JOIN quality_control t2 ON t1.origin_lot = t2.lot_number
                        WHERE t1.return_date BETWEEN '2024-{date1}-01' AND '2024-{date2}-{calendar.monthrange(2024, date2)[1]}')
                        GROUP BY evaluated_by

                        """)
                        result = self.cursor.fetchall()


                        x = []
                        y = []
                        for i, j in result:
                            x.append(i)
                            y.append(j)

                        self.graph5.clear()
                        self.graph5.bar(x, y)
                        self.graph5.set_yticks(np.arange(0, 110, 10))
                        self.graph5.set_ylim(0, 110)
                        self.graph5.set_xticklabels(x, rotation=30)
                        self.graph5.set_title("Highest Returns By QC Analyst", fontsize=15)
                        self.graph5.set_position([0.4, 0.08, 0.228, 0.35])


                        self.canvas.draw()

                    elif graph5_selections.currentText() == 'Extruder':

                        self.cursor.execute(f"""
                        SELECT machine, COUNT(machine)
                        FROM
                        (SELECT t1.lot_number, t2.machine
                        FROM returns t1
                        JOIN extruder t2 ON t1.origin_lot = ANY(t2.lot_number)
                        WHERE t1.return_date BETWEEN '2024-{date1}-01' AND '2024-{date2}-{calendar.monthrange(2024, date2)[1]}')
                        GROUP BY machine
                        """)

                        result = self.cursor.fetchall()

                        x = []
                        y = []
                        for i, j in result:
                            x.append(i)
                            y.append(j)

                        self.graph5.clear()
                        self.graph5.bar(x, y)
                        self.graph5.set_yticks(np.arange(0, 110, 10))
                        self.graph5.set_ylim(0, 110)
                        self.graph5.set_xticklabels(x, rotation=30)
                        self.graph5.set_title("Highest Returns By Extruder", fontsize=15)
                        self.graph5.set_position([0.4, 0.08, 0.228, 0.35])

                        self.canvas.draw()

                    elif graph5_selections.currentText() == 'Total Kg':

                        self.cursor.execute(f"""
                        SELECT product_code, SUM(quantity) as total_kg
                        FROM returns
                        WHERE return_date BETWEEN '2024-{date1}-01' AND '2024-{date2}-{calendar.monthrange(2024, date2)[1]}'
                        GROUP BY product_code
                        
                        ORDER BY total_kg
                        """)

                        result = self.cursor.fetchall()

                        x = []
                        y = []
                        for i, j in result:
                            x.append(i)
                            y.append(j)

                        self.graph5.clear()
                        self.graph5.bar(x, y)
                        self.graph5.set_xticklabels(x, rotation=30)
                        self.graph5.set_title("Total Returns(kg) By Product Code", fontsize=15)
                        self.graph5.set_position([0.4, 0.08, 0.228, 0.35])

                        self.canvas.draw()



                date1 = dateFrom_widget.currentIndex()+1
                date2 = dateTo_widget.currentIndex()+1

                # Get the last day of Month
                ph_holiday = hd.country_holidays('PH')
                self.cursor.execute(f"""
                            SELECT original_lot, MIN(evaluation_date)::DATE as min_date, MAX(evaluation_date)::DATE as max_date
                            FROM qc_num_days  
                            WHERE evaluation_date::DATE BETWEEN '2024-{date1}-01' AND '2024-{date2}-{calendar.monthrange(2024, date2)[1]}'
                            GROUP BY original_lot

                            """)
                result = self.cursor.fetchall()

                dayoff = []
                original_lot = []

                for entry in result:
                    prod_code = entry[0]
                    min_date = entry[1]
                    max_date = entry[2]

                    date_range = pd.date_range(start=f'{min_date}', end=f'{max_date}').strftime('%m-%d-%Y')

                    holidays = []
                    sundays = []
                    for i in list(date_range):
                        if i in ph_holiday:
                            holidays.append(i)
                        if datetime.strptime(i, '%m-%d-%Y').weekday() == 6:
                            sundays.append(i)

                    no_operation = holidays + sundays
                    no_operation = len(set(no_operation))
                    dayoff.append(no_operation)
                    original_lot.append(prod_code)

                data = [(x, y) for x, y in zip(original_lot, dayoff)]

                # DELETE THE TABLE TO CLEAR THE CONTENT FOR UPDATE
                self.cursor.execute("""DELETE FROM qc_dayoff
                     """)
                self.conn.commit()
                insert_query = sql.SQL("""
                INSERT INTO qc_dayoff
                VALUES(%s, %s)
                """)

                self.cursor.executemany(insert_query, data)
                self.conn.commit()

                # Query For Getting the AVERAGE qc_days PER PRODUCT_CODE
                self.cursor.execute(f"""
                 SELECT product_code,
    round(avg(days_float), 4) AS avg_qcdays
   FROM ( SELECT t1.product_code,
            EXTRACT(epoch FROM t1.qc_days - ((t2.dayoff || ' day'::text)::interval)) / 86400.0 AS days_float
           FROM ( WITH aggregated_materials AS (
         SELECT max(quality_control_tbl2.evaluation_date) - min(quality_control_tbl2.evaluation_date) AS qc_days,
            quality_control_tbl2.original_lot
           FROM quality_control_tbl2
		   WHERE evaluation_date::DATE BETWEEN '2024-{date1}-01' AND '2024-{date2}-{calendar.monthrange(2024, date2)[1]}'
          GROUP BY quality_control_tbl2.original_lot
          
        )
 SELECT q.id AS qc_id,
    q.lot_number,
    q.evaluation_date,
    q.original_lot,
    q.status,
    q.product_code,
    a.qc_days,
    q.qc_type
	
   FROM aggregated_materials a
     JOIN quality_control_tbl2 q ON a.original_lot::text = q.original_lot::text) t1
             JOIN qc_dayoff t2 ON t1.original_lot::text = t2.original_lot::text
          GROUP BY t1.product_code, (EXTRACT(epoch FROM t1.qc_days - ((t2.dayoff || ' day'::text)::interval)) / 86400.0)) subquery
  GROUP BY product_code
  ORDER BY avg_qcdays DESC;
                
                """)
                layout = QHBoxLayout(self.body_widget)

                self.figure = plt.figure(figsize=(2, 2), dpi=60)
                self.canvas = FigureCanvas(self.figure)
                layout.addWidget(self.canvas)
                self.figure.patch.set_facecolor("#eff3fe")

                font = QtGui.QFont("Arial", 14)

                dateRangeLabel = QLabel(self.body_widget)
                dateRangeLabel.setGeometry(450, 25, 320, 30)
                dateRangeLabel.setAlignment(Qt.AlignCenter)
                if dateTo_widget.currentText() == dateFrom_widget.currentText():
                    dateRangeLabel.setText(dateFrom_widget.currentText())
                else:
                    dateRangeLabel.setText(dateFrom_widget.currentText() + " - " + dateTo_widget.currentText())
                dateRangeLabel.setFont(font)
                dateRangeLabel.setStyleSheet("color: rgb(0, 109, 189); border: none;")
                dateRangeLabel.show()

                self.graph1 = self.figure.add_subplot(231)
                self.graph2 = self.figure.add_subplot(232)
                self.graph3 = self.figure.add_subplot(233)
                self.graph4 = self.figure.add_subplot(234)
                self.graph5 = self.figure.add_subplot(235)
                self.graph6 = self.figure.add_subplot(236)

                result = self.cursor.fetchall()
                x = []
                y = []
                # Unpack the List of tuple
                for item, value in result:
                    x.append(item)
                    y.append(value)

                # slice to only top 5
                x = x[:5]
                y = y[:5]
                self.graph1.bar(x, y)
                self.graph1.set_xticklabels(x, rotation=30)
                self.graph1.set_ylim(0, 110)
                self.graph1.set_yticks(np.arange(0, 110, 10))

                self.graph1.set_title("Highest AVG QC days", fontsize=15)

                # Getting the Pass to Fail And Fail TO pass QC
                self.cursor.execute(f"""
                            WITH numbered_row AS (SELECT * , ROW_NUMBER() OVER (PARTITION BY original_lot order by evaluation_date) AS rn
            FROM (SELECT * FROM quality_control_tbl2 WHERE evaluation_date BETWEEN '2024-{date1}-01' AND '2024-{date2}-{calendar.monthrange(2024, date2)[1]}'))
            SELECT numbered_row.original_lot, numbered_row.status, product_code, numbered_row.rn 
            FROM numbered_row			
                            """)

                result = self.cursor.fetchall()
                # Error Handling
                if len(result) == 0:
                    print("NO DATA")
                    return 0

                df = pd.DataFrame(result)
                df.columns = ["original_lot", "status", "product_code", "row_number"]
                passToFail = {}
                failToPass = {}
                failed_firstrun = {}

                # Getting the number of Lot Number with Passed Status and then become Failed Later
                for index, row in df.iterrows():
                    original_lot = row['original_lot']
                    status = row['status']
                    product_code = row['product_code']
                    row_number = row['row_number']
                    if status == "Passed" and row_number == 1:
                        try:
                            if df.iat[index + 1, 1] == 'Failed' and df.iat[index + 1, 0] == original_lot:
                                if product_code not in passToFail.keys():
                                    passToFail[product_code.strip()] = 1
                                else:
                                    passToFail[product_code.strip()] += 1
                        except Exception as e:
                            print(e)
                    elif status == "Failed" and row_number == 1:
                        if product_code not in failToPass.keys():
                            failToPass[product_code.strip()] = 1
                            failed_firstrun[product_code.strip()] = 1
                        else:
                            failToPass[product_code.strip()] += 1
                            failed_firstrun[product_code.strip()] += 1

                # Get the Total Product Code runs from data x to date y
                self.cursor.execute(f"""
                SELECT product_code, COUNT(*) AS total_quantity
            FROM (SELECT DISTINCT ON (product_code, original_lot) *
            FROM quality_control_tbl2
		    WHERE evaluation_date BETWEEN '2024-{date1}-01' AND '2024-{date2}-{calendar.monthrange(2024, date2)[1]}'
            ORDER BY product_code, original_lot, evaluation_date ) AS distinct_lots
            GROUP BY product_code
            ORDER BY product_code

                """)
                result = self.cursor.fetchall()

                total_productCodes = {}
                for i in result:
                    total_productCodes[i[0].strip()] = i[1]

                total_changeProductCodes = {}
                # get the percentage of each product code From PASS TO FAIL AND FAIL TO PAS
                for key,value in passToFail.items():
                    passToFail[key] = passToFail[key] / total_productCodes[key]

                for key, value in failToPass.items():
                    failToPass[key] = failToPass[key] / total_productCodes[key]

                # Combine the codes to see the most change
                for key in total_productCodes.keys():
                    if key in passToFail.keys():
                        total_changeProductCodes[key] = passToFail[key]
                    if key in failToPass.keys():
                        if key in total_changeProductCodes.keys():
                            total_changeProductCodes[key] += failToPass[key]
                        else:
                            total_changeProductCodes[key] = failToPass[key]

                # Sort the total_changeProductCodes
                total_changeProductCodes = sorted(total_changeProductCodes.items(), key=lambda x:x[1], reverse=True)

                total_changeProductCodes = dict(total_changeProductCodes)

                # Graph2
                x = []
                y1 = []
                y2 = []

                for key in total_changeProductCodes.keys():
                    try:
                        y1.append(passToFail[key])
                    except:
                        y1.append(0)
                    try:
                        y2.append(failToPass[key])
                    except:
                        y2.append(0)
                    x.append(key)

                for i in range(len(x)):
                    y1[i] = y1[i] * 100
                    y2[i] = y2[i] * 100

                std_ProductCodesTotal = {}
                for key, value in total_productCodes.items():
                    if total_productCodes[key] >= 5:
                        std_ProductCodesTotal[key] = value
                    else:
                        pass

                x_axis = []
                y1_axis = []
                y2_axis = []
                for i in range(len(x)):
                    if x[i] in std_ProductCodesTotal.keys():
                        x_axis.append(x[i])
                        y1_axis.append(y1[i])
                        y2_axis.append(y2[i])

                self.graph2.bar(x_axis, y1_axis)
                self.graph2.bar(x_axis, y2_axis, bottom=y1_axis)
                self.graph2.set_yticks(np.arange(0, 110, 10))
                self.graph2.set_ylim(0, 110)
                self.graph2.set_xticklabels(x_axis, rotation=30)
                self.graph2.set_title("Most Change", fontsize = 15)
                self.graph2.legend(["Pass to Fail", "Fail to pass"], loc = "upper right")

                # Getting the Percentage of Each Product Codes Failed in First Run
                for key, value in failed_firstrun.items():
                    failed_firstrun[key] = failed_firstrun[key] / total_productCodes[key]

                # Visual 4

                self.cursor.execute(f"""
                SELECT product_code, COUNT(*) 
                FROM returns
				WHERE return_date BETWEEN '2024-{date1}-01' AND '2024-{date2}-{calendar.monthrange(2024, date2)[1]}'
                GROUP BY product_code				
                LIMIT 5
                """)

                result = self.cursor.fetchall()

                productCodeReturns = {}


                for i, j in result:
                    productCodeReturns[i] = j

                # Getting the Percentage of Product Code returns for each Product Codes

                for key in productCodeReturns.keys():
                    productCodeReturns[key] = (productCodeReturns[key] / total_productCodes[key]) * 100

                x = []
                y = []
                # Getting the x and y from the Returns dictionary
                for key, value in productCodeReturns.items():
                    x.append(key)
                    y.append(value)

                # Plot the Visual 4

                self.graph4.bar(x, y)
                self.graph4.set_yticks(np.arange(0, 110, 10))
                self.graph4.set_ylim(0,110)
                self.graph4.set_xticklabels(x, rotation=30)
                self.graph4.set_title("Highest Return Percentage", fontsize=15)
                self.graph4.set_position([0.125, 0.08, 0.228, 0.35])


                # Getting the data For Visualization 5

                # Getting the Data for how many times an operator have a Returned product
                self.cursor.execute(f"""
                                    SELECT operator, COUNT(operator)
                                    FROM
                                    (SELECT t1.lot_number, t2.operator, t2.supervisor
                                    FROM returns t1
                                    JOIN extruder t2 ON t1.origin_lot = ANY(t2.lot_number)
                                    WHERE t1.return_date BETWEEN '2024-{date1}-01' AND '2024-{date2}-{calendar.monthrange(2024, date2)[1]}')
                                    GROUP BY operator
                
                """)
                result = self.cursor.fetchall()
                x = []
                y = []
                for i, j in result:
                    x.append(i)
                    y.append(j)

                # Set Page for Graph 5
                graph5_page = 1


                self.graph5.bar(x, y)
                self.graph5.set_yticks(np.arange(0, 110, 10))
                self.graph5.set_ylim(0, 110)
                self.graph5.set_xticklabels(x, rotation = 30)
                self.graph5.set_title("Highest AVG Returns By Operator", fontsize = 15)
                self.graph5.set_position([0.4, 0.08, 0.228, 0.35])

                graph5_selections = QComboBox(self.body_widget)
                graph5_selections.setGeometry(650, 682, 100, 20)

                graph5_selections.addItem('Operator')
                graph5_selections.addItem('Supervisor')
                graph5_selections.addItem('Total Kg')
                graph5_selections.addItem("QC Analyst")
                graph5_selections.addItem('Extruder')
                graph5_selections.currentIndexChanged.connect(change_graph)
                graph5_selections.show()



            self.qc_widget.deleteLater()

            self.qc_widget = QtWidgets.QWidget(self.main_widget)
            self.qc_widget.setGeometry(0, 0, 991, 751)
            self.qc_widget.setStyleSheet("background-color: rgb(240,240,240);")
            self.qc_widget.show()

            self.qcBtn_topBorder = QtWidgets.QWidget(self.qc_widget)
            self.qcBtn_topBorder.setGeometry(0, 0, 991, 30)
            self.qcBtn_topBorder.setStyleSheet("border-bottom: 1px solid rgb(160, 160, 160)")
            self.qcBtn_topBorder.show()

            self.qc_TableBtn = QtWidgets.QPushButton(self.qcBtn_topBorder)
            self.qc_TableBtn.setGeometry(0, 0, 150, 30)
            self.qc_TableBtn.setText("Evaluated Products")
            self.qc_TableBtn.setCursor(Qt.PointingHandCursor)
            self.qc_TableBtn.setFont(QtGui.QFont("Arial", 11))
            self.qc_TableBtn.setStyleSheet("border: 1px solid rgb(160, 160, 160); text-align: top;")
            self.qc_TableBtn.clicked.connect(self.quality_control)
            self.qc_TableBtn.setShortcut("F1")
            self.qc_TableBtn.show()

            self.qc_addEntryBtn = QtWidgets.QPushButton(self.qcBtn_topBorder)
            self.qc_addEntryBtn.setGeometry(150, 0, 150, 30)
            self.qc_addEntryBtn.setText("Evaluation Entry")
            self.qc_addEntryBtn.setCursor(Qt.PointingHandCursor)
            self.qc_addEntryBtn.setStyleSheet("border: 1px solid rgb(160, 160, 160); text-align: top;")
            self.qc_addEntryBtn.clicked.connect(evaluation_entry)
            self.qc_addEntryBtn.setFont(QtGui.QFont("Arial", 11))
            self.qc_addEntryBtn.setShortcut("F2")
            self.qc_addEntryBtn.show()

            self.qc_dataBtn = QtWidgets.QPushButton(self.qcBtn_topBorder)
            self.qc_dataBtn.setGeometry(300, 0, 150, 30)
            self.qc_dataBtn.setText("QC Data")
            self.qc_dataBtn.setCursor(Qt.PointingHandCursor)
            self.qc_dataBtn.setStyleSheet("border: 1px solid rgb(160, 160, 160); text-align: top;")
            self.qc_dataBtn.clicked.connect(show_qc_data)
            self.qc_dataBtn.setFont(QtGui.QFont("Arial", 11))
            self.qc_dataBtn.setShortcut("F3")
            self.qc_dataBtn.show()

            self.dashboardBtn = QtWidgets.QPushButton(self.qcBtn_topBorder)
            self.dashboardBtn.setGeometry(450, 0, 150, 30)
            self.dashboardBtn.setText("Dashboard")
            self.dashboardBtn.setCursor(Qt.PointingHandCursor)
            self.dashboardBtn.setStyleSheet("border: 1px solid rgb(160, 160, 160); color: rgb(0,109,189)")
            self.dashboardBtn.clicked.connect(show_dashboards)
            self.dashboardBtn.setFont(QtGui.QFont("Arial", 11))
            self.dashboardBtn.setShortcut("F4")
            self.dashboardBtn.show()

            self.qc_returns = QtWidgets.QPushButton(self.qcBtn_topBorder)
            self.qc_returns.setGeometry(600, 0, 150, 30)
            self.qc_returns.setText("Returns")
            self.qc_returns.setCursor(Qt.PointingHandCursor)
            self.qc_returns.setStyleSheet("border: 1px solid rgb(160, 160, 160); text-align: top;")
            self.qc_returns.setFont(QtGui.QFont("Arial", 11))
            self.qc_returns.clicked.connect(qc_returns)
            self.qc_returns.setShortcut("F5")
            self.qc_returns.show()

            self.body_widget = QtWidgets.QWidget(self.qc_widget)
            self.body_widget.setGeometry(-120, 30, 1200, 721)
            self.body_widget.setStyleSheet(
                "background-color: rgb(239, 243, 254); border-top : 1px solid rgb(160, 160, 160);")
            self.body_widget.show()

            dateFrom_widget = QtWidgets.QComboBox(self.body_widget)
            dateFrom_widget.setGeometry(200, 30, 100, 20)
            dateFrom_widget.setStyleSheet("background-color: rgb(137, 137, 161)")
            dateFrom_widget.addItem("JANUARY")
            dateFrom_widget.addItem("FEBRUARY")
            dateFrom_widget.addItem("MARCH")
            dateFrom_widget.addItem("APRIL")
            dateFrom_widget.addItem("MAY")
            dateFrom_widget.addItem("JUNE")
            dateFrom_widget.addItem("JULY")
            dateFrom_widget.addItem("AUGUST")
            dateFrom_widget.addItem("SEPTEMBER")
            dateFrom_widget.addItem("OCTOBER")
            dateFrom_widget.addItem("NOVEMBER")
            dateFrom_widget.addItem("DECEMBER")
            dateFrom_widget.show()

            dateTo_widget = QtWidgets.QComboBox(self.body_widget)
            dateTo_widget.setGeometry(350, 30, 100, 20)
            dateTo_widget.setStyleSheet("background-color: rgb(137, 137, 161)")
            dateTo_widget.addItem("JANUARY")
            dateTo_widget.addItem("FEBRUARY")
            dateTo_widget.addItem("MARCH")
            dateTo_widget.addItem("APRIL")
            dateTo_widget.addItem("MAY")
            dateTo_widget.addItem("JUNE")
            dateTo_widget.addItem("JULY")
            dateTo_widget.addItem("AUGUST")
            dateTo_widget.addItem("SEPTEMBER")
            dateTo_widget.addItem("OCTOBER")
            dateTo_widget.addItem("NOVEMBER")
            dateTo_widget.addItem("DECEMBER")
            dateTo_widget.activated.connect(get_data)
            dateTo_widget.show()

        def show_items():
            try:
                # Querying the selected item in the database
                id = self.qc_table.selectedItems()[0].text()
                self.cursor.execute(f"""
                            SELECT * FROM quality_control
                            WHERE id = '{id}'

                            """)
                result = self.cursor.fetchall()[0]

                # Unpacking the items

                lot_num, product_code, customer, status, remarks, action = result[1:7]
                evaluated_by, evaluated_date, encoded_on, updated_by, updated_on = result[8:13]
                time_endorsed = result[13]
                qc_type = result[15]
                formula_id = result[16]

                # showing the Selected Items
                customer_selected.setText(customer)
                productCode_selected.setText(product_code)
                result_selected.setText(status)
                evaluatedBy_selected.setText(evaluated_by)
                evaluatedDate_selected.setText(str(evaluated_date))
                encodedDate_selected.setText(str(encoded_on))
                remarks_box.setText(remarks)

                updatedBy_val1.setText(updated_by)
                time_endorsed_val.setText(str(time_endorsed))
                qc_type_selected.setText(qc_type)
            except:
                QMessageBox.information(self.qc_widget, "Selection Error", "No Items Selected")
                self.qc_table.clearSelection()

        def qc_returns():

            def insert_entry():
                # Check if the lot number is not entered yet in the Database
                self.cursor.execute(f"""
                SELECT * FROM returns
                WHERE  lot_number = '{lot_input.text()}'
                
                """)
                result = self.cursor.fetchall()

                try:
                    if len(result) == 1:
                        QMessageBox.information(self.body_widget, "Data Exist", "Lot Number already exist.")
                        lot_input.clear()
                        quantity_input.clear()
                        remarks_input.clear()
                        lot_input.setFocus()
                        origin_lot.clear()
                        return
                    else:
                        self.cursor.execute(f"""
                                            INSERT INTO returns (lot_number, quantity, product_code, customer, formula_id,
                                            remarks, return_date, origin_lot)
                                            VALUES('{lot_input.text()}', '{quantity_input.text()}', '{product_code_input.text()}', 
                                            '{customer_input.text()}','{formulaID_input.text()}', '{remarks_input.toPlainText()}',
                                            '{date_return.text()}', '{origin_lot.text()}')
                                            """)
                        self.conn.commit()
                        # Clear the Widgets
                        lot_input.clear()
                        quantity_input.clear()
                        remarks_input.clear()
                        origin_lot.clear()
                        QtWidgets.QMessageBox.information(self.qc_widget, "SUCCESS", "Successfully Inserted Data")
                        lot_input.setFocus()
                        show_table()

                except psycopg2.Error as e:
                    self.conn.rollback()
                    print(e)

            def autofill():
                if lot_input.text() in lot_list:
                    self.cursor.execute(f"""
                    SELECT t2.lot_number as origin_lot, t1.lot_number, t1.product_code, t1.formula_id, t2.customer
                    FROM quality_control_tbl2 t1
                    JOIN quality_control t2 ON t1.id = t2.id
                    WHERE t1.lot_number = '{lot_input.text()}'
                    
                    """)
                    result = self.cursor.fetchall()[0]
                    # Set the Text
                    origin_lot.setText(result[0])
                    product_code_input.setText(result[2])
                    formulaID_input.setText(str(result[3]))
                    customer_input.setText(result[4])


                else:

                    product_code_input.clear()
                    formulaID_input.clear()
                    customer_input.clear()

            def show_table():

                self.cursor.execute("""
                SELECT lot_number, quantity, product_code, customer, formula_id, return_date, remarks  
                FROM returns
                ORDER BY return_date DESC
                
                """)
                result = self.cursor.fetchall()

                for i in range(len(result)):
                    for j in range(len(result[i])):
                        item = QTableWidgetItem(str(result[i][j]))
                        item.setTextAlignment(Qt.AlignCenter)
                        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                        returns_table.setItem(i, j, item)

            def search_table():

                returns_table.clear()
                returns_table.setHorizontalHeaderLabels(["Lot Number", "Quantity", "Product Code", "Customer", "Formula ID",
                                                     "Return Date", "Remarks"])


                self.cursor.execute(f"""
                
                SELECT lot_number, quantity, product_code, customer, formula_id, return_date, remarks  
                FROM returns
                WHERE lot_number ILIKE '%{search_bar.text()}%'
                
                
                """)
                result = self.cursor.fetchall()

                for i in range(len(result)):
                    for j in range(len(result[i])):
                        item = QTableWidgetItem(str(result[i][j]))
                        item.setTextAlignment(Qt.AlignCenter)
                        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                        returns_table.setItem(i, j, item)

                returns_table.show()

            def filter_data():
                if masterbatch_checkbox.isChecked() == True and dryColor_checkbox.isChecked() == False:
                    self.cursor.execute("""
                    SELECT lot_number, quantity, product_code, customer, formula_id, return_date, remarks
                    FROM returns
                    WHERE RIGHT(lot_number, 2) ~ '^[A-Za-z]';
                    """)
                    result = self.cursor.fetchall()

                    if len(result) > 20 :
                        returns_table.setColumnCount(len(result))

                    returns_table.clear()
                    returns_table.setHorizontalHeaderLabels(
                        ["Lot Number", "Quantity", "Product Code", "Customer", "Formula ID",
                         "Return Date", "Remarks"])

                    for i in range(len(result)):
                        for j in range(len(result[i])):
                            item = QTableWidgetItem(str(result[i][j]))
                            item.setTextAlignment(Qt.AlignCenter)
                            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                            returns_table.setItem(i, j, item)

                    returns_table.show()
                elif masterbatch_checkbox.isChecked() == False and dryColor_checkbox.isChecked() == True:
                    self.cursor.execute("""
                                        SELECT lot_number, quantity, product_code, customer, formula_id, return_date, remarks
                                        FROM returns
                                        WHERE RIGHT(lot_number, 2) !~ '^[A-Za-z]';
                                        """)
                    result = self.cursor.fetchall()

                    if len(result) > 20:
                        returns_table.setColumnCount(len(result))

                    returns_table.clear()
                    returns_table.setHorizontalHeaderLabels(
                        ["Lot Number", "Quantity", "Product Code", "Customer", "Formula ID",
                         "Return Date", "Remarks"])

                    for i in range(len(result)):
                        for j in range(len(result[i])):
                            item = QTableWidgetItem(str(result[i][j]))
                            item.setTextAlignment(Qt.AlignCenter)
                            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                            returns_table.setItem(i, j, item)

                    returns_table.show()
                else:
                    self.cursor.execute("""
                                        SELECT lot_number, quantity, product_code, customer, formula_id, return_date, remarks
                                        FROM returns
                                        """)
                    result = self.cursor.fetchall()

                    if len(result) > 20:
                        returns_table.setColumnCount(len(result))

                    returns_table.clear()
                    returns_table.setHorizontalHeaderLabels(
                        ["Lot Number", "Quantity", "Product Code", "Customer", "Formula ID",
                         "Return Date", "Remarks"])

                    for i in range(len(result)):
                        for j in range(len(result[i])):
                            item = QTableWidgetItem(str(result[i][j]))
                            item.setTextAlignment(Qt.AlignCenter)
                            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                            returns_table.setItem(i, j, item)

                    returns_table.show()


            def clear_entry():
                lot_input.clear()
                quantity_input.clear()
                remarks_input.clear()



            self.qc_widget.deleteLater()

            self.qc_widget = QtWidgets.QWidget(self.main_widget)
            self.qc_widget.setGeometry(0, 0, 991, 751)
            self.qc_widget.setStyleSheet("background-color: rgb(240,240,240);")
            self.qc_widget.show()

            self.qcBtn_topBorder = QtWidgets.QWidget(self.qc_widget)
            self.qcBtn_topBorder.setGeometry(0, 0, 991, 30)
            self.qcBtn_topBorder.show()

            self.qc_TableBtn = QtWidgets.QPushButton(self.qcBtn_topBorder)
            self.qc_TableBtn.setGeometry(0, 0, 150, 30)
            self.qc_TableBtn.setText("Evaluated Products")
            self.qc_TableBtn.setCursor(Qt.PointingHandCursor)
            self.qc_TableBtn.setFont(QtGui.QFont("Arial", 11))
            self.qc_TableBtn.setStyleSheet("border: 1px solid rgb(160, 160, 160); text-align: top;")
            self.qc_TableBtn.clicked.connect(self.quality_control)
            self.qc_TableBtn.setShortcut("F1")
            self.qc_TableBtn.show()

            self.qc_addEntryBtn = QtWidgets.QPushButton(self.qcBtn_topBorder)
            self.qc_addEntryBtn.setGeometry(150, 0, 150, 30)
            self.qc_addEntryBtn.setText("Evaluation Entry")
            self.qc_addEntryBtn.setCursor(Qt.PointingHandCursor)
            self.qc_addEntryBtn.setStyleSheet("border: 1px solid rgb(160, 160, 160); text-align: top;")
            self.qc_addEntryBtn.clicked.connect(evaluation_entry)
            self.qc_addEntryBtn.setFont(QtGui.QFont("Arial", 11))
            self.qc_addEntryBtn.setShortcut("F2")
            self.qc_addEntryBtn.show()

            self.qc_dataBtn = QtWidgets.QPushButton(self.qcBtn_topBorder)
            self.qc_dataBtn.setGeometry(300, 0, 150, 30)
            self.qc_dataBtn.setText("QC Data")
            self.qc_dataBtn.setCursor(Qt.PointingHandCursor)
            self.qc_dataBtn.setStyleSheet("border: 1px solid rgb(160, 160, 160); text-align: top;")
            self.qc_dataBtn.clicked.connect(show_qc_data)
            self.qc_dataBtn.setFont(QtGui.QFont("Arial", 11))
            self.qc_dataBtn.setShortcut("F3")
            self.qc_dataBtn.show()

            self.dashboardBtn = QtWidgets.QPushButton(self.qcBtn_topBorder)
            self.dashboardBtn.setGeometry(450, 0, 150, 30)
            self.dashboardBtn.setText("Dashboard")
            self.dashboardBtn.setCursor(Qt.PointingHandCursor)
            self.dashboardBtn.setStyleSheet("border: 1px solid rgb(160, 160, 160); text-align: top;")
            self.dashboardBtn.clicked.connect(show_dashboards)
            self.dashboardBtn.setFont(QtGui.QFont("Arial", 11))
            self.dashboardBtn.setShortcut("F4")
            self.dashboardBtn.show()

            self.body_widget = QtWidgets.QWidget(self.qc_widget)
            self.body_widget.setGeometry(0, 30, 991, 721)
            self.body_widget.setStyleSheet(
                "background-color:rgb(239, 243, 254); border-top : 1px solid rgb(160, 160, 160);")
            self.body_widget.show()

            self.qc_returns = QtWidgets.QPushButton(self.qcBtn_topBorder)
            self.qc_returns.setGeometry(600, 0, 150, 30)
            self.qc_returns.setText("Returns")
            self.qc_returns.setCursor(Qt.PointingHandCursor)
            self.qc_returns.setStyleSheet("border: 1px solid rgb(160, 160, 160); color: rgb(0,109,189)")
            self.qc_returns.setFont(QtGui.QFont("Arial", 11))
            self.qc_returns.clicked.connect(qc_returns)
            self.qc_returns.setShortcut("F5")
            self.qc_returns.show()

            header_widget = QWidget(self.body_widget)
            header_widget.setGeometry(0, 0, 991, 60)
            header_widget.setStyleSheet('border: 1px solid black; background-color:rgb(239, 243, 254)')
            header_widget.show()

            edited_checkbox = QCheckBox(header_widget)
            edited_checkbox.move(355, 5)
            edited_checkbox.setStyleSheet("border: none;")
            edited_checkbox.show()

            edited_label = QLabel(header_widget)
            edited_label.setGeometry(370, 4, 90, 15)
            edited_label.setStyleSheet("border: none;")
            edited_label.setText("EDITED RECORDS")
            edited_label.setFont(QtGui.QFont("Arial", 8))
            edited_label.show()

            masterbatch_checkbox = QCheckBox(header_widget)
            masterbatch_checkbox.move(470, 5)
            masterbatch_checkbox.setStyleSheet('border:none')
            masterbatch_checkbox.stateChanged.connect(filter_data)
            masterbatch_checkbox.show()

            masterbatch_label = QLabel(header_widget)
            masterbatch_label.setGeometry(485, 4, 90, 15)
            masterbatch_label.setStyleSheet("border: none;")
            masterbatch_label.setText("MASTERBATCH")
            masterbatch_label.setFont(QtGui.QFont("Arial", 8))
            masterbatch_label.show()

            dryColor_checkbox = QCheckBox(header_widget)
            dryColor_checkbox.move(580, 5)
            dryColor_checkbox.setStyleSheet('border:none')
            dryColor_checkbox.stateChanged.connect(filter_data)
            dryColor_checkbox.show()

            drycolor_label = QLabel(header_widget)
            drycolor_label.setGeometry(595, 4, 90, 15)
            drycolor_label.setStyleSheet("border: none;")
            drycolor_label.setText("DRYCOLOR")
            drycolor_label.setFont(QtGui.QFont("Arial", 8))
            drycolor_label.show()

            search_bar = QtWidgets.QLineEdit(header_widget)
            search_bar.setGeometry(770, 5, 150, 25)
            search_bar.setStyleSheet("border: 1px solid rgb(171, 173, 179); background-color: rgb(255, 255, 17);")
            search_bar.setFont(QtGui.QFont("Arial", 9))
            search_bar.setPlaceholderText("Lot Number")
            search_bar.show()

            search_btn = QtWidgets.QPushButton(header_widget)
            search_btn.setGeometry(925, 5, 60, 25)
            search_btn.setStyleSheet("border: 1px solid rgb(171, 173, 179);")
            search_btn.setText("Search")
            search_btn.setShortcut("Ctrl+Return")
            search_btn.clicked.connect(search_table)
            search_btn.show()

            self.body_widget.setStyleSheet("border: none")
            entry_widget = QtWidgets.QWidget(self.body_widget)
            entry_widget.setGeometry(0, 60, 340, 600)
            entry_widget.setStyleSheet("background-color: rgb(187, 228, 252)")
            entry_widget.show()

            form_layout = QFormLayout(entry_widget)
            form_layout.setVerticalSpacing(20)

            # Set Font
            font = QtGui.QFont("Arial", 9)


            lot_label = QLabel()
            lot_label.setText("LOT NUMBER")
            lot_label.setFont(font)

            lot_input = QLineEdit()
            lot_input.setStyleSheet("background-color: rgb(255, 255, 0)")
            lot_input.setAlignment(Qt.AlignCenter)
            lot_input.setFixedHeight(25)
            lot_input.textChanged.connect(autofill)

            quantity_label = QLabel()
            quantity_label.setText("QUANTITY")
            quantity_label.setFont(font)

            quantity_input = QLineEdit()
            quantity_input.setStyleSheet("background-color: rgb(255, 255, 0)")
            quantity_input.setAlignment(Qt.AlignCenter)
            quantity_input.setFixedHeight(25)

            product_code_label = QLabel()
            product_code_label.setText("PRODUCT CODE")
            product_code_label.setFont(font)

            product_code_input = QLineEdit()
            product_code_input.setStyleSheet("background-color: rgb(240, 240, 240)")
            product_code_input.setAlignment(Qt.AlignCenter)
            product_code_input.setFixedHeight(25)
            product_code_input.setEnabled(False)

            formulaID_label = QLabel()
            formulaID_label.setText("FORMULA ID")
            formulaID_label.setFont(font)

            formulaID_input = QLineEdit()
            formulaID_input.setStyleSheet("background-color: rgb(240, 240, 240)")
            formulaID_input.setAlignment(Qt.AlignCenter)
            formulaID_input.setFixedHeight(25)
            formulaID_input.setEnabled(False)

            customer_label = QLabel()
            customer_label.setText("CUSTOMER")
            customer_label.setFont(font)

            customer_input = QLineEdit()
            customer_input.setStyleSheet("background-color: rgb(240, 240, 240)")
            customer_input.setFixedHeight(25)
            customer_input.setEnabled(False)
            customer_input.setAlignment(Qt.AlignCenter)

            remarks_label = QLabel()
            remarks_label.setText("REMARKS")
            remarks_label.setFont(font)
            remarks_label.setAlignment(Qt.AlignRight)

            remarks_input = QTextEdit()
            remarks_input.setStyleSheet("background-color: rgb(255, 255, 0)")
            remarks_input.setFont(font)
            remarks_input.setFixedHeight(160)

            date_label = QLabel()
            date_label.setText("RETURN DATE")
            date_label.setFont(font)
            date_label.setAlignment(Qt.AlignRight)

            date_return = QDateEdit()
            date_return.setStyleSheet("background-color: rgb(255, 255, 0)")
            date_return.setFont(font)
            date_return.setFixedHeight(25)
            today = date.today()
            date_return.setDate(today)
            date_return.setDisplayFormat("MM-dd-yyyy")

            origin_lot_label = QLabel()
            origin_lot_label.setText("ORIGIN LOT")
            origin_lot_label.setFont(font)
            origin_lot_label.setAlignment(Qt.AlignRight)

            origin_lot = QLineEdit()
            origin_lot.setStyleSheet("background-color: rgb(240, 240, 240)")
            origin_lot.setFixedHeight(25)
            origin_lot.setEnabled(False)
            origin_lot.setAlignment(Qt.AlignCenter)

            form_layout.addRow(lot_label,lot_input)
            form_layout.addRow(quantity_label, quantity_input)
            form_layout.addRow(product_code_label, product_code_input)
            form_layout.addRow(formulaID_label, formulaID_input)
            form_layout.addRow(customer_label, customer_input)
            form_layout.addRow(origin_lot_label, origin_lot)
            form_layout.addRow(date_label, date_return)
            form_layout.addRow(remarks_label, remarks_input)

            # Create Return Table

            returns_table = QTableWidget(self.body_widget)
            returns_table.setGeometry(340, 60, 651, 600)
            returns_table.setColumnCount(7)
            returns_table.setRowCount(20)
            returns_table.verticalHeader().setVisible(False)
            returns_table.setHorizontalHeaderLabels(["Lot Number", "Quantity", "Product Code", "Customer", "Formula ID",
                                                     "Return Date", "Remarks"])
            returns_table.setColumnWidth(0, 120)
            returns_table.setColumnWidth(2, 120)
            returns_table.setColumnWidth(3, 200)
            returns_table.setColumnWidth(4, 115)
            returns_table.setColumnWidth(5, 200)
            returns_table.setStyleSheet("gridline-color: rgb(138, 199, 235); border: 1px solid black")
            returns_table.horizontalHeader().setStyleSheet("""
            QHeaderView::section{
            font-weight: bold;
            background-color: rgb(187, 228, 252);
            color: black;
                }  
            
            """)
            show_table()
            returns_table.show()

            save_button = QPushButton(self.body_widget)
            save_button.setGeometry(50, 680, 60, 25)
            save_button.setText("Save")
            save_button.setStyleSheet("background-color: rgb(150, 147, 147); ")
            save_button.setCursor(Qt.PointingHandCursor)
            save_button.clicked.connect(insert_entry)
            save_button.setShortcut("Return")
            save_button.show()

            clear_button = QPushButton(self.body_widget)
            clear_button.setGeometry(150, 680, 60, 25)
            clear_button.setText("Clear")
            clear_button.setStyleSheet("background-color: rgb(150, 147, 147); ")
            clear_button.setCursor(Qt.PointingHandCursor)
            clear_button.clicked.connect(clear_entry)

            edit_button = QPushButton(self.body_widget)
            edit_button.setGeometry(250, 680, 60, 25)
            edit_button.setText("Edit")
            edit_button.setStyleSheet("background-color: rgb(150, 147, 147);")
            edit_button.show()

            clear_button.show()

            # Get the Lot List
            self.cursor.execute("""
            SELECT lot_number FROM quality_control_tbl2
            
            """)
            result = self.cursor.fetchall()
            lot_list = []
            for i in result:
                lot_list.append(i[0])

        def update_entry():
            def save_update():

                self.cursor.execute(f"""
                UPDATE quality_control 
                SET customer = '{customer_list.currentText()}', formula_id = '{formulaID_input.text()}',
                evaluated_on = '{date_started.text()}', status = '{test_result_dropdown.currentText()}',
                remarks = '{remarks_box.toPlainText()}'
                WHERE lot_number = '{lot_number}'
                
                
                """)

                self.conn.commit()
                self.update_qc_widget.close()
                QMessageBox.information(self.qc_widget, "Success", "Update Successful")

            selected = self.qc_table.selectedItems()
            lot_number = selected[1].text()

            self.cursor.execute(f"""
            SELECT * FROM quality_control
            WHERE lot_number = '{lot_number}'
            
            """)
            result = self.cursor.fetchone()

            lot_number = result[1].strip()
            product_code = result[2].strip()
            customer = result[3]
            status = result[4]
            remarks = result[5].strip()
            date_evaluated = result[9]
            formula_id = result[-1]



            if selected:
                self.update_qc_widget = QWidget()
                self.update_qc_widget.setGeometry(0, 0, 350, 550)
                self.update_qc_widget.setFixedSize(350, 550)
                self.update_qc_widget.setWindowModality(Qt.ApplicationModal)
                self.update_qc_widget.setStyleSheet("background-color: ")
                self.update_qc_widget.setWindowTitle("Edit Entry")

                widget1 = QWidget(self.update_qc_widget)
                widget1.setGeometry(0, 0, 350, 500)

                form_layout = QFormLayout(widget1)

                font = QtGui.QFont("Arial", 11)

                lotnumber_label = QLabel()
                lotnumber_label.setText("Lot Number")
                lotnumber_label.setFixedHeight(30)
                lotnumber_label.setFixedWidth(100)
                lotnumber_label.setFont(font)
                lotnumber_label.setStyleSheet("background-color: rgb(240, 240, 240)")

                lotnumber_input = QLineEdit()
                lotnumber_input.setFixedHeight(30)
                lotnumber_input.setText(lot_number)
                lotnumber_input.setEnabled(False)

                customer_label = QLabel()
                customer_label.setText("Customer")
                customer_label.setFont(font)
                customer_label.setFixedWidth(100)

                customer_list = QComboBox()
                customer_list.setFixedHeight(30)


                self.cursor.execute("""
                            SELECT customers FROM customer
                            ORDER BY customers
                            """)
                result = self.cursor.fetchall()

                for i in result:
                    customer_list.addItem(i[0])
                customer_list.setCurrentText(customer)

                productCode_label = QLabel()
                productCode_label.setText("Product Code")
                productCode_label.setFont(font)
                productCode_label.setFixedWidth(100)

                productCode_input = QLineEdit()
                productCode_input.setFixedHeight(30)
                productCode_input.setText(product_code)
                productCode_input.setEnabled(False)

                formulaID_label = QLabel()
                formulaID_label.setText("Formula ID")
                formulaID_label.setFixedWidth(100)
                formulaID_label.setFont(font)

                formulaID_input = QLineEdit()
                formulaID_input.setFixedHeight(30)
                formulaID_input.setText(str(formula_id))

                date_started_label = QLabel()
                date_started_label.setText("Date Evaluated")
                date_started_label.setFont(font)

                date_started = QDateTimeEdit()
                date_started.setDisplayFormat("MM-dd-yyyy HH:mm")
                date_started.setFixedHeight(30)
                date_started.setDateTime(date_evaluated)

                test_result_label = QLabel()
                test_result_label.setText("Test Result")
                test_result_label.setFont(font)
                test_result_label.setFixedWidth(100)

                test_result_dropdown = QComboBox()
                test_result_dropdown.addItem("Passed")
                test_result_dropdown.addItem("Failed")
                test_result_dropdown.setFixedHeight(30)
                test_result_dropdown.setCurrentText(status)

                remarks_label = QLabel()
                remarks_label.setText("Remarks")
                remarks_label.setFont(font)
                remarks_label.setFixedWidth(100)

                remarks_box = QTextEdit()
                remarks_box.setFixedHeight(120)
                remarks_box.setText(remarks)

                form_layout.addRow(lotnumber_label, lotnumber_input)
                form_layout.addRow(customer_label, customer_list)
                form_layout.addRow(productCode_label, productCode_input)
                form_layout.addRow(formulaID_label, formulaID_input)
                form_layout.addRow(date_started_label, date_started)
                form_layout.addRow(test_result_label, test_result_dropdown)
                form_layout.addRow(remarks_label, remarks_box)

                save_button = QPushButton(self.update_qc_widget)
                save_button.setGeometry(100, 500, 60, 25)
                save_button.setText("Save")
                save_button.clicked.connect(save_update)
                save_button.show()

                cancel_button = QPushButton(self.update_qc_widget)
                cancel_button.setGeometry(200, 500, 60, 25)
                cancel_button.setText("Cancel")
                cancel_button.show()

                self.update_qc_widget.show()

            else:
                QMessageBox.information(self.qc_widget, "No Selected", "No Item Selected")


        self.qc_widget = QtWidgets.QWidget(self.main_widget)
        self.qc_widget.setGeometry(0, 0, 991, 751)
        self.qc_widget.setStyleSheet("background-color: rgb(240,240,240);")
        self.qc_widget.show()

        self.qcBtn_topBorder = QtWidgets.QWidget(self.qc_widget)
        self.qcBtn_topBorder.setGeometry(0, 0, 991, 30)
        self.qcBtn_topBorder.show()

        self.qc_topBorder = QtWidgets.QWidget(self.qc_widget)
        self.qc_topBorder.setGeometry(0, 30, 991, 60)
        self.qc_topBorder.setStyleSheet("border-top: 1px solid black; background-color: rgb(239, 243, 254)")
        self.qc_topBorder.show()

        self.qc_table = QtWidgets.QTableWidget(self.qc_widget)
        self.qc_table.setGeometry(0, 90, 991, 350)
        self.qc_table.setColumnCount(7)
        self.qc_table.setRowCount(30)

        # Set Column Width
        self.qc_table.setColumnWidth(0, 80)
        self.qc_table.setColumnWidth(1, 120)
        self.qc_table.setColumnWidth(2, 300)
        self.qc_table.setColumnWidth(3, 120)
        self.qc_table.setColumnWidth(4, 80)
        self.qc_table.setColumnWidth(5, 170)

        self.qc_table.verticalHeader().setVisible(False)
        self.qc_table.setHorizontalHeaderLabels(["ID", "Lot Number", "Customer", "Product Code", "Status", "Remarks", "Action Taken"])

        self.qc_TableBtn = QtWidgets.QPushButton(self.qcBtn_topBorder)
        self.qc_TableBtn.setGeometry(0, 0, 150, 30)
        self.qc_TableBtn.setText("Evaluated Products")
        self.qc_TableBtn.setCursor(Qt.PointingHandCursor)
        self.qc_TableBtn.setFont(QtGui.QFont("Arial", 11))
        self.qc_TableBtn.setStyleSheet("color: rgb(0,109,189); border: 1px solid rgb(160, 160, 160);")
        self.qc_TableBtn.clicked.connect(self.quality_control)
        self.qc_TableBtn.setShortcut("F1")

        # Change the Row height of the table
        for i in range(self.qc_table.rowCount()):
            self.qc_table.setRowHeight(i, 22)

        table_header = self.qc_table.horizontalHeader()
        table_header.setFixedHeight(25)

        # Get the table Items from database
        self.cursor.execute("""
        SELECT id, lot_number, customer, product_code, status, remarks, action
        FROM quality_control
        ORDER BY id DESC
        
        """)

        result = self.cursor.fetchall()

        # Populate the table
        for i in range(len(result)):
            for j in range(len(result[i])):
                item = QTableWidgetItem(str(result[i][j]))
                item.setTextAlignment(Qt.AlignCenter)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                self.qc_table.setItem(i, j, item)

        self.qc_table.itemSelectionChanged.connect(show_items)

        self.qc_TableBtn.show()

        self.qc_addEntryBtn = QtWidgets.QPushButton(self.qcBtn_topBorder)
        self.qc_addEntryBtn.setGeometry(150, 0, 150, 30)
        self.qc_addEntryBtn.setText("Evaluation Entry")
        self.qc_addEntryBtn.setCursor(Qt.PointingHandCursor)
        self.qc_addEntryBtn.setStyleSheet("border: 1px solid rgb(160, 160, 160);  text-align: top;")
        self.qc_addEntryBtn.clicked.connect(evaluation_entry)
        self.qc_addEntryBtn.setFont(QtGui.QFont("Arial", 11))
        self.qc_addEntryBtn.setShortcut("F2")
        self.qc_addEntryBtn.show()

        self.qc_dataBtn = QtWidgets.QPushButton(self.qcBtn_topBorder)
        self.qc_dataBtn.setGeometry(300, 0, 150, 30)
        self.qc_dataBtn.setText("QC Data")
        self.qc_dataBtn.setCursor(Qt.PointingHandCursor)
        self.qc_dataBtn.setStyleSheet("border: 1px solid rgb(160, 160, 160); text-align: top;")
        self.qc_dataBtn.clicked.connect(show_qc_data)
        self.qc_dataBtn.setFont(QtGui.QFont("Arial", 11))
        self.qc_dataBtn.setShortcut("F3")
        self.qc_dataBtn.show()

        self.dashboardBtn = QtWidgets.QPushButton(self.qcBtn_topBorder)
        self.dashboardBtn.setGeometry(450, 0, 150, 30)
        self.dashboardBtn.setText("Dashboard")
        self.dashboardBtn.setCursor(Qt.PointingHandCursor)
        self.dashboardBtn.setStyleSheet("border: 1px solid rgb(160, 160, 160); text-align: top;")
        self.dashboardBtn.clicked.connect(show_dashboards)
        self.dashboardBtn.setFont(QtGui.QFont("Arial", 11))
        self.dashboardBtn.setShortcut("F4")
        self.dashboardBtn.show()

        self.qc_returns  = QtWidgets.QPushButton(self.qcBtn_topBorder)
        self.qc_returns.setGeometry(600, 0, 150, 30)
        self.qc_returns.setText("Returns")
        self.qc_returns.setCursor(Qt.PointingHandCursor)
        self.qc_returns.setStyleSheet("border: 1px solid rgb(160, 160, 160); text-align: top;")
        self.qc_returns.setFont(QtGui.QFont("Arial", 11))
        self.qc_returns.clicked.connect(qc_returns)
        self.qc_returns.setShortcut("F5")
        self.qc_returns.show()

        # Top Border Widgets
        evaluation_lbl = QLabel(self.qc_topBorder)
        evaluation_lbl.setGeometry(10, 30, 100, 30)
        evaluation_lbl.setFont(QtGui.QFont("Arial", 9))
        evaluation_lbl.setText("Evaluation For:")
        evaluation_lbl.setStyleSheet("border: none;")
        evaluation_lbl.show()

        evaluation_text = QLabel(self.qc_topBorder)
        evaluation_text.setGeometry(120, 30, 500, 30)
        evaluation_text.setStyleSheet("border: none;")
        evaluation_text.show()

        loadAll_checkbox = QCheckBox(self.qc_topBorder)
        loadAll_checkbox.move(5, 5)
        loadAll_checkbox.setStyleSheet("border: none")
        loadAll_checkbox.show()

        loadAll_label = QLabel(self.qc_topBorder)
        loadAll_label.setGeometry(20, 4, 80, 15)
        loadAll_label.setStyleSheet("border: none;")
        loadAll_label.setText("LOAD ALL DATA")
        loadAll_label.setFont(QtGui.QFont("Arial", 8))
        loadAll_label.show()

        edited_checkbox = QCheckBox(self.qc_topBorder)
        edited_checkbox.move(105, 5)
        edited_checkbox.setStyleSheet("border: none;")
        edited_checkbox.stateChanged.connect(edited_checkbox_changed)
        edited_checkbox.show()

        edited_label = QLabel(self.qc_topBorder)
        edited_label.setGeometry(120, 4, 90, 15)
        edited_label.setStyleSheet("border: none;")
        edited_label.setText("EDITED RECORDS")
        edited_label.setFont(QtGui.QFont("Arial", 8))
        edited_label.show()

        search_bar = QtWidgets.QLineEdit(self.qc_topBorder)
        search_bar.setGeometry(770, 5, 150, 25)
        search_bar.setStyleSheet("border: 1px solid rgb(171, 173, 179); background-color: rgb(255, 255, 17);")
        search_bar.setFont(QtGui.QFont("Arial", 9))
        search_bar.setPlaceholderText("Lot Number")
        search_bar.show()

        search_btn = QtWidgets.QPushButton(self.qc_topBorder)
        search_btn.setGeometry(925, 5, 60, 25)
        search_btn.setStyleSheet("border: 1px solid rgb(171, 173, 179);")
        search_btn.setText("Search")
        search_btn.clicked.connect(lotNumber_search)
        search_btn.show()

        # Bottom Widgets
        bottom_widget = QtWidgets.QWidget(self.qc_widget)
        bottom_widget.setGeometry(0, 440, 991, 311)
        bottom_widget.setStyleSheet("background-color : rgb(239, 243, 254)")
        bottom_widget.show()

        leftSide1_widget = QtWidgets.QWidget(bottom_widget)
        leftSide1_widget.setGeometry(0, 0, 140, 225)
        leftSide1_widget.show()

        leftSide2_widget = QtWidgets.QWidget(bottom_widget)
        leftSide2_widget.setGeometry(140, 0, 455, 225)
        leftSide2_widget.show()

        rightSide_widget = QtWidgets.QWidget(bottom_widget)
        rightSide_widget.setGeometry(595, 0, 130, 114)
        rightSide_widget.show()


        label_font = QtGui.QFont("Segoe UI", 11)

        # Create Vertical Box layout

        leftVBox1_layout = QVBoxLayout(leftSide1_widget)
        leftVBox2_layout = QVBoxLayout(leftSide2_widget)
        rightVBo1_layout = QVBoxLayout(rightSide_widget)

        # Left Side Labels
        customer_label = QLabel()
        customer_label.setText("Customer            :")
        customer_label.setFont(label_font)

        productCode_label = QLabel()
        productCode_label.setText("Product Code     :")
        productCode_label.setFont(label_font)

        result_label = QLabel()
        result_label.setText("Result                 :")
        result_label.setFont(label_font)

        evaluatedBy_label = QLabel()
        evaluatedBy_label.setText("Evaluated By      :")
        evaluatedBy_label.setFont(label_font)

        evaluatedDate_label = QLabel()
        evaluatedDate_label.setText("Evaluated On     :")
        evaluatedDate_label.setFont(label_font)

        encodedDate_label = QLabel()
        encodedDate_label.setText("Encoded On       :")
        encodedDate_label.setFont(label_font)

        # Left Side Outputs
        customer_selected = QLabel()
        productCode_selected = QLabel()
        result_selected = QLabel()
        evaluatedBy_selected = QLabel()
        evaluatedDate_selected = QLabel()
        encodedDate_selected = QLabel()


        # Right Side Labels
        updatedBy_label = QLabel()
        updatedBy_label.setText("Updated By :")
        updatedBy_label.setAlignment(Qt.AlignRight)
        updatedBy_label.setFont(label_font)

        remarks_label = QLabel(bottom_widget)
        remarks_label.setText("Remarks :")
        remarks_label.setAlignment(Qt.AlignRight)
        remarks_label.setFont(label_font)
        remarks_label.setGeometry(612, 110, 100, 30)
        remarks_label.show()

        remarks_box = QTextEdit(bottom_widget)
        remarks_box.setGeometry(722, 110, 200, 80)
        remarks_box.setStyleSheet("background-color: rgb(227, 227, 227)")
        remarks_box.setEnabled(False)
        remarks_box.show()

        actionTaken_label = QLabel(bottom_widget)
        actionTaken_label.setText("Action Taken :")
        actionTaken_label.setAlignment(Qt.AlignRight)
        actionTaken_label.setFont(label_font)
        actionTaken_label.setGeometry(612, 200, 100, 30)
        actionTaken_label.show()

        actionTake_box = QTextEdit(bottom_widget)
        actionTake_box.setGeometry(722, 200, 200, 60)
        actionTake_box.setStyleSheet("background-color: rgb(227, 227, 227)")
        actionTake_box.setEnabled(False)
        actionTake_box.show()

        time_endorsed = QLabel()
        time_endorsed.setText("Time Endorsed :")
        time_endorsed.setAlignment(Qt.AlignRight)
        time_endorsed.setFont(label_font)

        time_endorsed_val = QLabel(self.qc_widget)
        time_endorsed_val.setGeometry(721, 483, 120, 25)
        time_endorsed_val.setFont(QtGui.QFont("Arial", 9))
        time_endorsed_val.setStyleSheet("background-color: rgb(239, 243, 254)")
        time_endorsed_val.show()

        qc_type_label = QLabel()
        qc_type_label.setText("QC Type : ")
        qc_type_label.setAlignment(Qt.AlignRight)
        qc_type_label.setFont(label_font)

        qc_type_selected = QLabel(self.qc_widget)
        qc_type_selected.setGeometry(721, 515, 120, 25)
        qc_type_selected.setStyleSheet("background-color: rgb(239, 243, 254)")
        qc_type_selected.show()

        # Adding the widgets to the Layout
        leftVBox1_layout.addWidget(customer_label)
        leftVBox1_layout.addWidget(productCode_label)
        leftVBox1_layout.addWidget(result_label)
        leftVBox1_layout.addWidget(evaluatedBy_label)
        leftVBox1_layout.addWidget(evaluatedDate_label)
        leftVBox1_layout.addWidget(encodedDate_label)

        leftVBox2_layout.addWidget(customer_selected)
        leftVBox2_layout.addWidget(productCode_selected)
        leftVBox2_layout.addWidget(result_selected)
        leftVBox2_layout.addWidget(evaluatedBy_selected)
        leftVBox2_layout.addWidget(evaluatedDate_selected)
        leftVBox2_layout.addWidget(encodedDate_selected)

        rightVBo1_layout.addWidget(updatedBy_label)
        rightVBo1_layout.addWidget(time_endorsed)
        rightVBo1_layout.addWidget(qc_type_label)

        updatedBy_val1 = QtWidgets.QLineEdit(self.qc_widget)
        updatedBy_val1.setGeometry(721, 445, 130, 25)
        updatedBy_val1.setStyleSheet("background-color: rgb(227, 227, 227)")
        updatedBy_val1.setEnabled(False)
        updatedBy_val1.show()

        updatedBy_val2 = QtWidgets.QLineEdit(self.qc_widget)
        updatedBy_val2.setGeometry(851, 445, 130, 25)
        updatedBy_val2.setEnabled(False)
        updatedBy_val2.setStyleSheet("background-color: rgb(227, 227, 227)")
        updatedBy_val2.show()

        export_label = QLabel(self.qc_widget)
        export_label.setGeometry(10, 680, 100, 20)
        export_label.setText("Export Database")
        export_label.setStyleSheet("background-color: rgb(239, 243, 254);")
        export_label.show()

        label1 = QLabel(self.qc_widget)
        label1.setText("FROM")
        label1.setGeometry(30, 700, 70, 25)
        label1.setStyleSheet("background-color: rgb(239, 243, 254)")
        label1.setFont(label_font)
        label1.show()

        date1 = QDateEdit(self.qc_widget)
        date1.setGeometry(75, 700, 100, 25)
        date1.setStyleSheet("background-color: rgb(239, 243, 254)")
        date1.setDisplayFormat("yyyy-MM-dd")
        date1.show()

        label2 = QLabel(self.qc_widget)
        label2.setGeometry(185, 700, 30, 25)
        label2.setText("TO")
        label2.setFont(label_font)
        label2.setStyleSheet("background-color: rgb(239, 243, 254)")
        label2.show()

        now = QtCore.QDate.currentDate() # get the current date

        date2 = QDateEdit(self.qc_widget)
        date2.setGeometry(215, 700, 100, 25)
        date2.setStyleSheet("background-color: rgb(239, 243, 254)")
        date2.setDisplayFormat("yyyy-MM-dd")
        date2.setDate(now)
        date2.show()

        export_btn = ClickableLabel(self.qc_widget)
        export_btn.setGeometry(320, 703, 20, 20)
        export_btn.setPixmap(QtGui.QIcon('export.png').pixmap(20, 20))
        export_btn.setCursor(Qt.PointingHandCursor)
        export_btn.clicked.connect(exportBtn_clicked)
        export_btn.setToolTip("Export")
        export_btn.show()

        update_btn = QtWidgets.QPushButton(self.qc_widget)
        update_btn.setGeometry(650, 703, 60, 25)
        update_btn.setText("UPDATE")
        update_btn.clicked.connect(update_entry)
        update_btn.show()

        delete_btn = QtWidgets.QPushButton(self.qc_widget)
        delete_btn.setGeometry(710, 703, 60, 25)
        delete_btn.setText("DELETE")
        delete_btn.show()

        self.qc_table.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)
        self.qc_table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.qc_table.show()


if __name__ == "__main__":
    import sys

    app = QtWidgets.QApplication(sys.argv)
    LoginWindow = QtWidgets.QMainWindow()
    ui = Ui_LoginWindow()
    ui.setupUi(LoginWindow)
    LoginWindow.show()
    sys.exit(app.exec_())
