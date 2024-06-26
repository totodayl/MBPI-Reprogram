import datetime
from openpyxl import Workbook, load_workbook
import json
import psycopg2
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import Qt, pyqtSignal
from PyQt5.QtWidgets import *
from datetime import timedelta, datetime, time

# For Clickable Icons
class ClickableLabel(QtWidgets.QLabel):
    clicked = pyqtSignal()

    def mousePressEvent(self, event):
        self.clicked.emit()

class Ui_LoginWindow(object):
    def setupUi(self, LoginWindow):
        LoginWindow.setObjectName("LoginWindow")
        LoginWindow.resize(800, 600)
        self.login_window = QtWidgets.QWidget(LoginWindow)
        self.login_window.setStyleSheet("""background-color : qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:1, 
        stop:0 rgba(176, 0, 0, 255), stop:0.738636 rgba(255, 113, 250, 255))""")
        self.login_window.setObjectName("MainWindow")
        self.username = QtWidgets.QLineEdit(self.login_window)
        self.username.setGeometry(QtCore.QRect(310, 140, 171, 31))
        self.username.setAutoFillBackground(False)
        self.username.setStyleSheet("background-color: rgb(255, 255, 255); border-radius: 5px")
        self.username.setObjectName("username")
        self.username.setText("postgres")
        self.password = QtWidgets.QLineEdit(self.login_window)
        self.password.setGeometry(QtCore.QRect(310, 230, 171, 31))
        self.password.setAutoFillBackground(False)
        self.password.setStyleSheet("background-color: rgb(255, 255, 255); border-radius: 5px")
        self.password.setObjectName("password")
        self.password.setEchoMode(QtWidgets.QLineEdit.Password)  # Make the input hidden
        self.password.setText("mbpi")
        self.login_btn = QtWidgets.QPushButton(self.login_window)
        self.login_btn.setGeometry(QtCore.QRect(360, 340, 75, 23))
        self.login_btn.setStyleSheet("\n"
                                     "color: rgb(0, 0, 0);\n"
                                     "background-color: rgb(255, 255, 255); \n"
                                     "border-radius: 10px;")
        self.login_btn.setObjectName("Login")
        self.login_btn.clicked.connect(self.login)
        LoginWindow.setCentralWidget(self.login_window)
        self.retranslateUi(LoginWindow)
        QtCore.QMetaObject.connectSlotsByName(LoginWindow)

    def retranslateUi(self, LoginWindow):
        _translate = QtCore.QCoreApplication.translate
        LoginWindow.setWindowTitle(_translate("LoginWindow", "MBPI"))
        self.login_btn.setText(_translate("LoginWindow", "Login"))

    def login(self):
        username = self.username.text()
        pass1 = self.password.text()

        # Connect to the Database
        try:
            self.conn = psycopg2.connect(
                host="192.168.1.13",
                port=5432,
                dbname='postgres',
                user=f'{username}',
                password=f'{pass1}'
            )

            self.username.deleteLater()
            self.password.deleteLater()
            self.login_btn.deleteLater()
        except psycopg2.Error:
            QMessageBox.information(self.login_window, "INVALID CREDENTIALS", "Check your Username and Password")
            return
        self.cursor = self.conn.cursor()
        self.launch_main()

    # This is the main window after login screen
    def launch_main(self):
        LoginWindow.move(37, 100)
        LoginWindow.setFixedSize(1200, 750)
        self.login_window.setStyleSheet("background-color: rgb(60,60,60);")
        self.main_widget = QtWidgets.QWidget(self.login_window)
        self.main_widget.setStyleSheet("""
        background-color: rgb(240,240,240);     
        
        """)
        self.main_widget.setGeometry(210, 0, 991, 751)
        self.main_widget.show()

        self.production_btn = QtWidgets.QPushButton(self.login_window)
        self.production_btn.setGeometry(30, 200, 180, 40)
        self.production_btn.setCursor(Qt.PointingHandCursor)
        self.production_btn.setStyleSheet("""
        border-top-left-radius: 10px; 
        border-bottom-left-radius: 10px; 
        background-color: rgb(125,125,125);
        """)
        self.production_btn.clicked.connect(self.production)
        self.production_btn.show()

        self.production_lbl = QtWidgets.QLabel(self.production_btn)
        self.production_lbl.setText("Extruder")
        self.production_lbl.setGeometry(50, 5, 100, 30)
        self.production_lbl.setFont(QtGui.QFont("Arial", 13))
        self.production_lbl.setStyleSheet("color: blue;")
        self.production_lbl.setCursor(Qt.PointingHandCursor)
        self.production_lbl.show()

        self.production_icon = ClickableLabel(self.production_btn)
        self.production_icon.setGeometry(10, 3, 30, 30)
        self.production_icon.setPixmap(QtGui.QIcon('setting.png').pixmap(30, 30))  # Set icon
        self.production_icon.setScaledContents(True)  # Scale icon to fit the label
        self.production_icon.setCursor(Qt.PointingHandCursor)
        self.production_icon.show()

    def production(self):

        # Delete If there are existing Widgets
        try:
            self.info_widget.deleteLater()
            self.temp_table.deleteLater()
            self.time_table.deleteLater()
            self.material_table.deleteLater()
            self.group_box.deleteLater()
            self.export_btn.deleteLater()


        except Exception as e:
            print(e)

        def show_form():


            try:
                # Clear all the widget first
                self.extruder_table.setVisible(False)
                self.view_btn.setVisible(False)
                self.add_btn.setVisible(False)
                self.update_btn.setVisible(False)
            except:
                pass

            try:
                selected = self.extruder_table.selectedItems()
                selected = [i.text() for i in selected]

                # Query the whole columns
                self.cursor.execute(f"SELECT * FROM extruder WHERE process_id = {selected[0]}")
                selected = self.cursor.fetchall()
            except Exception as e:
                print(e)
                QMessageBox.information(self.main_widget, "ERROR", "No Selected Item")
                return

            selected = selected[0]

            # Unpack all the items for convenience
            machine, ordered_qty, product_output, customer = selected[1:5]
            formula_id, product_code, order_id, total_time, time_start = selected[5:10]
            time_end, output_percent, loss, loss_percent, purging = selected[10:15]
            resin, remarks, screw_config, feed_rate, rpm = selected[15:20]
            screen_size, operator, supervisor, materials, temperature = selected[20:25]
            purge_duration, outputs, output_per_hour = selected[25:28]
            production_ID = selected[28]
            total_input = selected[29]

            def exportToExcel():
                from openpyxl.styles import Font
                from openpyxl.styles import Alignment


                process_id = selected[0]

                self.cursor.execute(f"""
                           SELECT * FROM extruder 
                           WHERE process_id = '{process_id}';

                           """)
                items = self.cursor.fetchall()[0]
                # Unpack the Items
                machine_number = items[1]
                quantity_order = items[2]
                customer = items[4]
                code = items[6]
                total_input = items[-2]
                total_time = items[8]
                time_start = items[9]
                time_end = items[10]
                outputPerHour = items[27]
                total_output = items[3]
                outputPercent = items[11]
                loss = items[12]
                lossPercent = items[13]
                purging = items[14]
                resin = items[15]
                remarks = items[16]
                operator = items[21]
                supervisor = items[22]
                outputs = items[-5]
                materials = items[-8]
                lot_number = items[-1][0]
                purge_duration = time(hour=items[-6])

                wb = load_workbook(
                    r"\\mbpi-server-01\IT\AMIEL\Extruder System\dist")
                worksheet = wb.active

                font = Font(size=8, bold=True, name='Arial')
                center_Alignment = Alignment(horizontal='center', vertical='center')

                worksheet["F5"] = "Extruder Machine No. " + machine_number[-1]
                worksheet["A8"] = machine_number[-1]
                worksheet["B8"] = quantity_order  # quantity order
                worksheet["C8"].font = font
                worksheet["C8"].alignment = center_Alignment
                worksheet["C8"] = customer  # customer

                worksheet["F8"] = code  # product code
                worksheet["G9"] = total_input  # total input
                worksheet["H9"] = total_time  # total time used
                worksheet["I9"] = outputPerHour  # output Per Hour
                worksheet["K9"] = total_output  # total Output
                worksheet["L9"] = outputPercent  # Total Output Percentage
                worksheet["M9"] = loss
                worksheet["N9"] = lossPercent

                total_sec = timedelta()
                for row in range(len(time_start)):
                    worksheet["A" + str(12 + row)] = time_start[row].strftime("%b-%d-%Y %H:%M")
                    worksheet["D" + str(12 + row)] = time_end[row].strftime("%b-%d-%Y %H:%M")
                    worksheet["F" + str(12 + row)] = time_end[row] - time_start[row]
                    worksheet["G" + str(12 + row)] = outputs[row]
                    total_sec = total_sec + (time_end[row] - time_start[row])

                try:
                    hour = str(int(total_sec.total_seconds() // 3600))
                    minute = str((int(total_sec.total_seconds() % 3600) // 60))
                    print(hour, minute)
                    total_time_used = time(int(hour), int(minute))

                    worksheet["F25"] = total_time_used
                except ValueError:
                    worksheet["F25"] = hour + ":" + minute

                for key in list(materials.keys()):
                    worksheet["I" + str(12 + list(materials.keys()).index(key))] = key
                    worksheet["K" + str(12 + list(materials.keys()).index(key))] = materials[key]

                for ln in range(len(lot_number)):
                    worksheet["M" + str(12 + ln)] = lot_number[ln]

                worksheet["B27"] = purging
                worksheet["E28"] = purge_duration
                worksheet["B29"] = resin
                worksheet["G26"] = remarks
                worksheet["M28"] = operator
                worksheet["M29"] = supervisor

                wb.save("text.xlsx")
                print("load successful")

            # Convert string of json to JSON
            materials = str(materials).replace("'", '"')
            materials = json.loads(materials)

            # Main Widget
            self.info_widget = QtWidgets.QWidget(self.main_widget)
            self.info_widget.setGeometry(20, 0, 951, 450)
            self.info_widget.setStyleSheet("background-color: rgb(0,109,189);")
            self.info_widget.show()

            # Set Font Style and Size
            label_stylesheet = "color: rgb(195, 164, 86)"
            font = QtGui.QFont("Arial", 14)  # Set Font for Labels

            # Create 3 widgets for division
            info_widget1 = QtWidgets.QWidget(self.info_widget)
            info_widget1.setGeometry(0, 80, 158, 380)

            info_widget2 = QtWidgets.QWidget(self.info_widget)
            info_widget2.setGeometry(158, 80, 158, 380)

            info_widget3 = QtWidgets.QWidget(self.info_widget)
            info_widget3.setGeometry(316, 80, 158, 380)

            info_widget4 = QtWidgets.QWidget(self.info_widget)
            info_widget4.setGeometry(474, 80, 158, 380)

            info_widget5 = QtWidgets.QWidget(self.info_widget)
            info_widget5.setGeometry(632, 80, 158, 380)

            info_widget6 = QtWidgets.QWidget(self.info_widget)
            info_widget6.setGeometry(790, 80, 158, 380)

            # create Vertical Layouts
            left_vertical_layout = QVBoxLayout()

            self.ordered_company = QtWidgets.QLabel(self.info_widget)
            self.ordered_company.setFont(QtGui.QFont("Arial", 30))
            self.ordered_company.setGeometry(0, 30, 950, 50)
            self.ordered_company.setText(customer)
            self.ordered_company.setAlignment(Qt.AlignCenter)
            self.ordered_company.show()

            # Extruder Label
            self.machine_label = QtWidgets.QLabel(self.info_widget)
            self.machine_label.setText("Extruder:")
            self.machine_label.setStyleSheet(label_stylesheet)
            self.machine_label.setFont(font)

            # Show Extruder Value
            self.extruder_val = QtWidgets.QLabel(self.info_widget)
            self.extruder_val.setText(machine)
            self.extruder_val.setFont(font)

            # Product Code Label
            self.code_label = QtWidgets.QLabel(self.info_widget)
            self.code_label.setText("Product Code:")
            self.code_label.setFont(font)
            self.code_label.setStyleSheet(label_stylesheet)

            # Show Product Code Value
            self.product_code_val = QtWidgets.QLabel(self.info_widget)
            self.product_code_val.setText(product_code)
            self.product_code_val.setFont(font)

            # Quantity Order Label
            self.order_label = QtWidgets.QLabel(self.info_widget)
            self.order_label.setText("Quantity Order:")
            self.order_label.setFont(font)
            self.order_label.setStyleSheet(label_stylesheet)

            # Show Order Value
            self.order_val = QtWidgets.QLabel(self.info_widget)
            self.order_val.setText(str(ordered_qty))
            self.order_val.setFont(font)

            # Show Output Label
            self.output_label = QtWidgets.QLabel(self.info_widget)
            self.output_label.setText("Output:")
            self.output_label.setFont(font)
            self.output_label.setStyleSheet(label_stylesheet)

            # Show Output Value
            self.output_val = QtWidgets.QLabel(self.info_widget)
            self.output_val.setText(str(product_output))
            self.output_val.setFont(font)

            # Show formula Label
            self.formula_label = QtWidgets.QLabel(self.info_widget)
            self.formula_label.setText("Formula ID:")
            self.formula_label.setFont(font)
            self.formula_label.setStyleSheet(label_stylesheet)

            # Show Formula ID Value
            self.formulaID_val = QtWidgets.QLabel(self.info_widget)
            self.formulaID_val.setText(str(formula_id))
            self.formulaID_val.setFont(font)

            self.resin_label = QtWidgets.QLabel(self.info_widget)
            self.resin_label.setText("Resin:")
            self.resin_label.setFont(font)
            self.resin_label.setStyleSheet(label_stylesheet)

            # Show Resin Value
            self.resin_val = QtWidgets.QLabel(self.info_widget)
            self.resin_val.setText(str(resin))
            self.resin_val.setFont(font)

            # Lot Label
            self.lot_label = QtWidgets.QLabel(self.info_widget)
            self.lot_label.setText("LOT Number:")
            self.lot_label.setFont(font)
            self.lot_label.setStyleSheet(label_stylesheet)

            # Show Lot Number Value
            self.lotNum_val = QtWidgets.QLabel(self.info_widget)
            self.lotNum_val.setText(str(resin))
            self.lotNum_val.setFont(font)

            self.feedrate_label = QtWidgets.QLabel(self.info_widget)
            self.feedrate_label.setText("Feed Rate:")
            self.feedrate_label.setFont(font)
            self.feedrate_label.setStyleSheet(label_stylesheet)

            # Show Feed Rate Value
            self.feedrate_val = QtWidgets.QLabel(self.info_widget)
            self.feedrate_val.setText(str(feed_rate))
            self.feedrate_val.setFont(font)

            # RPM label
            self.rpm_label = QtWidgets.QLabel(self.info_widget)
            self.rpm_label.setText("RPM:")
            self.rpm_label.setFont(font)
            self.rpm_label.setStyleSheet(label_stylesheet)

            # Show RPM Value
            self.rpm_val = QtWidgets.QLabel(self.info_widget)
            self.rpm_val.setText(str(rpm))
            self.rpm_val.setFont(font)

            # Screen Size Label
            self.screen_size_label = QtWidgets.QLabel(self.info_widget)
            self.screen_size_label.setFont(font)
            self.screen_size_label.setText("Screen Size:")
            self.screen_size_label.setStyleSheet(label_stylesheet)

            # Show Screen Size Value
            self.screenSize_val = QtWidgets.QLabel(self.info_widget)
            self.screenSize_val.setText(str(screen_size))
            self.screenSize_val.setFont(font)

            self.screwconfig_label = QtWidgets.QLabel(self.info_widget)
            self.screwconfig_label.setFont(font)
            self.screwconfig_label.setText("Screw Config:")
            self.screwconfig_label.setStyleSheet(label_stylesheet)

            # Show Screw Config Value
            self.screwConf_val = QtWidgets.QLabel(self.info_widget)
            self.screwConf_val.setText(str(screw_config))
            self.screwConf_val.setFont(font)

            # Output % Label
            self.output_percentage_lbl = QtWidgets.QLabel(self.info_widget)
            self.output_percentage_lbl.setText("Output %:")
            self.output_percentage_lbl.setFont(font)
            self.output_percentage_lbl.setStyleSheet(label_stylesheet)

            # Show Output Percentage Value
            self.outputPercent_val = QtWidgets.QLabel(self.info_widget)
            self.outputPercent_val.setText(str(output_percent))
            self.outputPercent_val.setFont(font)

            # Loss Label
            self.loss_label = QtWidgets.QLabel(self.info_widget)
            self.loss_label.setText("Loss:")
            self.loss_label.setFont(QtGui.QFont(font))
            self.loss_label.setStyleSheet(label_stylesheet)

            # Show Loss Value
            self.loss_val = QtWidgets.QLabel(self.info_widget)
            self.loss_val.setText(str(loss))
            self.loss_val.setFont(font)

            # loss percentage Label
            self.loss_percent_label = QtWidgets.QLabel(self.info_widget)
            self.loss_percent_label.setText("Loss %:")
            self.loss_percent_label.setFont(font)
            self.loss_percent_label.setStyleSheet(label_stylesheet)

            # Show Loss Percentage Value
            self.lossPercent_val = QtWidgets.QLabel(self.info_widget)
            self.lossPercent_val.setText(str(loss_percent))
            self.lossPercent_val.setFont(font)

            # Purge Duration Label
            self.purge_duration_label = QtWidgets.QLabel(self.info_widget)
            self.purge_duration_label.setText("Purge Duration:")
            self.purge_duration_label.setFont(font)
            self.purge_duration_label.setStyleSheet(label_stylesheet)

            # Show Resin Value
            self.purgeDuration_val = QtWidgets.QLabel(self.info_widget)
            self.purgeDuration_val.setText(str(purge_duration))
            self.purgeDuration_val.setFont(font)

            # Production ID
            self.productionID_value = QtWidgets.QLabel(self.info_widget)
            self.productionID_value.setText(str(production_ID))
            self.productionID_value.setFont(font)

            self.productionID_label = QtWidgets.QLabel(self.info_widget)
            self.productionID_label.setText("Production ID:")
            self.productionID_label.setFont(font)
            self.productionID_label.setStyleSheet(label_stylesheet)

            # Order ID
            self.orderID_label = QtWidgets.QLabel(self.info_widget)
            self.orderID_label.setText("Order ID:")
            self.orderID_label.setFont(font)
            self.orderID_label.setStyleSheet(label_stylesheet)

            self.orderID_value = QtWidgets.QLabel(self.info_widget)
            self.orderID_value.setText(str(order_id))
            self.orderID_value.setFont(font)

            # Total Time
            self.total_time = QtWidgets.QLabel(self.info_widget)
            self.total_time.setText("Total Hours:")
            self.total_time.setFont(font)
            self.total_time.setStyleSheet(label_stylesheet)

            self.total_time_value = QtWidgets.QLabel(self.info_widget)
            self.total_time_value.setText(str(total_time))
            self.total_time_value.setFont(font)

            # Output Per Hour
            self.outputPerHour_label = QtWidgets.QLabel(self.info_widget)
            self.outputPerHour_label.setText("Output / Hour:")
            self.outputPerHour_label.setFont(font)
            self.outputPerHour_label.setStyleSheet(label_stylesheet)

            self.outputPerHour_val = QtWidgets.QLabel(self.info_widget)
            self.outputPerHour_val.setText(str(output_per_hour))
            self.outputPerHour_val.setFont(font)

            # Total Input
            self.total_input = QtWidgets.QLabel(self.info_widget)
            self.total_input.setFont(font)
            self.total_input.setText(str(total_input))

            self.totalInput_label = QtWidgets.QLabel(self.info_widget)
            self.totalInput_label.setFont(font)
            self.totalInput_label.setText("Total Input:")
            self.totalInput_label.setStyleSheet(label_stylesheet)

            # Purging Labels
            self.purging_label = QtWidgets.QLabel(self.info_widget)
            self.purging_label.setText("Purging To:")
            self.purging_label.setFont(font)
            self.purging_label.setStyleSheet(label_stylesheet)

            self.purging_val = QtWidgets.QLabel(self.info_widget)
            self.purging_val.setFont(font)
            self.purging_val.setText(purging)

            # Operator Labels
            self.operator_label = QtWidgets.QLabel(self.info_widget)
            self.operator_label.setFont(font)
            self.operator_label.setText("Operator:")
            self.operator_label.setStyleSheet(label_stylesheet)

            self.operator_value = QtWidgets.QLabel(self.info_widget)
            self.operator_value.setFont(font)
            self.operator_value.setText(operator)

            # Supervisor Labels
            self.supervisor_label = QtWidgets.QLabel(self.info_widget)
            self.supervisor_label.setText("Supervisor")
            self.supervisor_label.setFont(font)
            self.supervisor_label.setStyleSheet(label_stylesheet)

            self.supervisor_value = QtWidgets.QLabel(self.info_widget)
            self.supervisor_value.setText(supervisor)
            self.supervisor_value.setFont(font)

            info_vbox1 = QVBoxLayout(info_widget1)
            info_vbox2 = QVBoxLayout(info_widget2)
            info_vbox3 = QVBoxLayout(info_widget3)
            info_vbox4 = QVBoxLayout(info_widget4)
            info_vbox5 = QVBoxLayout(info_widget5)
            info_vbox6 = QVBoxLayout(info_widget6)

            # First VBOX for Labels
            info_vbox1.addWidget(self.machine_label)
            info_vbox1.addWidget(self.productionID_label)
            info_vbox1.addWidget(self.code_label)
            info_vbox1.addWidget(self.orderID_label)
            info_vbox1.addWidget(self.formula_label)
            info_vbox1.addWidget(self.order_label)
            info_vbox1.addWidget(self.total_time)

            # Second VBOX for Inputs
            info_vbox2.addWidget(self.extruder_val)
            info_vbox2.addWidget(self.productionID_value)
            info_vbox2.addWidget(self.product_code_val)
            info_vbox2.addWidget(self.orderID_value)
            info_vbox2.addWidget(self.formulaID_val)
            info_vbox2.addWidget(self.order_val)
            info_vbox2.addWidget(self.total_time_value)

            # third VBOX for Second Label
            info_vbox3.addWidget(self.totalInput_label)
            info_vbox3.addWidget(self.output_label)
            info_vbox3.addWidget(self.output_percentage_lbl)
            info_vbox3.addWidget(self.outputPerHour_label)
            info_vbox3.addWidget(self.loss_label)
            info_vbox3.addWidget(self.loss_percent_label)
            info_vbox3.addWidget(self.feedrate_label)
            info_vbox3.addWidget(self.rpm_label)

            # Fourth Vbox for the Second Row Inputs Value
            info_vbox4.addWidget(self.total_input)
            info_vbox4.addWidget(self.output_val)
            info_vbox4.addWidget(self.outputPercent_val)
            info_vbox4.addWidget(self.outputPerHour_val)
            info_vbox4.addWidget(self.loss_val)
            info_vbox4.addWidget(self.lossPercent_val)
            info_vbox4.addWidget(self.feedrate_val)
            info_vbox4.addWidget(self.rpm_val)

            # Fifth VBOX for Labels
            info_vbox5.addWidget(self.purging_label)
            info_vbox5.addWidget(self.resin_label)
            info_vbox5.addWidget(self.screwconfig_label)
            info_vbox5.addWidget(self.screen_size_label)
            info_vbox5.addWidget(self.purge_duration_label)
            info_vbox5.addWidget(self.operator_label)
            info_vbox5.addWidget(self.supervisor_label)

            info_vbox6.addWidget(self.purging_val)
            info_vbox6.addWidget(self.resin_val)
            info_vbox6.addWidget(self.screwConf_val)
            info_vbox6.addWidget(self.screenSize_val)
            info_vbox6.addWidget(self.purgeDuration_val)
            info_vbox6.addWidget(self.operator_value)
            info_vbox6.addWidget(self.supervisor_value)

            info_widget1.show()
            info_widget2.show()
            info_widget3.show()
            info_widget4.show()
            info_widget5.show()
            info_widget6.show()

            # Create 3 tables for Time, Materials and Temperature
            try:
                self.time_table = QtWidgets.QTableWidget(self.main_widget)
                self.time_table.setGeometry(20, 450, 330, 300)
                self.time_table.setColumnCount(3)
                self.time_table.setRowCount(len(time_start) + 2)
                self.time_table.setRowCount(9)
                self.time_rows = len(time_end)
                self.time_cols = 3
                self.time_table.setHorizontalHeaderLabels(["Time Start", "Time End", "Time Diff"])
                total_timediff = timedelta()

                for i in range(self.time_rows):
                    for j in range(self.time_cols):
                        if j == 0:
                            item = QtWidgets.QTableWidgetItem(time_start[i].strftime("%b-%d-%Y %H:%M"))  # Convert to string
                            item.setFlags(item.flags() & ~Qt.ItemIsEditable)  # Make the cells unable to be edited
                            self.time_table.setItem(i, j, item)
                        elif j == 1:
                            item = QtWidgets.QTableWidgetItem(time_end[i].strftime("%b-%d-%Y %H:%M"))  # Convert to string
                            item.setFlags(item.flags() & ~Qt.ItemIsEditable)  # Make the cells unable to be edited
                            self.time_table.setItem(i, j, item)
                        else:
                            datetime1 = time_start[i]
                            datetime2 = time_end[i]
                            t_diff = datetime2 - datetime1
                            # Convert timedelta to a string representation (e.g., "X days, HH:MM:SS")
                            total_timediff = total_timediff + t_diff
                            t_diff_str = str(t_diff)
                            item = QtWidgets.QTableWidgetItem(t_diff_str)
                            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                            self.time_table.setItem(i, j, item)

                # Convert it to Hours and Minutes only
                total_hours = total_timediff.days * 24 + total_timediff.seconds // 3600
                total_minutes = (total_timediff.seconds % 3600) // 60

                # Populate the Time Table
                self.time_table.setItem(self.time_rows + 1, 1, QtWidgets.QTableWidgetItem("Total"))
                self.time_table.setItem(self.time_rows + 1, 2,
                                        QtWidgets.QTableWidgetItem(str(total_hours) + ":" + str(total_minutes)))
                self.time_table.show()

                # Material Table
                self.material_table = QtWidgets.QTableWidget(self.main_widget)
                self.material_table.setRowCount(len(materials))
                self.material_table.setColumnCount(2)
                self.material_table.setRowCount(9)
                self.material_table.setGeometry(350, 450, 225, 300)
                self.material_table.setHorizontalHeaderLabels(["Materials", "Quantity(Kg)"])

                # Populate the Materials Table
                for key in list(materials.keys()):
                    key_item = QtWidgets.QTableWidgetItem(str(key))
                    value_item = QtWidgets.QTableWidgetItem(str(materials[key]))
                    key_item.setFlags(key_item.flags() & ~Qt.ItemIsEditable)  # Make the cells unable to be edited
                    value_item.setFlags(value_item.flags() & ~Qt.ItemIsEditable)  # Make the cells unable to be edited
                    self.material_table.setItem(list(materials.keys()).index(key), 0, key_item)
                    self.material_table.setItem(list(materials.keys()).index(key), 1, value_item)

                self.material_table.show()

                # Temperature Table
                self.temp_table = QtWidgets.QTableWidget(self.main_widget)
                self.temp_table.setGeometry(575, 450, 140, 300)
                self.temp_table.setRowCount(12)
                self.temp_table.setColumnCount(1)
                self.temp_table.setVerticalHeaderLabels(["Z" + str(i + 1) for i in range(12)])
                self.temp_table.setHorizontalHeaderLabels(["Temperature"])


                # Populate the Table
                for i in range(len(temperature)):
                    item = QtWidgets.QTableWidgetItem(str(temperature[i]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                    self.temp_table.setItem(i, 0, item)

                self.temp_table.show()
            except Exception as e:
                print(e)

            self.group_box = QtWidgets.QGroupBox(self.main_widget)
            self.group_box.setGeometry(730, 460, 230, 150)
            self.group_box.setTitle("Remarks")
            self.group_box.setFont(QtGui.QFont("Arial", 10))

            self.show_remarks = QtWidgets.QTextEdit(self.group_box)
            self.show_remarks.setGeometry(0, 20, 230, 130)
            self.show_remarks.setAutoFillBackground(True)
            self.show_remarks.setText(remarks)
            self.show_remarks.setEnabled(False)
            self.show_remarks.show()

            self.group_box.show()

            self.export_btn = QtWidgets.QPushButton(self.main_widget)
            self.export_btn.setGeometry(730, 610, 100, 30)
            self.export_btn.setText("Export")
            self.export_btn.clicked.connect(exportToExcel)
            self.export_btn.show()

        def add_entry():

            self.entry_widget = QtWidgets.QWidget()
            self.entry_widget.setGeometry(300, 100, 800, 750)
            self.entry_widget.setWindowIcon(QtGui.QIcon("setting.png"))
            self.entry_widget.setWindowTitle("ADD EXTRUDER DATA")
            self.entry_widget.setStyleSheet("background-color : rgb(240,240,240);")
            self.entry_widget.setWindowModality(Qt.ApplicationModal)
            self.entry_widget.show()


            def get_entries():
                # get the data from the tables

                try:

                    print(type(self.remarks_textBox.toPlainText()))
                    # Time Table
                    temp_row = time_table.rowCount()
                    time_start = []
                    time_end = []
                    outputs = []

                    # Getting the data from the Time Table
                    for i in range(self.time_entry):
                        time_start.append(time_table.item(i, 0))  # time start
                        time_end.append(time_table.item(i, 1))  # time end
                        outputs.append(time_table.item(i, 2))

                    # Removing Null Values
                    time_start = [i for i in time_start if i is not None]
                    time_start = [i.text() for i in time_start]
                    time_end = [i for i in time_end if i is not None]
                    time_end = [i.text() for i in time_end]
                    outputs = [i for i in outputs if i is not None]
                    outputs = [i.text() for i in outputs]

                    total_time = timedelta()
                    try:
                        for i in range(len(time_start)):
                            t_start = datetime.strptime(time_start[i], "%m-%d-%Y %H:%M")
                            t_end = datetime.strptime(time_end[i], "%m-%d-%Y %H:%M")
                            total_time = total_time + (t_start - t_end)
                    except Exception as e:
                        print(e)

                    hours = str(int(total_time.total_seconds() // 3600))
                    minutes = str((int(total_time.total_seconds() % 3600) // 60))
                    seconds = str(int(total_time.total_seconds() % 60))

                    total_hours = round(abs(total_time.total_seconds() / 3600), 2)

                    time_start = ', '.join(["'{}'".format(time) for time in time_start])
                    time_end = ', '.join(["'{}'".format(time) for time in time_end])

                    # Getting the Data for temperature
                    temperature = []
                    for i in range(temperature_table.rowCount()):
                        temperature.append(temperature_table.item(i, 0))

                    temperature = [i for i in temperature if i is not None]
                    temperature = [i.text() for i in temperature]

                    # Declare additional variables need here like loss percentage
                    output_percent = round((float(product_output_input.text()) / float(product_input.text())) * 100,
                                           4)  # Round to the 4th decimal
                    loss_percent = round((float(loss_input.text()) / float(product_input.text())) * 100,
                                         4)  # Round to the 4th decimal
                    purge_duration = timedelta()
                    outputPerHour = round(float(product_output_input.text()) / total_hours, 4)


                    try:
                        purge_start = datetime.strptime(purgeStart_input.text(), "%Y-%m-%d %H:%M")
                        purge_end = datetime.strptime(purgeEnd_input.text(), "%Y-%m-%d %H:%M")
                        purge_duration = purge_end - purge_start


                    except:

                        purge_start = datetime.strptime(
                            datetime.today().strftime("%Y-%m-%d") + " " + purgeStart_input.text(), "%Y-%m-%d %H:%M")
                        purge_end = datetime.strptime(
                            datetime.today().strftime("%Y-%m-%d") + " " + purgeEnd_input.text(),
                            "%Y-%m-%d %H:%M")
                        purge_duration = (purge_end - purge_start).total_seconds()

                    purge_duration = purge_duration // 60
                    # SQL command here to insert Items
                    self.cursor.execute(
                        f"SELECT materials FROM production_merge WHERE production_id = '{productionID_input.text()}'")

                    self.total_mats = json.dumps(self.total_mats)
                    self.total_mats = self.total_mats.replace('\\',"")
                    print(type(self.total_mats), self.total_mats)


                    # Convert the list to string
                    temperature = str(temperature).replace("[", "").replace("]", "")
                    outputs = str(outputs).replace("[", "").replace("]", "")

                    try:

                        self.cursor.execute(f"""
                                        INSERT INTO extruder( machine, qty_order, total_output, customer,
                                        formula_id, product_code, order_id, total_time, time_start, time_end, output_percent,
                                        loss, loss_percent, materials, purging, resin, purge_duration, screw_config, feed_rate, 
                                        rpm, screen_size, operator, supervisor, temperature, outputs, output_per_hour, production_id, total_input,
                                        remarks, lot_number) 
                                        VALUES('{machine_input.text()}', '{orderedQuantity_input.text()}', '{product_output_input.text()}',
                                        '{customer_input.text().replace("'", "''")}', '{self.formulaID_input.text()}', '{productCode_input.text()}',
                                        '{order_number_input.text()}', '{total_hours}', ARRAY[{time_start}]::timestamp[], ARRAY[{time_end}]::timestamp[], 
                                        '{str(output_percent)}', '{loss_input.text()}', '{loss_percent}', '{self.total_mats}', '{purging_input.text()}',
                                         '{resin_input.text()}', {purge_duration}, '{screwConf_input.text()}', '{feedRate_input.text()}',
                                         '{rpm_input.text()}','{screenSize_input.text()}', '{operator_input.text()}', '{supervisor_input.text()}',
                                         ARRAY[{temperature}]::INTEGER[], ARRAY[{outputs}]::FLOAT[], {outputPerHour}, {productionID_input.text()},
                                         {product_input.text()},'{self.remarks_textBox.toPlainText()}', 
                                         ARRAY[{str(self.lot_numberList)}]::VARCHAR[])

                                                """)
                        print("query successful")
                        self.conn.commit()
                        self.entry_widget.close()
                    except Exception as e:
                        print("Insert Failed")
                        QMessageBox.critical(self.entry_widget, "ERROR", "INVALID ENTRY")
                        print(e)
                        self.conn.rollback()
                except Exception as e:
                    print(e)
                    QMessageBox.critical(self.entry_widget, "ERROR", "INVALID ENTRY")
                    return

            def select_production():

                try:
                    self.table.deleteLater()
                    self.table2.deleteLater()
                    self.table3.deleteLater()
                except:
                    pass

                def add_data():
                    pass

                def show_table():
                    self.table2.clearContents()

                    item = self.table.selectedItems()
                    item = [i.text() for i in item]
                    self.cursor.execute(f"""
                    SELECT production_id, lot_number, t_qtyreq, materials
                    FROM production_merge
                    WHERE production_id = '{item[0]}'
                    """)

                    result = self.cursor.fetchall()
                    result = result[0]

                    material = result[-1]

                    try:

                        for keys in list(material.keys()):
                            key = QTableWidgetItem(str(keys))
                            value = QTableWidgetItem(str(material[keys]))
                            self.table2.setItem(list(material.keys()).index(keys), 0, key)
                            key.setFlags(key.flags() & ~Qt.ItemIsEditable)
                            self.table2.setItem(list(material.keys()).index(keys), 1, value)
                            value.setFlags(value.flags() & ~Qt.ItemIsEditable)
                            self.table2.show()

                        label2.setText(str(result[2]))
                        prod_id = QTableWidgetItem(str(result[0]))
                        prod_id.setTextAlignment(Qt.AlignCenter)  # Align Center
                        lot_num = QTableWidgetItem(str(result[1]))
                        lot_num.setTextAlignment(Qt.AlignCenter)
                        self.table3.setItem(0, 0, prod_id)
                        prod_id.setFlags(prod_id.flags() & ~Qt.ItemIsEditable)
                        self.table3.setItem(0, 1, lot_num)
                        lot_num.setFlags(lot_num.flags() & ~Qt.ItemIsEditable)

                    except Exception as e:
                        print(e)

                self.added_entry = 0  # for counting the lot numbers added
                self.total_mats = {}
                self.total_quantity_order = 0
                self.total_output = 0
                self.lot_numberList = []
                self.total_outputPercent = 0
                def push_data():
                    item = self.table.selectedItems()
                    item = [i.text() for i in item]

                    self.cursor.execute(f"""
                                                                SELECT * FROM production_merge
                                                                WHERE production_id = '{item[0]}' 

                                                                """)
                    result = self.cursor.fetchall()
                    result = result[0]

                    # Unpack the result
                    prod_id = result[0]
                    customer = result[2]
                    formula_id = result[3]
                    product_code = result[5]
                    product_color = result[6]
                    lot_number = result[9]
                    order_number = result[10]
                    machine_name = result[14]
                    quantity_order = result[15]
                    output_quantity = result[17]
                    remarks = result[18]
                    materials = result[-1]

                    self.added_entry += 1
                    self.lot_numberList.append(lot_number)
                    self.total_output += output_quantity
                    self.total_quantity_order += quantity_order

                    for key in materials.keys():
                        if key in list(self.total_mats.keys()):
                            self.total_mats[key] = self.total_mats[key] + materials[key]
                        else:
                            self.total_mats[key] = materials[key]

                    # Set the Text to the Extruder Entry Form
                    productionID_input.setText(prod_id)
                    customer_input.setText(customer)
                    productCode_input.setText(product_code)
                    lot_number_input.setText('/'.join(self.lot_numberList))
                    # product_output_input.setText(str(round(self.total_output,2)))
                    machine_input.setText(machine_name)
                    self.formulaID_input.setText(str(formula_id))
                    order_number_input.setText(str(order_number))

                    self.cursor.execute(f"""
                    SELECT production_id, lot_number
                    FROM production_merge
                    WHERE product_code = '{product_code}' AND 
                    machine = '{machine_name}' AND t_fid = '{formula_id}';
                    
                    """)
                    query_result = self.cursor.fetchall()
                    self.table.itemSelectionChanged.disconnect(show_table)
                    self.table.clearSelection()
                    self.table.clear()
                    self.table.setRowCount(len(query_result))
                    self.table.setHorizontalHeaderLabels(["Production ID", "Lot Number"])

                    for row in range(len(query_result)):
                        prod = QTableWidgetItem(query_result[row][0])
                        prod.setFlags(prod.flags() & ~Qt.ItemIsEditable)
                        lot = QTableWidgetItem(query_result[row][1])
                        lot.setFlags(lot.flags() & ~Qt.ItemIsEditable)
                        self.table.setItem(row, 0, prod)
                        self.table.setItem(row, 1, lot)
                    self.table.itemSelectionChanged.connect(show_table)
                    self.table.show()

                    self.table2.clearSelection()
                    self.table2.clear()
                    self.table2.setHorizontalHeaderLabels(["Materials", "Quantity"])
                    for keys in list(self.total_mats.keys()):
                        key = QTableWidgetItem(str(keys))
                        value = QTableWidgetItem(str(self.total_mats[keys]))
                        self.table2.setItem(list(self.total_mats.keys()).index(keys), 0, key)
                        key.setFlags(key.flags() & ~Qt.ItemIsEditable)
                        self.table2.setItem(list(self.total_mats.keys()).index(keys), 1, value)
                        value.setFlags(value.flags() & ~Qt.ItemIsEditable)
                    self.table2.show()

                def search():
                    try:
                        self.table.itemSelectionChanged.disconnect(show_table)
                        self.table.clearContents()
                        self.cursor.execute(f"""
                                            SELECT production_id, lot_number
                                            FROM production_merge
                                            WHERE lot_number ILIKE '%{search_bar.text()}%'
                                            """)
                        search_result = self.cursor.fetchall()

                        for i in range(len(search_result)):

                            item_pair = search_result[i]

                            item = QTableWidgetItem(item_pair[0])
                            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                            item2 = QTableWidgetItem(item_pair[1])
                            item2.setFlags(item2.flags() & ~Qt.ItemIsEditable)
                            self.table.setItem(i, 0, item)
                            self.table.setItem(i, 1, item2)

                        self.table.itemSelectionChanged.connect(show_table)
                    except Exception as e:
                        print(e)

                def clear():

                    try:
                        self.cursor.execute("""
                                            SELECT production_id, lot_number
                                            FROM production_merge

                                                                                """)
                        result = self.cursor.fetchall()
                        self.table.setRowCount(len(result))
                        self.table.setColumnCount(2)
                        self.table.setHorizontalHeaderLabels(["Production ID", "Lot Number"])

                        self.table.itemSelectionChanged.disconnect(show_table)
                        self.table.clearSelection()
                        self.table.clear()
                        search_bar.clear()

                        self.table2.clearContents()
                        self.table3.clearContents()
                        self.table.itemSelectionChanged.connect(show_table)

                    except Exception as e:
                        self.table.setHorizontalHeaderLabels(["Production ID", "Lot Number"])
                        self.table.itemSelectionChanged.connect(show_table)

                    try:
                        self.table.setHorizontalHeaderLabels(["Production ID", "Lot Number"])
                        for i in range(len(result)):
                            prod_id = QTableWidgetItem(str(result[i][0]))
                            lot_num = QTableWidgetItem(str(result[i][1]))
                            self.table.setItem(i, 0, prod_id)
                            prod_id.setFlags(prod_id.flags() & ~Qt.ItemIsEditable)
                            self.table.setItem(i, 1, lot_num)
                            lot_num.setFlags(lot_num.flags() & ~Qt.ItemIsEditable)
                            self.table.show()

                    except Exception as e:
                        print(e)

                self.cursor.execute("""
                SELECT production_id, lot_number
                FROM production_merge

                """)
                result = self.cursor.fetchall()

                self.selectProd_widget = QtWidgets.QWidget()
                self.selectProd_widget.setGeometry(400, 200, 800, 630)
                self.selectProd_widget.setFixedSize(800, 600)

                search_bar = QtWidgets.QLineEdit(self.selectProd_widget)
                search_bar.setGeometry(40, 25, 170, 25)
                search_bar.setFont(QtGui.QFont("Arial", 10))
                search_bar.setPlaceholderText("Search Lot Number")
                search_bar.show()

                search_btn = QtWidgets.QPushButton(self.selectProd_widget)
                search_btn.setGeometry(210, 25, 70, 27)
                search_btn.setText("Search")
                search_btn.clicked.connect(search)
                search_btn.show()

                clear_btn = QtWidgets.QPushButton(self.selectProd_widget)
                clear_btn.setGeometry(280, 25, 70, 27)
                clear_btn.setText("Clear")
                clear_btn.clicked.connect(clear)
                clear_btn.show()

                # ProductionId and Lot Number Table widget
                self.table = QtWidgets.QTableWidget(self.selectProd_widget)
                self.table.setGeometry(0, 70, 350, 500)
                self.table.setColumnCount(2)
                self.table.setRowCount(len(result))
                self.table.setHorizontalHeaderLabels(["Production ID", "Lot Number"])
                self.table.setColumnWidth(1, 180)
                self.table.setColumnWidth(0, 150)
                self.table.setFont(QtGui.QFont("Arial", 12))
                self.table.setStyleSheet('color: rgb(0,109,189);')
                self.table.verticalHeader().setVisible(False)
                try:
                    for i in range(len(result)):
                        prod_id = QTableWidgetItem(str(result[i][0]))
                        lot_num = QTableWidgetItem(str(result[i][1]))
                        self.table.setItem(i, 0, prod_id)
                        prod_id.setFlags(prod_id.flags() & ~Qt.ItemIsEditable)
                        self.table.setItem(i, 1, lot_num)
                        lot_num.setFlags(lot_num.flags() & ~Qt.ItemIsEditable)
                except Exception as e:
                    print(e)

                self.table.show()
                self.table.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)
                self.table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
                self.table.itemSelectionChanged.connect(show_table)

                # Materials Table
                self.table2 = QtWidgets.QTableWidget(self.selectProd_widget)
                self.table2.setGeometry(350, 70, 450, 500)
                self.table2.setColumnCount(2)
                self.table2.setRowCount(16)
                self.table2.setColumnWidth(0, 205)
                self.table2.setColumnWidth(1, 225)
                self.table2.verticalHeader().setVisible(False)
                self.table2.setHorizontalHeaderLabels(["Materials", "Quantity"])
                self.table2.show()

                # Table 3 For showing Selected ProdID and Lot Number
                self.table3 = QtWidgets.QTableWidget(self.selectProd_widget)
                self.table3.setGeometry(380, 0, 402, 60)
                self.table3.setColumnCount(2)
                self.table3.setRowCount(1)
                self.table3.setHorizontalHeaderLabels(["Production ID", "Lot Number"])
                self.table3.setColumnWidth(0, 200)
                self.table3.setColumnWidth(1, 200)
                self.table3.setRowHeight(0, 35)
                font = QtGui.QFont("Arial", 12)
                font.setBold(True)
                self.table3.setFont(font)
                self.table3.setStyleSheet("color: rgb(0,109,189) ")
                self.table3.horizontalHeader().setStyleSheet("""
                    QHeaderView::section{
                        font-weight: bold;
                        background-color: rgb(0,109,189);
                        color: white;
                    }
                """)

                for i in range(2):
                    item = QTableWidgetItem("")
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)  # Make the cells unable to be edited
                    self.table3.setItem(0, i, item)

                self.table3.verticalHeader().setVisible(False)
                self.table3.show()

                label1 = QtWidgets.QLabel(self.selectProd_widget)
                label1.setGeometry(230, 570, 120, 30)
                label1.setText("Total QTY (Kg)")
                label1.setFont(QtGui.QFont("Arial", 11))
                label1.setAlignment(Qt.AlignCenter)
                label1.setStyleSheet('border: 1px solid black;')
                label1.show()

                label2 = QtWidgets.QLabel(self.selectProd_widget)
                label2.setGeometry(350, 570, 200, 30)
                label2.setFont(QtGui.QFont("Arial", 15))
                label2.setAlignment(Qt.AlignCenter)
                label2.setStyleSheet("background-color: white; color: blue;")
                label2.setText("Test")
                label2.show()

                # Save Button
                save_prod = QtWidgets.QPushButton(self.selectProd_widget)
                save_prod.setGeometry(590, 570, 70, 30)
                save_prod.setText("Push Data")
                save_prod.clicked.connect(push_data)
                save_prod.show()

                refresh = QtWidgets.QPushButton(self.selectProd_widget)
                refresh.setGeometry(660, 570, 70, 30)
                refresh.setText("ADD")
                refresh.show()

                close = QtWidgets.QPushButton(self.selectProd_widget)
                close.setGeometry(730, 570, 70, 30)
                close.setText("Close")
                close.show()

                self.selectProd_widget.setWindowModality(Qt.ApplicationModal)
                self.selectProd_widget.show()

            self.time_entry = 0

            def add_time():

                product_output_input.setText(str(float(product_output_input.text()) + float(output_lineEdit.text())))

                item1 = QTableWidgetItem(time_start_input.text())
                item2 = QTableWidgetItem(time_end_input.text())
                item3 = QTableWidgetItem(output_lineEdit.text())

                item3.setTextAlignment(Qt.AlignCenter)

                time_table.setItem(self.time_entry, 0, item1)
                time_table.setItem(self.time_entry, 1, item2)
                time_table.setItem(self.time_entry, 2, item3)

                self.time_entry += 1

                output_lineEdit.clear()

            def reset_table():
                time_table.clearContents()
                product_output_input.setText("0.0")

            def loss_auto():
                if product_output_input.text() != "":
                    try:
                        loss_input.setText(
                            str(round(float(product_input.text()) - float(product_output_input.text()), 4)))
                    except:
                        loss_input.setText("INVALID")

            # Create two new widget for the VBOX Layout
            self.leftInput_side = QtWidgets.QWidget(self.entry_widget)
            self.leftInput_side.setGeometry(0, 0, 400, 450)
            self.leftInput_side.show()

            self.right_side = QtWidgets.QWidget(self.entry_widget)
            self.right_side.setGeometry(400, 0, 400, 450)
            self.right_side.show()

            # Create Vertical Box Layout
            self.left_vbox = QtWidgets.QFormLayout(self.leftInput_side)
            self.left_vbox.setSpacing(20)
            self.right_vbox = QtWidgets.QFormLayout(self.right_side)
            self.right_vbox.setSpacing(20)

            font = QtGui.QFont("Berlin Sans FB", 14)

            productionID_label = QtWidgets.QLabel()
            productionID_label.setText("Production ID")
            productionID_label.setFont(font)

            customer_label = QtWidgets.QLabel()
            customer_label.setText("Customer")
            customer_label.setFont(font)

            machine_label = QtWidgets.QLabel()
            machine_label.setText("Machine No.")
            machine_label.setFont(font)

            productCode_label = QtWidgets.QLabel()
            productCode_label.setText("Product Code")
            productCode_label.setFont(font)

            productOutput_label = QtWidgets.QLabel()
            productOutput_label.setText("Output (kg)")
            productOutput_label.setFont(font)

            formulaID_label = QtWidgets.QLabel()
            formulaID_label.setText("Formula ID")
            formulaID_label.setFont(font)

            lotnumber_label = QtWidgets.QLabel()
            lotnumber_label.setText("Lot Number")
            lotnumber_label.setFont(font)

            orderedQuantity_label = QtWidgets.QLabel()
            orderedQuantity_label.setText("Ordered Qty")
            orderedQuantity_label.setFont(font)

            feedrate_label = QtWidgets.QLabel()
            feedrate_label.setText("Feed Rate")
            feedrate_label.setFont(font)

            rpm_label = QtWidgets.QLabel()
            rpm_label.setText("RPM")
            rpm_label.setFont(font)

            screenSize_label = QtWidgets.QLabel()
            screenSize_label.setText("Screen Size")
            screenSize_label.setFont(font)

            screwConf_label = QtWidgets.QLabel()
            screwConf_label.setText("Screw Config")
            screwConf_label.setFont(font)

            loss_label = QtWidgets.QLabel()
            loss_label.setText("Loss")
            loss_label.setFont(font)

            purgeStart_label = QtWidgets.QLabel()
            purgeStart_label.setText("Purge Start")
            purgeStart_label.setFont(font)

            purgeEnd_label = QtWidgets.QLabel()
            purgeEnd_label.setText("Purge End")
            purgeEnd_label.setFont(font)

            remarks_label = QtWidgets.QLabel()

            operator_label = QtWidgets.QLabel()
            operator_label.setText("Operator")
            operator_label.setFont(font)

            supervisor_label = QtWidgets.QLabel()
            supervisor_label.setText("Supervisor")
            supervisor_label.setFont(font)

            order_number_lbl = QtWidgets.QLabel()
            order_number_lbl.setText("Order Number")
            order_number_lbl.setFont(font)

            resin_label = QtWidgets.QLabel()
            resin_label.setText("Resin")
            resin_label.setFont(font)

            purging_label = QtWidgets.QLabel()
            purging_label.setText("Purging")
            purging_label.setFont(font)

            product_input_label = QtWidgets.QLabel()
            product_input_label.setText("Input")
            product_input_label.setFont(font)

            # QLineEdit Boxes
            productionID_input = QtWidgets.QLineEdit()
            productionID_input.setFixedHeight(25)
            productionID_input.setEnabled(False)
            productionID_input.setAlignment(Qt.AlignCenter)
            productionID_input.setStyleSheet("background-color: white; border: 1px solid black")

            machine_input = QtWidgets.QLineEdit()
            machine_input.setFixedHeight(25)
            machine_input.setEnabled(False)
            machine_input.setAlignment(Qt.AlignCenter)
            machine_input.setStyleSheet("background-color: white; border: 1px solid black")

            customer_input = QtWidgets.QLineEdit()
            customer_input.setFixedHeight(25)
            customer_input.setEnabled(False)
            customer_input.setAlignment(Qt.AlignCenter)
            customer_input.setStyleSheet("background-color: white; border: 1px solid black")

            orderedQuantity_input = QtWidgets.QLineEdit()
            orderedQuantity_input.setAlignment(Qt.AlignCenter)
            orderedQuantity_input.setFixedHeight(25)
            orderedQuantity_input.setStyleSheet("background-color: white; border: 1px solid black")

            productCode_input = QtWidgets.QLineEdit()
            productCode_input.setFixedHeight(25)
            productCode_input.setEnabled(False)
            productCode_input.setAlignment(Qt.AlignCenter)
            productCode_input.setStyleSheet("background-color: white; border: 1px solid black")

            product_output_input = QtWidgets.QLineEdit()
            product_output_input.setFixedHeight(25)
            product_output_input.setEnabled(False)
            product_output_input.setText("0.0")
            product_output_input.setAlignment(Qt.AlignCenter)
            product_output_input.setStyleSheet("background-color: white; border: 1px solid black")

            self.formulaID_input = QtWidgets.QLineEdit()
            self.formulaID_input.setAlignment(Qt.AlignCenter)
            self.formulaID_input.setFixedHeight(25)
            self.formulaID_input.setEnabled(False)
            self.formulaID_input.setStyleSheet("background-color: white; border: 1px solid black")

            lot_number_input = QtWidgets.QLineEdit()
            lot_number_input.setAlignment(Qt.AlignCenter)
            lot_number_input.setFixedHeight(25)
            lot_number_input.setEnabled(False)
            lot_number_input.setStyleSheet("background-color: white; border: 1px solid black")

            feedRate_input = QtWidgets.QLineEdit()
            feedRate_input.setFixedHeight(25)
            feedRate_input.setStyleSheet("background-color: white; border: 1px solid black")

            rpm_input = QtWidgets.QLineEdit()
            rpm_input.setFixedHeight(25)
            rpm_input.setStyleSheet("background-color: white; border: 1px solid black")

            screenSize_input = QtWidgets.QLineEdit()
            screenSize_input.setFixedHeight(25)
            screenSize_input.setStyleSheet("background-color: white; border: 1px solid black")

            screwConf_input = QtWidgets.QLineEdit()
            screwConf_input.setFixedHeight(25)
            screwConf_input.setStyleSheet("background-color: white; border: 1px solid black ")

            loss_input = QtWidgets.QLineEdit()
            loss_input.setFixedHeight(25)
            loss_input.setEnabled(False)
            loss_input.setAlignment(Qt.AlignCenter)
            loss_input.setStyleSheet("background-color: white; border: 1px solid black")

            purgeStart_input = QtWidgets.QLineEdit()
            purgeStart_input.setFixedHeight(25)
            purgeStart_input.setAlignment(Qt.AlignCenter)
            purgeStart_input.setStyleSheet("background-color: white; border: 1px solid black")
            purgeStart_input.setText("00:00")

            purgeEnd_input = QtWidgets.QLineEdit()
            purgeEnd_input.setFixedHeight(25)
            purgeEnd_input.setAlignment(Qt.AlignCenter)
            purgeEnd_input.setStyleSheet("background-color: white; border: 1px solid black")
            purgeEnd_input.setText("00:00")

            remarks = QtWidgets.QTextEdit()

            operator_input = QtWidgets.QLineEdit()
            operator_input.setFixedHeight(25)
            operator_input.setAlignment(Qt.AlignCenter)
            operator_input.setStyleSheet("background-color: white; border: 1px solid black")

            supervisor_input = QtWidgets.QLineEdit()
            supervisor_input.setFixedHeight(25)
            supervisor_input.setAlignment(Qt.AlignCenter)
            supervisor_input.setStyleSheet("background-color: white; border: 1px solid black")

            order_number_input = QtWidgets.QLineEdit()
            order_number_input.setFixedHeight(25)
            order_number_input.setEnabled(False)
            order_number_input.setAlignment(Qt.AlignCenter)
            order_number_input.setStyleSheet("background-color: white; border: 1px solid black")

            resin_input = QtWidgets.QLineEdit()
            resin_input.setFixedHeight(25)
            resin_input.setAlignment(Qt.AlignCenter)
            resin_input.setStyleSheet("background-color: white; border: 1px solid black")

            purging_input = QtWidgets.QLineEdit()
            purging_input.setFixedHeight(25)
            purging_input.setAlignment(Qt.AlignCenter)
            purging_input.setStyleSheet("background-color: white; border: 1px solid black")

            product_input = QtWidgets.QLineEdit()
            product_input.setFixedHeight(25)
            product_input.setAlignment(Qt.AlignCenter)
            product_input.setStyleSheet("background-color: white; border: 1px solid black")
            product_input.textChanged.connect(loss_auto)

            self.groupBoxRemarks = QtWidgets.QGroupBox(self.entry_widget)
            self.groupBoxRemarks.setGeometry(600, 500, 200, 150)
            self.groupBoxRemarks.setTitle("Remarks")
            self.groupBoxRemarks.show()

            self.remarks_textBox = QtWidgets.QTextEdit(self.groupBoxRemarks)
            self.remarks_textBox.setGeometry(0, 20, 200, 130)
            self.remarks_textBox.show()

            # Left Side of Vertical Box
            self.left_vbox.addRow(productionID_label, productionID_input)
            self.left_vbox.addRow(productCode_label, productCode_input)
            self.left_vbox.addRow(customer_label, customer_input)
            self.left_vbox.addRow(orderedQuantity_label, orderedQuantity_input)
            self.left_vbox.addRow(lotnumber_label, lot_number_input)
            self.left_vbox.addRow(product_input_label, product_input)
            self.left_vbox.addRow(productOutput_label, product_output_input)
            self.left_vbox.addRow(loss_label, loss_input)
            self.left_vbox.addRow(machine_label, machine_input)
            self.left_vbox.addRow(formulaID_label, self.formulaID_input)
            self.left_vbox.addRow(order_number_lbl, order_number_input)

            # Add widgets to the right Form Box
            self.right_vbox.addRow(feedrate_label, feedRate_input)
            self.right_vbox.addRow(rpm_label, rpm_input)
            self.right_vbox.addRow(screenSize_label, screenSize_input)
            self.right_vbox.addRow(screwConf_label, screwConf_input)
            self.right_vbox.addRow(purging_label, purging_input)
            self.right_vbox.addRow(resin_label, resin_input)
            self.right_vbox.addRow(purgeStart_label, purgeStart_input)
            self.right_vbox.addRow(purgeEnd_label, purgeEnd_input)
            self.right_vbox.addRow(operator_label, operator_input)
            self.right_vbox.addRow(supervisor_label, supervisor_input)

            # Time Table Entry
            time_table = QtWidgets.QTableWidget(self.entry_widget)
            time_table.setGeometry(0, 500, 450, 200)
            time_table.setColumnCount(3)
            time_table.setRowCount(8)
            time_table.setColumnWidth(0, 150)
            time_table.setColumnWidth(1, 150)
            time_table.setStyleSheet("background-color: white;")
            time_table.setFont(QtGui.QFont("Arial", 10))
            time_table.setEnabled(False)

            time_table.setHorizontalHeaderLabels(["Time Start", "Time End", "Output"])
            time_table.show()

            # Temperature Table Entry
            temperature_table = QtWidgets.QTableWidget(self.entry_widget)
            temperature_table.setGeometry(450, 500, 150, 200)
            temperature_table.setColumnCount(1)
            temperature_table.setRowCount(12)
            temperature_table.setStyleSheet("background-color: white;")
            temperature_table.setHorizontalHeaderLabels(["Temperature"])
            temperature_index = ["Z" + str(i + 1) for i in range(12)]  # set the index
            temperature_table.setVerticalHeaderLabels(temperature_index)
            temperature_table.show()

            # Select Production Data Button
            select_prod = QtWidgets.QPushButton(self.entry_widget)
            select_prod.setGeometry(600, 705, 60, 25)
            select_prod.setText("Select")
            select_prod.clicked.connect(select_production)
            select_prod.setCursor(Qt.PointingHandCursor)
            select_prod.show()

            save_btn = QtWidgets.QPushButton(self.entry_widget)
            save_btn.setGeometry(540, 705, 60, 25)
            save_btn.clicked.connect(get_entries)
            save_btn.setText("Save")
            save_btn.setCursor(Qt.PointingHandCursor)
            save_btn.show()

            default_date = QtCore.QDateTime(2024, 1, 1, 0, 0)

            time_start_input = QtWidgets.QDateTimeEdit(self.entry_widget)
            time_start_input.setGeometry(30, 475, 120, 25)
            time_start_input.setDisplayFormat("MM-dd-yyyy HH:mm")
            time_start_input.setDateTime(default_date)
            time_start_input.show()

            time_end_input = QtWidgets.QDateTimeEdit(self.entry_widget)
            time_end_input.setGeometry(180, 475, 120, 25)
            time_end_input.setDisplayFormat("MM-dd-yyyy HH:mm")
            time_end_input.setDateTime(default_date)
            time_end_input.show()

            output_lineEdit = QtWidgets.QLineEdit(self.entry_widget)
            output_lineEdit.setGeometry(310, 475, 80, 25)
            output_lineEdit.setAlignment(Qt.AlignCenter)
            output_lineEdit.setStyleSheet("background-color: white; border: 1px solid black")
            output_lineEdit.show()

            self.plus_icon = ClickableLabel(self.entry_widget)
            self.plus_icon.setGeometry(390, 475, 25, 25)
            self.plus_icon.setPixmap(QtGui.QIcon('plus.png').pixmap(25, 25))
            self.plus_icon.setCursor(Qt.PointingHandCursor)
            self.plus_icon.clicked.connect(add_time)
            self.plus_icon.show()

            self.reset_icon = ClickableLabel(self.entry_widget)
            self.reset_icon.setGeometry(425, 475, 25, 25)
            self.reset_icon.setPixmap(QtGui.QIcon('reset.png').pixmap(20, 20))
            self.reset_icon.setCursor(Qt.PointingHandCursor)
            self.reset_icon.clicked.connect(reset_table)
            self.reset_icon.show()

        def update_entry():
            try:
                selected = self.extruder_table.selectedItems()
                selected = [i.text() for i in selected]
                self.cursor.execute(f"SELECT * FROM extruder WHERE process_id = {selected[0]}")
                result = self.cursor.fetchall()
                result = result[0]

            except:
                QMessageBox.critical(self.main_widget,"ERROR", "No Data Selected")
                return

            self.entry_widget = QtWidgets.QWidget()
            self.entry_widget.setGeometry(300, 100, 800, 750)
            self.entry_widget.setStyleSheet("background-color : rgb(240,240,240);")
            self.entry_widget.setWindowModality(Qt.ApplicationModal)
            self.entry_widget.show()

            def update():
                # get the data from the tables

                # Time Table
                temp_row = time_table.rowCount()
                time_start = []
                time_end = []
                outputs = []

                # Getting the data from the Time Table
                for i in range(8):
                    if time_table.item(i,0) == None:
                        break
                    else:
                        time_start.append(time_table.item(i, 0))  # time start
                        time_end.append(time_table.item(i, 1))  # time end
                        outputs.append(time_table.item(i, 2))

                # Removing Null Values
                time_start = [i for i in time_start if i is not None]
                time_start = [i.text() for i in time_start]
                time_end = [i for i in time_end if i is not None]
                time_end = [i.text() for i in time_end]
                outputs = [i for i in outputs if i is not None]
                outputs = [i.text() for i in outputs]

                total_time = timedelta()

                try:
                    for i in range(len(time_start)):
                        t_start = datetime.strptime(time_start[i], "%Y-%m-%d %H:%M:%S")
                        t_end = datetime.strptime(time_end[i], "%Y-%m-%d %H:%M:%S")
                        total_time = total_time + (t_end - t_start)

                except Exception as e:
                    print(e)
                hours = str(int(total_time.total_seconds() // 3600)).replace('-', '')
                minutes = str((int(total_time.total_seconds() % 3600) // 60))
                seconds = str(int(total_time.total_seconds() % 60))

                total_hours = round(abs(total_time.total_seconds() / 3600), 2)

                time_start = ', '.join(["'{}'".format(time) for time in time_start])
                time_end = ', '.join(["'{}'".format(time) for time in time_end])

                # Getting the Data for temperature
                temperature = []
                for i in range(temperature_table.rowCount()):
                    temperature.append(temperature_table.item(i, 0))

                temperature = [i for i in temperature if i is not None]
                temperature = [i.text() for i in temperature]

                # Declare additional variables need here like loss percentage
                output_percent = round((float(product_output_input.text()) / float(product_input.text())) * 100,
                                       4)  # Round to the 4th decimal
                loss_percent = round((float(loss_input.text()) / float(product_input.text())) * 100,
                                     4)  # Round to the 4th decimal

                outputPerHour = round(float(product_output_input.text()) / total_hours, 4)

                try:
                    purge_duration = timedelta()
                    try:
                        purge_start = datetime.strptime(purgeStart_input.text(), "%Y-%m-%d %H:%M")
                        purge_end = datetime.strptime(purgeEnd_input.text(), "%Y-%m-%d %H:%M")
                        purge_duration = abs(purge_end - purge_start)

                    except:
                        purge_start = datetime.strptime(
                            datetime.today().strftime("%Y-%m-%d") + " " + purgeStart_input.text(), "%Y-%m-%d %H:%M")
                        purge_end = datetime.strptime(
                            datetime.today().strftime("%Y-%m-%d") + " " + purgeEnd_input.text(),
                            "%Y-%m-%d %H:%M")
                        purge_duration = (purge_end - purge_start).total_seconds()

                    purge_duration = purge_duration // 60
                except:
                    purge_duration = result[25]


                # SQL command here to insert Items
                self.cursor.execute(
                    f"SELECT materials FROM production_merge WHERE production_id = '{productionID_input.text()}'")
                material = self.cursor.fetchall()
                material = material[0][0]
                material = json.dumps(material)

                # Convert the list to string
                temperature = str(temperature).replace("[", "").replace("]", "")
                outputs = str(outputs).replace("[", "").replace("]", "")

                try:

                    self.cursor.execute(f"""
                                UPDATE extruder
                                SET  total_time = {total_hours}, total_input = {product_input.text()}, outputs = ARRAY[{outputs}]::FLOAT[],
                                temperature = ARRAY[{temperature}]::INTEGER[], remarks = '{self.remarks_textBox.toPlainText()}',
                                feed_rate = '{feedRate_input.text()}', rpm = '{rpm_input.text()}', screen_size = '{screenSize_input.text()}',
                                screw_config = '{screwConf_input.text()}', purging = '{purging_input.text()}', resin = '{resin_input.text()}',
                                purge_duration = {purge_duration}, operator = '{operator_input.text()}', supervisor = '{supervisor_input.text()}',
                                time_start = ARRAY[{time_start}]::timestamp[], time_end =  ARRAY[{time_end}]::timestamp[],
                                output_percent = '{str(output_percent)}', loss = '{loss_input.text()}', loss_percent = '{loss_percent}',
                                output_per_hour = '{outputPerHour}' 
                                WHERE process_id = {selected[0]};
                                ;      
                                """)
                    QMessageBox.information(self.entry_widget, "UPDATE SUCCESSFUL", f"Successfully Updated \n Form No. {selected[0]}")
                    print("query successful")
                    self.conn.commit()
                    self.entry_widget.close()
                except Exception as e:
                    print("Insert Failed")
                    print(e)
                    self.conn.rollback()

            self.time_entry = 0

            def add_time():

                item1 = QTableWidgetItem(time_start_input.text())
                item2 = QTableWidgetItem(time_end_input.text())
                item3 = QTableWidgetItem(output_lineEdit.text())

                item3.setTextAlignment(Qt.AlignCenter)

                time_table.setItem(self.time_entry, 0, item1)
                time_table.setItem(self.time_entry, 1, item2)
                time_table.setItem(self.time_entry, 2, item3)

                self.time_entry += 1

                output_lineEdit.clear()

            def loss_auto():
                if product_output_input.text() != "":
                    try:
                        loss_input.setText(
                            str(round(float(product_input.text()) - float(product_output_input.text()), 4)))
                    except:
                        loss_input.setText("INVALID")

            # Create two new widget for the VBOX Layout
            self.leftInput_side = QtWidgets.QWidget(self.entry_widget)
            self.leftInput_side.setGeometry(0, 0, 400, 450)
            self.leftInput_side.show()

            self.right_side = QtWidgets.QWidget(self.entry_widget)
            self.right_side.setGeometry(400, 0, 400, 450)
            self.right_side.show()

            # Create Vertical Box Layout
            self.left_vbox = QtWidgets.QFormLayout(self.leftInput_side)
            self.left_vbox.setSpacing(20)
            self.right_vbox = QtWidgets.QFormLayout(self.right_side)
            self.right_vbox.setSpacing(20)

            font = QtGui.QFont("Berlin Sans FB", 14)

            productionID_label = QtWidgets.QLabel()
            productionID_label.setText("Production ID")
            productionID_label.setFont(font)

            customer_label = QtWidgets.QLabel()
            customer_label.setText("Customer")
            customer_label.setFont(font)

            machine_label = QtWidgets.QLabel()
            machine_label.setText("Machine No.")
            machine_label.setFont(font)

            productCode_label = QtWidgets.QLabel()
            productCode_label.setText("Product Code")
            productCode_label.setFont(font)

            productOutput_label = QtWidgets.QLabel()
            productOutput_label.setText("Output (kg)")
            productOutput_label.setFont(font)

            formulaID_label = QtWidgets.QLabel()
            formulaID_label.setText("Formula ID")
            formulaID_label.setFont(font)

            lotnumber_label = QtWidgets.QLabel()
            lotnumber_label.setText("Lot Number")
            lotnumber_label.setFont(font)

            orderedQuantity_label = QtWidgets.QLabel()
            orderedQuantity_label.setText("Ordered Qty")
            orderedQuantity_label.setFont(font)

            feedrate_label = QtWidgets.QLabel()
            feedrate_label.setText("Feed Rate")
            feedrate_label.setFont(font)

            rpm_label = QtWidgets.QLabel()
            rpm_label.setText("RPM")
            rpm_label.setFont(font)

            screenSize_label = QtWidgets.QLabel()
            screenSize_label.setText("Screen Size")
            screenSize_label.setFont(font)

            screwConf_label = QtWidgets.QLabel()
            screwConf_label.setText("Screw Config")
            screwConf_label.setFont(font)

            loss_label = QtWidgets.QLabel()
            loss_label.setText("Loss")
            loss_label.setFont(font)

            purgeStart_label = QtWidgets.QLabel()
            purgeStart_label.setText("Purge Start")
            purgeStart_label.setFont(font)

            purgeEnd_label = QtWidgets.QLabel()
            purgeEnd_label.setText("Purge End")
            purgeEnd_label.setFont(font)

            remarks_label = QtWidgets.QLabel()

            operator_label = QtWidgets.QLabel()
            operator_label.setText("Operator")
            operator_label.setFont(font)

            supervisor_label = QtWidgets.QLabel()
            supervisor_label.setText("Supervisor")
            supervisor_label.setFont(font)

            order_number_lbl = QtWidgets.QLabel()
            order_number_lbl.setText("Order Number")
            order_number_lbl.setFont(font)

            resin_label = QtWidgets.QLabel()
            resin_label.setText("Resin")
            resin_label.setFont(font)

            purging_label = QtWidgets.QLabel()
            purging_label.setText("Purging")
            purging_label.setFont(font)

            product_input_label = QtWidgets.QLabel()
            product_input_label.setText("Input")
            product_input_label.setFont(font)

            # QLineEdit Boxes
            productionID_input = QtWidgets.QLineEdit()
            productionID_input.setFixedHeight(25)
            productionID_input.setEnabled(False)
            productionID_input.setAlignment(Qt.AlignCenter)
            productionID_input.setStyleSheet("background-color: white; border: 1px solid black")
            productionID_input.setText(str(result[-3]))
            productionID_input.setEnabled(False)

            machine_input = QtWidgets.QLineEdit()
            machine_input.setFixedHeight(25)
            machine_input.setEnabled(False)
            machine_input.setAlignment(Qt.AlignCenter)
            machine_input.setStyleSheet("background-color: white; border: 1px solid black")
            machine_input.setText(result[1])
            machine_input.setEnabled(False)

            customer_input = QtWidgets.QLineEdit()
            customer_input.setFixedHeight(25)
            customer_input.setEnabled(False)
            customer_input.setAlignment(Qt.AlignCenter)
            customer_input.setStyleSheet("background-color: white; border: 1px solid black")
            customer_input.setText(result[4])
            customer_input.setEnabled(False)

            orderedQuantity_input = QtWidgets.QLineEdit()
            orderedQuantity_input.setAlignment(Qt.AlignCenter)
            orderedQuantity_input.setFixedHeight(25)
            orderedQuantity_input.setEnabled(False)
            orderedQuantity_input.setStyleSheet("background-color: white; border: 1px solid black")
            orderedQuantity_input.setText(str(result[2]))

            productCode_input = QtWidgets.QLineEdit()
            productCode_input.setFixedHeight(25)
            productCode_input.setEnabled(False)
            productCode_input.setAlignment(Qt.AlignCenter)
            productCode_input.setStyleSheet("background-color: white; border: 1px solid black")
            productCode_input.setText(result[6])

            product_output_input = QtWidgets.QLineEdit()
            product_output_input.setFixedHeight(25)
            product_output_input.setAlignment(Qt.AlignCenter)
            product_output_input.setStyleSheet("background-color: white; border: 1px solid black")
            product_output_input.setText(str(result[3]))

            self.formulaID_input = QtWidgets.QLineEdit()
            self.formulaID_input.setAlignment(Qt.AlignCenter)
            self.formulaID_input.setFixedHeight(25)
            self.formulaID_input.setEnabled(False)
            self.formulaID_input.setStyleSheet("background-color: white; border: 1px solid black")
            self.formulaID_input.setText(str(result[5]))

            lot_number_input = QtWidgets.QLineEdit()
            lot_number_input.setAlignment(Qt.AlignCenter)
            lot_number_input.setFixedHeight(25)
            lot_number_input.setEnabled(False)
            lot_number_input.setStyleSheet("background-color: white; border: 1px solid black")
            try:
                lot_number_input.setText('/'.join(result[-1][0]))
            except:
                lot_number_input.setText(None)

            feedRate_input = QtWidgets.QLineEdit()
            feedRate_input.setFixedHeight(25)
            feedRate_input.setStyleSheet("background-color: white; border: 1px solid black")
            feedRate_input.setText(str(result[18]))

            rpm_input = QtWidgets.QLineEdit()
            rpm_input.setFixedHeight(25)
            rpm_input.setStyleSheet("background-color: white; border: 1px solid black")
            rpm_input.setText(str(result[19]))

            screenSize_input = QtWidgets.QLineEdit()
            screenSize_input.setFixedHeight(25)
            screenSize_input.setStyleSheet("background-color: white; border: 1px solid black")
            screenSize_input.setText(result[20])

            screwConf_input = QtWidgets.QLineEdit()
            screwConf_input.setFixedHeight(25)
            screwConf_input.setStyleSheet("background-color: white; border: 1px solid black ")
            screwConf_input.setText(result[17])

            loss_input = QtWidgets.QLineEdit()
            loss_input.setFixedHeight(25)
            loss_input.setEnabled(False)
            loss_input.setAlignment(Qt.AlignCenter)
            loss_input.setStyleSheet("background-color: white; border: 1px solid black")
            loss_input.setText(str(result[12]))

            purgeStart_input = QtWidgets.QLineEdit()
            purgeStart_input.setFixedHeight(25)
            purgeStart_input.setAlignment(Qt.AlignCenter)
            purgeStart_input.setStyleSheet("background-color: white; border: 1px solid black")

            purgeEnd_input = QtWidgets.QLineEdit()
            purgeEnd_input.setFixedHeight(25)
            purgeEnd_input.setAlignment(Qt.AlignCenter)
            purgeEnd_input.setStyleSheet("background-color: white; border: 1px solid black")

            remarks = QtWidgets.QTextEdit()

            operator_input = QtWidgets.QLineEdit()
            operator_input.setFixedHeight(25)
            operator_input.setAlignment(Qt.AlignCenter)
            operator_input.setStyleSheet("background-color: white; border: 1px solid black")
            operator_input.setText(result[21])

            supervisor_input = QtWidgets.QLineEdit()
            supervisor_input.setFixedHeight(25)
            supervisor_input.setAlignment(Qt.AlignCenter)
            supervisor_input.setStyleSheet("background-color: white; border: 1px solid black")
            supervisor_input.setText(result[22])

            order_number_input = QtWidgets.QLineEdit()
            order_number_input.setFixedHeight(25)
            order_number_input.setEnabled(False)
            order_number_input.setAlignment(Qt.AlignCenter)
            order_number_input.setStyleSheet("background-color: white; border: 1px solid black")
            order_number_input.setText(str(result[7]))

            resin_input = QtWidgets.QLineEdit()
            resin_input.setFixedHeight(25)
            resin_input.setAlignment(Qt.AlignCenter)
            resin_input.setStyleSheet("background-color: white; border: 1px solid black")
            resin_input.setText(result[15])

            purging_input = QtWidgets.QLineEdit()
            purging_input.setFixedHeight(25)
            purging_input.setAlignment(Qt.AlignCenter)
            purging_input.setStyleSheet("background-color: white; border: 1px solid black")
            purging_input.setText(result[14])

            product_input = QtWidgets.QLineEdit()
            product_input.setFixedHeight(25)
            product_input.setAlignment(Qt.AlignCenter)
            product_input.setStyleSheet("background-color: white; border: 1px solid black")
            product_input.textChanged.connect(loss_auto)
            product_input.setText(str(result[-2]))

            self.groupBoxRemarks = QtWidgets.QGroupBox(self.entry_widget)
            self.groupBoxRemarks.setGeometry(600, 500, 200, 150)
            self.groupBoxRemarks.setTitle("Remarks")
            self.groupBoxRemarks.show()

            self.remarks_textBox = QtWidgets.QTextEdit(self.groupBoxRemarks)
            self.remarks_textBox.setGeometry(0, 20, 200, 130)
            self.remarks_textBox.setText(result[16])
            self.remarks_textBox.show()

            # Left Side of Vertical Box
            self.left_vbox.addRow(productionID_label, productionID_input)
            self.left_vbox.addRow(productCode_label, productCode_input)
            self.left_vbox.addRow(customer_label, customer_input)
            self.left_vbox.addRow(orderedQuantity_label, orderedQuantity_input)
            self.left_vbox.addRow(lotnumber_label, lot_number_input)
            self.left_vbox.addRow(product_input_label, product_input)
            self.left_vbox.addRow(productOutput_label, product_output_input)
            self.left_vbox.addRow(loss_label, loss_input)
            self.left_vbox.addRow(machine_label, machine_input)
            self.left_vbox.addRow(formulaID_label, self.formulaID_input)
            self.left_vbox.addRow(order_number_lbl, order_number_input)

            # Add widgets to the right Form Box
            self.right_vbox.addRow(feedrate_label, feedRate_input)
            self.right_vbox.addRow(rpm_label, rpm_input)
            self.right_vbox.addRow(screenSize_label, screenSize_input)
            self.right_vbox.addRow(screwConf_label, screwConf_input)
            self.right_vbox.addRow(purging_label, purging_input)
            self.right_vbox.addRow(resin_label, resin_input)
            self.right_vbox.addRow(purgeStart_label, purgeStart_input)
            self.right_vbox.addRow(purgeEnd_label, purgeEnd_input)
            self.right_vbox.addRow(operator_label, operator_input)
            self.right_vbox.addRow(supervisor_label, supervisor_input)

            # Time Table Entry
            time_table = QtWidgets.QTableWidget(self.entry_widget)
            time_table.setGeometry(0, 500, 450, 200)
            time_table.setColumnCount(3)
            time_table.setRowCount(8)
            time_table.setColumnWidth(0, 150)
            time_table.setColumnWidth(1, 150)
            time_table.setStyleSheet("background-color: white;")
            time_table.setFont(QtGui.QFont("Arial", 10))
            time_table.setHorizontalHeaderLabels(["Time Start", "Time End", "Output"])

            # Populate the Table
            for i in range(len(result[9])):
                item = QTableWidgetItem(str(result[9][i]))
                item2 = QTableWidgetItem(str(result[10][i]))
                item3 = QTableWidgetItem(str(result[-5][i]))
                time_table.setItem(i, 0 , item)
                time_table.setItem(i, 1, item2)
                time_table.setItem(i, 2, item3)

            time_table.show()

            # Temperature Table Entry
            temperature_table = QtWidgets.QTableWidget(self.entry_widget)
            temperature_table.setGeometry(450, 500, 150, 200)
            temperature_table.setColumnCount(1)
            temperature_table.setRowCount(12)
            temperature_table.setStyleSheet("background-color: white;")
            temperature_table.setHorizontalHeaderLabels(["Temperature"])
            temperature_index = ["Z" + str(i + 1) for i in range(12)]  # set the index
            temperature_table.setVerticalHeaderLabels(temperature_index)

            # Populate the table
            for i in range(len(result[-7])):
                item = QTableWidgetItem(str(result[-7][i]))
                temperature_table.setItem(i, 0, item)

            temperature_table.show()


            save_btn = QtWidgets.QPushButton(self.entry_widget)
            save_btn.setGeometry(540, 705, 60, 25)
            save_btn.clicked.connect(update)
            save_btn.setText("Save")
            save_btn.setCursor(Qt.PointingHandCursor)
            save_btn.show()

            default_date = QtCore.QDateTime(2024, 1, 1, 0, 0)

            time_start_input = QtWidgets.QDateTimeEdit(self.entry_widget)
            time_start_input.setGeometry(30, 475, 120, 25)
            time_start_input.setDisplayFormat("MM-dd-yyyy HH:mm")
            time_start_input.setDateTime(default_date)
            time_start_input.show()

            time_end_input = QtWidgets.QDateTimeEdit(self.entry_widget)
            time_end_input.setGeometry(180, 475, 120, 25)
            time_end_input.setDisplayFormat("MM-dd-yyyy HH:mm")
            time_end_input.setDateTime(default_date)
            time_end_input.show()

            output_lineEdit = QtWidgets.QLineEdit(self.entry_widget)
            output_lineEdit.setGeometry(340, 475, 80, 25)
            output_lineEdit.setAlignment(Qt.AlignCenter)
            output_lineEdit.setStyleSheet("background-color: white; border: 1px solid black")
            output_lineEdit.show()

            self.plus_icon = ClickableLabel(self.entry_widget)
            self.plus_icon.setGeometry(420, 475, 25, 25)
            self.plus_icon.setPixmap(QtGui.QIcon('plus.png').pixmap(25, 25))
            self.plus_icon.setCursor(Qt.PointingHandCursor)
            self.plus_icon.clicked.connect(add_time)
            self.plus_icon.show()

        def print_file():

            from openpyxl.styles import Font
            from openpyxl.styles import Alignment

            selected = self.extruder_table.selectedItems()

            process_id = selected[0].text()

            self.cursor.execute(f"""
            SELECT * FROM extruder 
            WHERE process_id = '{process_id}';
            
            """)
            items = self.cursor.fetchall()[0]
            # Unpack the Items
            machine_number = items[1]
            quantity_order = items[2]
            customer = items[4]
            code = items[6]
            total_input = items[-2]
            total_time = items[8]
            time_start = items[9]
            time_end = items[10]
            outputPerHour = items[27]
            total_output = items[3]
            outputPercent = items[11]
            loss = items[12]
            lossPercent = items[13]
            purging = items[14]
            resin = items[15]
            remarks = items[16]
            operator = items[21]
            supervisor = items[22]
            outputs = items[-5]
            materials = items[-8]
            lot_number = items[-1][0]

            wb = load_workbook(r"C:\Users\Administrator\PycharmProjects\3.12\MBPI system Reprogram\Extruder Template.xlsx")
            worksheet = wb.active

            font = Font(size=8, bold=True, name='Arial')
            center_Alignment = Alignment(horizontal='center', vertical='center')

            worksheet["F5"] = "Extruder Machine No. " + machine_number[-1]
            worksheet["A8"] = machine_number[-1]
            worksheet["B8"] = quantity_order  # quantity order
            worksheet["C8"].font = font
            worksheet["C8"].alignment = center_Alignment
            worksheet["C8"] = customer  # customer

            worksheet["F8"] = code  # product code
            worksheet["G9"] = total_input  # total input
            worksheet["H9"] = total_time  # total time used
            worksheet["I9"] = outputPerHour  # output Per Hour
            worksheet["K9"] = total_output  # total Output
            worksheet["L9"] = outputPercent  # Total Output Percentage
            worksheet["M9"] = loss
            worksheet["N9"] = lossPercent

            total_sec = timedelta()
            for row in range(len(time_start)):
                worksheet["A" + str(12 + row)] = time_start[row].strftime("%b-%d-%Y %H:%M")
                worksheet["D" + str(12 + row)] = time_end[row].strftime("%b-%d-%Y %H:%M")
                worksheet["F" + str(12 + row)] = time_end[row] - time_start[row]
                worksheet["G" + str(12 + row)] = outputs[row]
                total_sec = total_sec + (time_end[row] - time_start[row])

            try:
                hour = str(int(total_sec.total_seconds() // 3600))
                minute = str((int(total_sec.total_seconds() % 3600) // 60))
                print(hour, minute)
                total_time_used = time(int(hour), int(minute))

                worksheet["F25"] = total_time_used
            except ValueError:
                worksheet["F25"] = hour + ":" + minute

            for key in list(materials.keys()):
                worksheet["I" + str(12 + list(materials.keys()).index(key))] = key
                worksheet["K" + str(12 + list(materials.keys()).index(key))] = materials[key]

            for ln in range(len(lot_number)):
                worksheet["M" + str(12 + ln)] = lot_number[ln]

            worksheet["B27"] = purging
            worksheet["B29"] = resin
            worksheet["G26"] = remarks
            worksheet["M28"] = operator
            worksheet["M29"] = supervisor

            wb.save("text.xlsx")
            print("load successful")

        def filter_table():

            queryConList = []

            if self.machine_combo.currentText() != "-":
                queryConList.append(f"machine = '{self.machine_combo.currentText()}'")
            if self.company_combo.currentText() != "-":
                queryConList.append(f"customer = '{self.company_combo.currentText().replace("'","''")}'")


            query = f"""
                    SELECT 
                    process_id, machine, customer, qty_order, total_output, formula_id, product_code, total_time
                    FROM extruder WHERE
                                    """

            if len(queryConList) == 0:
                query = query.replace("WHERE", "")

            for condition in queryConList:
                if condition != queryConList[-1]:
                    query += condition + " AND "
                else:
                    query += condition

            query += "ORDER BY process_id"

            self.cursor.execute(query)
            result = self.cursor.fetchall()
            if self.extruder_table.rowCount() < len(result):
                self.extruder_table.setRowCount(len(result))


            self.extruder_table.clearContents()
            for i in range(len(result)):
                for j in range(len(column_names)):
                    item = QtWidgets.QTableWidgetItem(str(result[i][j]))  # Convert to string
                    # Set Alignment for specific columns
                    if j == 2 or j == 6 or j == 3 or j == 4 or j == 7:
                        item.setTextAlignment(Qt.AlignCenter)
                    else:
                        pass
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)  # Make the cells unable to be edited
                    self.extruder_table.setItem(i, j, item)

            self.extruder_table.show()
            pass

        self.extruder_table = QtWidgets.QTableWidget(self.main_widget)
        self.extruder_table.setGeometry(QtCore.QRect(20, 50, 900, 375))
        self.extruder_table.verticalHeader().setVisible(False)
        self.extruder_table.setSortingEnabled(True)

        self.cursor.execute("""
        SELECT column_name FROM information_schema.columns
        WHERE TABLE_NAME = 'extruder';
        """)
        # column_names = self.cursor.fetchall()
        # column_names = [i[0] for i in column_names]

        column_names = ["process_id", "machine", "customer", "qty_order", "total_output", "formula_id", "product_code",
                        "total time(hr)"]

        try:
            self.cursor.execute("""SELECT 
                        process_id, machine, customer, qty_order, total_output, formula_id, product_code, total_time
                        FROM extruder
                        ORDER BY process_id DESC;
                        """)
            result = self.cursor.fetchall()
        except Exception as e:
            self.cursor.execute("""
                        SELECT 
                        process_id, machine, customer, qty_order, total_output, formula_id, product_code, total_time
                        FROM extruder
                        ORDER BY process_id DESC
                        ; 

                        """)
            result = self.cursor.fetchall()

        # Set Column Count
        self.extruder_table.setColumnCount(len(column_names))
        # Set Row Count
        self.extruder_table.setRowCount(len(result))


        self.extruder_table.setStyleSheet("""
        gridline-color: rgb(0, 0, 127); 
        color : rgb(0, 121, 0);
        """)

        # Populate table with data
        for i in range(len(result)):
            for j in range(len(column_names)):
                item = QtWidgets.QTableWidgetItem(str(result[i][j]))  # Convert to string
                # Set Alignment for specific columns
                if j == 2 or j == 6 or j == 3 or j == 4 or j == 7:
                    item.setTextAlignment(Qt.AlignCenter)
                else:
                    pass
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)  # Make the cells unable to be edited
                self.extruder_table.setItem(i, j, item)

        bold_font = QtGui.QFont()
        bold_font.setBold(True)
        self.extruder_table.horizontalHeader().setFont(bold_font)
        self.extruder_table.horizontalHeader().setStyleSheet("""
        QHeaderView::section{
        font-weight: bold;
        background-color: black;
        color: white;
        }

        """)

        # Set Column Width
        self.extruder_table.setColumnWidth(2, 175)

        self.extruder_table.setHorizontalHeaderLabels([col.upper() for col in column_names])  # Set column names
        # Set selection mode to select entire rows and disable single item selection
        self.extruder_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.extruder_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.extruder_table.show()

        #Filters
        self.machine_combo = QtWidgets.QComboBox(self.main_widget)
        self.machine_combo.setGeometry(120, 10, 100, 30)
        self.cursor.execute("""
                    SELECT DISTINCT(machine) FROM extruder;
                """)
        machine = self.cursor.fetchall()
        self.machine_combo.addItem("-")
        for i in machine:
            self.machine_combo.addItem(i[0])
        self.machine_combo.currentIndexChanged.connect(filter_table)
        self.machine_combo.show()

        self.company_combo = QtWidgets.QComboBox(self.main_widget)
        self.company_combo.setGeometry(220, 10, 170, 30)
        self.cursor.execute("""
            SELECT DISTINCT(customer) FROM extruder;
        """)
        customers = self.cursor.fetchall()
        self.company_combo.addItem("-")
        for i in customers:
            self.company_combo.addItem(i[0])
        self.company_combo.setEditable(True)
        self.company_combo.currentIndexChanged.connect(filter_table)
        self.company_combo.show()

        self.view_btn = QtWidgets.QPushButton(self.main_widget)
        self.view_btn.setGeometry(100, 500, 100, 30)
        self.view_btn.setText("View")
        self.view_btn.setStyleSheet("background-color : red;")
        self.view_btn.clicked.connect(show_form)
        self.view_btn.show()

        self.add_btn = QtWidgets.QPushButton(self.main_widget)
        self.add_btn.setGeometry(250, 500, 100, 30)
        self.add_btn.setText("Add Entry")
        self.add_btn.setStyleSheet("background-color : red;")
        self.add_btn.clicked.connect(add_entry)
        self.add_btn.show()

        self.update_btn = QtWidgets.QPushButton(self.main_widget)
        self.update_btn.setGeometry(400, 500, 100, 30)
        self.update_btn.setText("Update")
        self.update_btn.setStyleSheet("background-color : red;")
        self.update_btn.clicked.connect(update_entry)
        self.update_btn.show()

        self.print_btn = QtWidgets.QPushButton(self.main_widget)
        self.print_btn.setGeometry(550, 500, 100, 30)
        self.print_btn.setText("Print")
        self.print_btn.setStyleSheet("background-color: red;")
        self.print_btn.clicked.connect(print_file)
        self.print_btn.show()

if __name__ == "__main__":
    import sys

    app = QtWidgets.QApplication(sys.argv)
    LoginWindow = QtWidgets.QMainWindow()
    ui = Ui_LoginWindow()
    ui.setupUi(LoginWindow)
    LoginWindow.show()
    sys.exit(app.exec_())